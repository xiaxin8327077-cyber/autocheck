from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, NamedTuple, Sequence

from openpyxl import load_workbook


LABEL_TO_PROCESS = {
    "人行模板\\逐笔": "pbc_template",
    "1104": "jr_1104",
    "全要素": "full_elements",
    "中信登定期": "citic_registration",
    "EAST5.0": "east5",
    "五篇大文章": "five_articles",
}

QUARTERLY_MONTHS = {1, 4, 7, 10}


class ScheduleRow(NamedTuple):
    report_month: str
    process_code: str
    report_date: date
    source_type: str
    source_year: int


PROCESSES = (
    ("pbc_central", "人行大集中", 10, 1, 1),
    ("pbc_template", "人行模板、逐笔报送", 20, 1, 1),
    ("jr_1104", "1104报送", 30, 1, 1),
    ("full_elements", "21、23版全要素报送", 40, 1, 1),
    ("citic_registration", "中信登定期报送", 50, 1, 1),
    ("east5", "East5报送", 60, 1, 1),
    ("five_articles", "五篇大文章报送", 70, 1, 1),
    ("supplement_tasks", "补录任务统计", 1000, 0, 0),
)

STEPS = (
    ("pbc_central_1", "pbc_central", "导入并生成存续回购业务明细信息", 1, "all_rows_match_report_date", 1, 0, 1),
    ("pbc_central_2", "pbc_central", "资产负债表的资产合计与负债及权益合计一致", 2, "amounts_equal", 1, 0, 1),
    ("pbc_central_3", "pbc_central", "募集信息表中的地区与客户类型不为空，且份额满足跨期校验", 3, "no_blank_fields_and_no_ck", 1, 0, 1),
    ("pbc_central_4", "pbc_central", "内部产品资金端客户与资产端交易对手校验一致", 4, "no_ck_and_min_time", 1, 0, 1),
    ("pbc_template_1", "pbc_template", "确认新增债券发行人、交易对手行业规模信息", 1, "no_pending_status", 1, 0, 1),
    ("pbc_template_2", "pbc_template", "确认各维度数据是否有变化需要报备", 2, "month_rows_or_dependency", 1, 0, 1),
    ("pbc_template_3", "pbc_template", "补录底表正常更新、补录任务正常触发", 3, "default_completed", 1, 1, 1),
    ("pbc_template_4", "pbc_template", "导入补全人行全量 SPV 码", 4, "minimum_time_in_current_month", 1, 0, 1),
    ("pbc_template_5", "pbc_template", "完成监管报送平台人行报表校验", 5, "dependency_completed", 1, 0, 1),
    ("pbc_template_6", "pbc_template", "制表人核验归档报表并上传至人行报送网站", 6, "all_versions_present", 1, 0, 1),
    ("pbc_template_7", "pbc_template", "制表人填写数据调整情况说明（如有）", 7, "date_reached", 1, 0, 1),
    ("jr_1104_1", "jr_1104", "导入报送外部数据", 1, "quarterly_rows_exist", 1, 0, 1),
    ("jr_1104_2", "jr_1104", "核查跨表资产大类科目余额与实收信托比对", 2, "dependency_completed", 1, 0, 1),
    ("jr_1104_3", "jr_1104", "完成监管报送平台 1104 报表校验", 3, "dependency_completed", 1, 0, 1),
    ("jr_1104_4", "jr_1104", "制表人核验归档报表并上传至金监报送网站", 4, "version_present", 1, 0, 1),
    ("full_elements_1", "full_elements", "归档并上传全要素报送", 1, "version_present", 1, 0, 1),
    ("citic_registration_1", "citic_registration", "补录底表正常更新、补录任务正常触发", 1, "default_completed", 1, 1, 1),
    ("citic_registration_2", "citic_registration", "导入报送外部数据", 2, "current_month_rows_in_all_sources", 1, 0, 1),
    ("citic_registration_3", "citic_registration", "确认各表不存在数据缺失、数据重复", 3, "dependency_completed", 1, 0, 1),
    ("citic_registration_4", "citic_registration", "完成监管报送平台中信登报表校验", 4, "dependency_completed", 1, 0, 1),
    ("citic_registration_5", "citic_registration", "制表人核验归档报表并上传中信登平台质检网站", 5, "version_present", 1, 0, 1),
    ("east5_1", "east5", "归档并上传 East5 报送", 1, "version_present", 1, 0, 1),
    ("five_articles_1", "five_articles", "归档并上传五篇大文章报送", 1, "version_present", 1, 0, 1),
    ("supplement_tasks_1", "supplement_tasks", "统计补录任务", 1, "supplement_task_counts", 0, 0, 0),
)

DEPENDENCIES = (
    ("pbc_template_2", "pbc_template_6"),
    ("pbc_template_5", "pbc_template_6"),
    ("jr_1104_2", "jr_1104_4"),
    ("jr_1104_3", "jr_1104_4"),
    ("citic_registration_3", "citic_registration_5"),
    ("citic_registration_4", "citic_registration_5"),
)

SOURCES = (
    ("pbc_central_1", "primary", "ass_man_reg", "ex_pledge_back", 1, 1),
    ("pbc_central_2", "primary", "reg-report-analysis", "zf_detail_2024", 1, 1),
    ("pbc_central_3", "duration", "currency_report_24", "currency_report_duration", 1, 1),
    ("pbc_central_3", "ck_result", "reg-report-analysis", "ck_result", 2, 1),
    ("pbc_central_4", "ck_result", "reg-report-analysis", "ck_result", 1, 1),
    ("pbc_central_4", "spv_detail", "ass_man_reg_24", "zgxgzh_spvdetail_zg08", 2, 1),
    ("pbc_template_1", "straight_flush", "currency_report_24", "straight_flush", 1, 1),
    ("pbc_template_1", "straight_flush_yxzgq", "currency_report_24", "straight_flush_yxzgq", 2, 1),
    ("pbc_template_2", "primary", "ass_man_reg", "product_change", 1, 1),
    ("pbc_template_4", "primary", "ass_man_reg_24", "zg08_wb", 1, 1),
    ("pbc_template_6", "primary", "reg-report-analysis", "xt_reg_version", 1, 1),
    ("jr_1104_1", "primary", "1104", "relation_ship_1104_dm", 1, 1),
    ("jr_1104_4", "primary", "reg-report-analysis", "xt_reg_version", 1, 1),
    ("full_elements_1", "primary", "reg-report-analysis", "xt_reg_version", 1, 1),
    ("citic_registration_2", "asset_credit", "zxd", "zxd_asset_credit_info", 1, 1),
    ("citic_registration_2", "external_data", "zxd", "result14_xtbzjj_external_data", 2, 1),
    ("citic_registration_2", "basic_info", "zxd", "jsxt_basic_info", 3, 1),
    ("citic_registration_5", "primary", "reg-report-analysis", "xt_reg_version", 1, 1),
    ("east5_1", "primary", "reg-report-analysis", "xt_reg_version", 1, 1),
    ("five_articles_1", "primary", "reg-report-analysis", "xt_reg_version", 1, 1),
    ("supplement_tasks_1", "primary", "bl", "jsxt_console.rep_data_task_detail", 1, 1),
)

SOURCE_FIELDS = (
    ("pbc_central_1", "primary", "period_field", "reporting_period"),
    ("pbc_central_2", "primary", "period_field", "caldate"),
    ("pbc_central_2", "primary", "left_amount_field", "a0001"),
    ("pbc_central_2", "primary", "right_amount_field", "d0000"),
    ("pbc_central_3", "duration", "period_field", "caldate"),
    ("pbc_central_3", "duration", "region_field", "c_regioncode"),
    ("pbc_central_3", "duration", "customer_type_field", "c_custtype"),
    ("pbc_central_3", "ck_result", "period_field", "period"),
    ("pbc_central_3", "ck_result", "check_id_field", "ck_id"),
    ("pbc_central_4", "ck_result", "period_field", "period"),
    ("pbc_central_4", "ck_result", "check_id_field", "ck_id"),
    ("pbc_central_4", "spv_detail", "time_field", "tbtime"),
    ("pbc_template_1", "straight_flush", "period_field", "caldate"),
    ("pbc_template_1", "straight_flush", "status_field", "status"),
    ("pbc_template_1", "straight_flush_yxzgq", "period_field", "caldate"),
    ("pbc_template_1", "straight_flush_yxzgq", "status_field", "status"),
    ("pbc_template_2", "primary", "date_field", "chdate"),
    ("pbc_template_4", "primary", "time_field", "tbtime"),
    ("pbc_template_6", "primary", "manage_code_field", "manage_code"),
    ("pbc_template_6", "primary", "version_field", "version_num"),
    ("jr_1104_1", "primary", "date_field", "createdate"),
    ("jr_1104_4", "primary", "manage_code_field", "manage_code"),
    ("jr_1104_4", "primary", "version_field", "version_num"),
    ("full_elements_1", "primary", "manage_code_field", "manage_code"),
    ("full_elements_1", "primary", "version_field", "version_num"),
    ("citic_registration_2", "asset_credit", "date_field", "createdate"),
    ("citic_registration_2", "external_data", "date_field", "createdate"),
    ("citic_registration_2", "basic_info", "date_field", "createdate"),
    ("citic_registration_5", "primary", "manage_code_field", "manage_code"),
    ("citic_registration_5", "primary", "version_field", "version_num"),
    ("east5_1", "primary", "manage_code_field", "manage_code"),
    ("east5_1", "primary", "version_field", "version_num"),
    ("five_articles_1", "primary", "manage_code_field", "manage_code"),
    ("five_articles_1", "primary", "version_field", "version_num"),
    ("supplement_tasks_1", "primary", "id_field", "id"),
    ("supplement_tasks_1", "primary", "date_field", "create_date"),
    ("supplement_tasks_1", "primary", "status_field", "status"),
    ("supplement_tasks_1", "primary", "deleted_field", "del_flag"),
)

VALUES = (
    ("pbc_central_3", "target_ck_id", "5677", "text", 1),
    ("pbc_central_4", "target_ck_id", "7118", "text", 1),
    ("pbc_template_1", "pending_status", "待确认", "text", 1),
    ("pbc_template_1", "pending_status", "待补充", "text", 2),
    ("pbc_template_6", "manage_code", "20002", "text", 1),
    ("pbc_template_6", "manage_code", "zbbs24", "text", 2),
    ("jr_1104_1", "active_month", "1", "integer", 1),
    ("jr_1104_1", "active_month", "4", "integer", 2),
    ("jr_1104_1", "active_month", "7", "integer", 3),
    ("jr_1104_1", "active_month", "10", "integer", 4),
    ("jr_1104_4", "manage_code", "system1104", "text", 1),
    ("full_elements_1", "manage_code", "qysnew", "text", 1),
    ("citic_registration_5", "manage_code", "zxdreport", "text", 1),
    ("east5_1", "manage_code", "east5", "text", 1),
    ("five_articles_1", "manage_code", "dwz5", "text", 1),
    ("supplement_tasks_1", "completed_status", "5", "text", 1),
    ("supplement_tasks_1", "valid_deleted_value", "0", "text", 1),
)


def _coerce_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
    return None


def load_schedule_rows(path: Path | str) -> list[ScheduleRow]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if "工作日数据" not in workbook.sheetnames:
        raise ValueError("工作簿缺少“工作日数据”工作表")
    grouped: dict[tuple[str, int, int], date] = {}
    sheet = workbook["工作日数据"]
    try:
        for values in sheet.iter_rows(values_only=True):
            label = next(
                (
                    str(value).strip()
                    for value in values
                    if value is not None and str(value).strip() in LABEL_TO_PROCESS
                ),
                "",
            )
            if not label:
                continue
            report_date = next((parsed for value in values if (parsed := _coerce_date(value)) is not None), None)
            if report_date is None:
                continue
            process_code = LABEL_TO_PROCESS[label]
            if process_code == "five_articles" and report_date.month not in QUARTERLY_MONTHS:
                continue
            key = (process_code, report_date.year, report_date.month)
            grouped[key] = max(grouped.get(key, report_date), report_date)
    finally:
        workbook.close()

    years = sorted({year for _, year, _ in grouped})
    for year in years:
        for month in range(1, 13):
            grouped[("pbc_central", year, month)] = date(year, month, 1)

    return [
        ScheduleRow(
            report_month=f"{year:04d}-{month:02d}",
            process_code=process_code,
            report_date=report_date,
            source_type="default" if process_code == "pbc_central" else "imported",
            source_year=year,
        )
        for (process_code, year, month), report_date in sorted(
            grouped.items(), key=lambda item: (item[0][1], item[0][2], item[0][0])
        )
    ]


def _sql_text(value: object) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _row(values: Iterable[object]) -> str:
    rendered = []
    for value in values:
        if value is None:
            rendered.append("NULL")
        elif isinstance(value, (int, float)):
            rendered.append(str(value))
        elif isinstance(value, date):
            rendered.append(_sql_text(value.isoformat()))
        else:
            rendered.append(_sql_text(value))
    return "(" + ", ".join(rendered) + ")"


def _upsert(table: str, columns: Sequence[str], rows: Sequence[Sequence[object]], update_columns: Sequence[str]) -> str:
    column_sql = ", ".join(f"`{column}`" for column in columns)
    values_sql = ",\n  ".join(_row(row) for row in rows)
    update_sql = ", ".join(f"`{column}`=VALUES(`{column}`)" for column in update_columns)
    return (
        f"INSERT INTO `{table}` ({column_sql}) VALUES\n  {values_sql}\n"
        f"ON DUPLICATE KEY UPDATE {update_sql};"
    )


def build_seed_sql(schedule_rows: Sequence[ScheduleRow]) -> str:
    sections = [
        _upsert(
            "report_nav_processes",
            ("process_code", "process_name", "display_order", "enabled", "allow_manual_step_completion"),
            PROCESSES,
            ("process_name", "display_order", "enabled", "allow_manual_step_completion"),
        ),
        _upsert(
            "report_nav_process_months",
            ("process_code", "month_no"),
            tuple((code, month) for code, *_ in PROCESSES if code != "supplement_tasks" for month in (QUARTERLY_MONTHS if code == "five_articles" else range(1, 13))),
            ("month_no",),
        ),
        _upsert(
            "report_nav_steps",
            ("step_code", "process_code", "step_name", "display_order", "evaluator_key", "enabled", "default_completed", "manual_completion_allowed"),
            STEPS,
            ("process_code", "step_name", "display_order", "evaluator_key", "enabled", "default_completed", "manual_completion_allowed"),
        ),
        _upsert(
            "report_nav_step_dependencies",
            ("step_code", "depends_on_step_code"),
            DEPENDENCIES,
            ("depends_on_step_code",),
        ),
    ]

    source_rows = tuple((index, *source) for index, source in enumerate(SOURCES, start=1))
    sections.append(
        _upsert(
            "report_nav_step_sources",
            ("id", "step_code", "source_role", "data_source_name", "table_name", "display_order", "enabled"),
            source_rows,
            ("step_code", "source_role", "data_source_name", "table_name", "display_order", "enabled"),
        )
    )
    source_ids = {(step_code, source_role): index for index, (step_code, source_role, *_rest) in enumerate(SOURCES, start=1)}
    field_rows = tuple(
        (index, source_ids[(step_code, source_role)], field_role, column_name)
        for index, (step_code, source_role, field_role, column_name) in enumerate(SOURCE_FIELDS, start=1)
    )
    sections.append(
        _upsert(
            "report_nav_step_fields",
            ("id", "step_source_id", "field_role", "column_name"),
            field_rows,
            ("step_source_id", "field_role", "column_name"),
        )
    )
    value_rows = tuple((index, *value) for index, value in enumerate(VALUES, start=1))
    sections.append(
        _upsert(
            "report_nav_step_values",
            ("id", "step_code", "value_role", "value_text", "value_type", "display_order"),
            value_rows,
            ("step_code", "value_role", "value_text", "value_type", "display_order"),
        )
    )
    if schedule_rows:
        sections.append(
            _upsert(
                "report_nav_monthly_schedules",
                ("report_month", "process_code", "report_date", "source_type", "source_year", "updated_by", "updated_at"),
                tuple((*row, "seed", "2026-07-16 00:00:00") for row in schedule_rows),
                ("report_date", "source_type", "source_year", "updated_by", "updated_at"),
            )
        )
    sections.append(
        _upsert(
            "report_nav_scheduler_state",
            ("id", "enabled", "interval_minutes", "next_run_at", "lock_owner", "lock_until", "last_started_at", "last_finished_at", "last_status", "last_error", "updated_at"),
            ((1, 1, 10, None, None, None, None, None, None, None, "2026-07-16 00:00:00"),),
            ("enabled", "interval_minutes", "updated_at"),
        )
    )
    return "\n\n".join(sections) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description="Build report navigation relational seed SQL")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    rows = load_schedule_rows(args.workbook)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(build_seed_sql(rows), encoding="utf-8")
    print(f"wrote {args.output.as_posix()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
