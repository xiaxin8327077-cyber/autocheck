from io import BytesIO

from openpyxl import load_workbook

from auto_check.db_validation.legacy_rules import ACTIVE_LEGACY_RULES, DISABLED_LEGACY_RULE_IDS
from auto_check.db_validation.rules_document import (
    DOCUMENT_ENABLED_RULE_IDS,
    RULE_DOCUMENT_FILENAME,
    build_rules_document,
)


def test_rules_document_is_user_readable_workbook():
    filename, payload = build_rules_document()

    assert filename == RULE_DOCUMENT_FILENAME
    assert payload.startswith(b"PK")

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    assert len(workbook.sheetnames) == 3

    overview = workbook[workbook.sheetnames[0]]
    overview_values = [cell.value for row in overview.iter_rows(values_only=False) for cell in row if cell.value]
    assert overview_values
    assert any(
        "\u516c\u5f00\u4fe1\u606f\u4ea4\u53c9\u6821\u9a8c" in str(value)
        and "\u6a21\u677f\u4ea4\u53c9\u6821\u9a8c" in str(value)
        for value in overview_values
    )

    detail = workbook[workbook.sheetnames[2]]
    rows = list(detail.iter_rows(values_only=True))
    assert rows[0][3] in {"\u89c4\u5219\u7f16\u53f7", "\u7470\u52cb\u5782\u7f16\u53f7"}
    flat_text = "\n".join(str(value) for row in rows for value in row if value)
    expected_rule_ids = {rule.rule_id for rule in ACTIVE_LEGACY_RULES}
    document_rule_ids = {row[3] for row in rows[1:]}
    assert expected_rule_ids <= document_rule_ids
    disabled_rule_ids = DISABLED_LEGACY_RULE_IDS - DOCUMENT_ENABLED_RULE_IDS
    disabled_rows = [row for row in rows[1:] if row[3] in disabled_rule_ids]
    if disabled_rows:
        assert {row[9] for row in disabled_rows} <= {"\u505c\u7528", "\u934b\u6ec5\u7528"}
    assert "SQL" not in flat_text

    workbook.close()


def test_rules_document_describes_template_cpkj_table_mapping():
    _, payload = build_rules_document()

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    rows = list(workbook[workbook.sheetnames[2]].iter_rows(values_only=True))
    detail_by_rule_id = {row[3]: row for row in rows[1:]}
    flat_text = "\n".join(str(value) for row in rows for value in row if value)

    assert detail_by_rule_id["Zg09_Rule3"][9] == "启用"
    assert detail_by_rule_id["Zg10_Rule1"][9] == "启用"
    assert "信托产品类型口径=1 对比字段映射解析出的 ZG09 口径 1 模板物理表" in flat_text
    assert "信托产品类型口径=1 对比字段映射解析出的 ZG10 口径 1 模板物理表" in flat_text
    assert "口径=2 对比口径 2 模板物理表" in flat_text

    workbook.close()


def test_rules_document_describes_zg05_zg07_loan_balance_mapping_dependency():
    _, payload = build_rules_document()

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    rows = list(workbook[workbook.sheetnames[2]].iter_rows(values_only=True))
    detail_by_rule_id = {row[3]: row for row in rows[1:]}

    assert "ZG07“贷款余额折人民币”字段映射" in str(detail_by_rule_id["Zg05_Rule3"][10])

    workbook.close()


def test_rules_document_describes_zg02_original_amount_mapping_dependency():
    _, payload = build_rules_document()

    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    rows = list(workbook[workbook.sheetnames[2]].iter_rows(values_only=True))
    detail_by_rule_id = {row[3]: row for row in rows[1:]}

    assert "“初始募集金额”和“初始募集金额折人民币”字段映射" in str(
        detail_by_rule_id["Zg02_Rule1"][10]
    )

    workbook.close()
