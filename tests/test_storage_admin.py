from io import BytesIO

from openpyxl import load_workbook

from auto_check.app.local_store import _connect, db_path_for_config
from auto_check.app.server import ApiRouter
from auto_check.app.storage_schema import CURRENT_SCHEMA_VERSION


ADMIN_USER = {"id": "u-admin", "username": "admin", "display_name": "管理员", "role": "admin"}
NORMAL_USER = {"id": "u-user", "username": "user", "display_name": "普通用户", "role": "user"}


def _seed_storage_rows(config_path):
    with _connect(db_path_for_config(config_path)) as connection:
        connection.execute(
            """
            INSERT INTO data_sources(
                id, name, db_type, host, port, database_name, schema_name,
                username, password_encrypted, is_default, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "ds-main",
                "DWS 生产只读",
                "mysql",
                "10.20.18.9",
                3306,
                "risk_dw",
                "public",
                "risk_reader",
                "aesgcm$very-secret",
                1,
                "2026-07-03 10:00:00",
                "2026-07-03 10:01:00",
            ),
        )
        connection.execute(
            """
            INSERT INTO app_settings(key, value_json, updated_at)
            VALUES (?, ?, ?)
            """,
            (
                "flow_tool",
                '{"service_url":"https://internal.example","token":"secret-token","nested":{"username":"operator"}}',
                "2026-07-03 10:02:00",
            ),
        )
        connection.execute(
            """
            INSERT INTO users(
                id, username, display_name, role, password_hash, enabled,
                created_at, updated_at, last_login_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "u-admin",
                "admin",
                "管理员",
                "admin",
                "pbkdf2_sha256$260000$v1Sz$raw-secret-hash",
                1,
                "2026-07-03 09:00:00",
                "2026-07-03 09:01:00",
                "2026-07-03 09:02:00",
            ),
        )


def test_admin_storage_router_requires_login_and_admin_role(tmp_path):
    router = ApiRouter(config_path=tmp_path / "config.json")

    assert router.handle("GET", "/api/admin/storage/tables", None, current_user=None)[0] == 401
    assert router.handle("GET", "/api/admin/storage/tables", None, current_user=NORMAL_USER)[0] == 403

    status, payload = router.handle("GET", "/api/admin/storage/tables", None, current_user=ADMIN_USER)

    assert status == 200
    names = {table["name"] for table in payload["tables"]}
    assert {"data_sources", "run_headers", "reconcile_results", "storage_migration_runs"} <= names
    assert any(table["cn_name"] == "数据源配置表" for table in payload["tables"])


def test_admin_storage_health_reports_schema_integrity_and_counts(tmp_path):
    config_path = tmp_path / "config.json"
    router = ApiRouter(config_path=config_path)
    _seed_storage_rows(config_path)

    status, payload = router.handle("GET", "/api/admin/storage/health", None, current_user=ADMIN_USER)

    assert status == 200
    health = payload["health"]
    assert health["schema_version"] == CURRENT_SCHEMA_VERSION
    assert health["integrity_check"] == "ok"
    assert health["foreign_key_issues"] == 0
    assert health["database"]["filename"] == "auto-check.db"
    assert "auto-check.db" in health["database"]["display_path"]
    assert str(tmp_path) not in health["database"]["display_path"]
    assert health["table_counts"]["data_sources"] == 1


def test_admin_storage_rows_are_paginated_whitelisted_and_mask_sensitive_values(tmp_path):
    config_path = tmp_path / "config.json"
    router = ApiRouter(config_path=config_path)
    _seed_storage_rows(config_path)

    status, payload = router.handle(
        "GET",
        "/api/admin/storage/tables/data_sources/rows",
        None,
        current_user=ADMIN_USER,
    )

    assert status == 200
    assert payload["page"] == 1
    assert payload["page_size"] == 20
    row = payload["rows"][0]
    assert row["name"] == "DWS 生产只读"
    assert row["host"] == "10.20.18.***"
    assert row["username"] == "ri***"
    assert row["password_encrypted"] == "******"
    assert payload["fields"]["password_encrypted"]["sensitive"] is True
    assert payload["fields"]["password_encrypted"]["display"] == "脱敏"

    router._query_string = "page=1&page_size=10"
    status, payload = router.handle(
        "GET",
        "/api/admin/storage/tables/app_settings/rows",
        None,
        current_user=ADMIN_USER,
    )
    assert status == 200
    value_json = payload["rows"][0]["value_json"]
    assert isinstance(value_json, dict)
    assert value_json["token"] == "******"
    assert value_json["nested"]["username"] == "op***"

    status, payload = router.handle(
        "GET",
        "/api/admin/storage/tables/users/rows",
        None,
        current_user=ADMIN_USER,
    )
    assert status == 200
    user_row = payload["rows"][0]
    assert user_row["username"] == "ad***"
    assert user_row["password_hash"] == "******"


def test_admin_storage_rejects_unlisted_tables_and_invalid_page_size(tmp_path):
    router = ApiRouter(config_path=tmp_path / "config.json")

    status, payload = router.handle(
        "GET",
        "/api/admin/storage/tables/sqlite_master/rows",
        None,
        current_user=ADMIN_USER,
    )
    assert status == 404
    assert "not found" in payload["error"]

    router._query_string = "page=1&page_size=999"
    status, payload = router.handle(
        "GET",
        "/api/admin/storage/tables/data_sources/rows",
        None,
        current_user=ADMIN_USER,
    )
    assert status == 400
    assert "page_size" in payload["error"]


def test_admin_storage_schema_export_is_valid_xlsx_with_catalog_and_table_sheets(tmp_path):
    config_path = tmp_path / "config.json"
    router = ApiRouter(config_path=config_path)
    _seed_storage_rows(config_path)

    filename, payload = router.get_storage_schema_export(current_user=ADMIN_USER)

    assert filename == "本地数据库表结构.xlsx"
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    assert "表目录" in workbook.sheetnames
    assert "data_sources" in workbook.sheetnames
    catalog = workbook["表目录"]
    header = [cell.value for cell in next(catalog.iter_rows(min_row=1, max_row=1))]
    assert header[:7] == ["分类", "表名", "中文名", "记录数", "字段数", "主键", "用途"]
    data_sources = workbook["data_sources"]
    rows = list(data_sources.iter_rows(values_only=True))
    assert rows[0][:6] == ("字段序号", "字段名", "类型", "约束", "中文说明", "展示策略")
    assert any(row[1] == "password_encrypted" and row[5] == "脱敏" for row in rows[1:])
    workbook.close()


def test_admin_storage_table_data_export_uses_cn_headers_and_masked_values(tmp_path):
    config_path = tmp_path / "config.json"
    router = ApiRouter(config_path=config_path)
    _seed_storage_rows(config_path)

    filename, payload = router.get_storage_table_data_export("data_sources", current_user=ADMIN_USER)

    assert filename == "data_sources-表数据.xlsx"
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    assert "data_sources" in workbook.sheetnames
    sheet = workbook["data_sources"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][:9] == ("数据源 ID", "数据源名称", "数据库类型", "主机地址", "端口", "数据库名", "Schema", "用户名", "加密密码")
    assert rows[1][:9] == ("ds-main", "DWS 生产只读", "mysql", "10.20.18.***", 3306, "risk_dw", "public", "ri***", "******")
    workbook.close()

    filename, payload = router.get_storage_table_data_export("app_settings", current_user=ADMIN_USER)
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    rows = list(workbook["app_settings"].iter_rows(values_only=True))
    assert rows[0][:2] == ("设置键", "设置内容")
    assert '"token": "******"' in rows[1][1]
    assert '"username": "op***"' in rows[1][1]
    workbook.close()
