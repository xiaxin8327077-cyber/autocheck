from datetime import date

from openpyxl import load_workbook

from auto_check.db_validation.excel import result_filename, write_result_excel
from auto_check.db_validation.models import ValidationResultRow


def test_excel_writer_matches_old_result_structure(tmp_path):
    path = tmp_path / "result.xlsx"
    rows = [
        ValidationResultRow(
            data_date="2026-05-31",
            org_code="D1003632000013",
            org_name="江苏省国际信托有限责任公司",
            manager_org="南京",
            detail="产品代码:P1",
            form="资管产品基本信息校验",
            value1="产品名称:X",
            value2="",
            mark="20260531-D1003632000013-ZG01-Zg01_Rule6",
            rule="Zg01_Rule6:产品名称长度小于等于5个字，有特殊符号，需核实",
            error="产品名称过于简单，含有特殊字符（？、！、^），需核实",
            note="",
        )
    ]

    write_result_excel(path, rows)
    wb = load_workbook(path)
    ws = wb.active

    assert wb.sheetnames == ["Sheet1"]
    assert [ws.cell(1, c).value for c in range(1, 13)] == [
        "数据日期",
        "金融机构编码",
        "法人金融机构名称",
        "数据管理机构",
        "明细数据相关信息",
        "校验表单",
        "数据值1",
        "数据值2",
        "校验标识",
        "校验规则",
        "错误描述",
        "情况说明",
    ]
    assert ws.cell(2, 9).value == "20260531-D1003632000013-ZG01-Zg01_Rule6"
    assert ws.freeze_panes is None
    assert ws.auto_filter.ref is None


def test_result_filename_uses_old_program_format():
    assert result_filename(date(2026, 5, 31)) == (
        "20260531-资管产品数据审核结果-模板校验（否）-公开信息校验（否）(Ver.20260202).xlsx"
    )
