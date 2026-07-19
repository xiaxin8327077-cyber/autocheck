from decimal import Decimal
from io import BytesIO
from datetime import date
from pathlib import Path
import errno
import socket
import sqlite3
import sys
import threading
import time
import zipfile

from openpyxl import load_workbook
import pytest

import auto_check.app.server as server_module
from auto_check.app.config import (
    ConfigStore,
    DataSourceEntry,
    DataSourceConfig,
    DbValidationDatasetSettings,
    DbValidationSettings,
    FlowChainConfig,
    FlowChainStep,
    FlowToolSettings,
    NamedConfig,
    PbcImportToolSettings,
    ReconcileDataSourceSettings,
    load_store,
    save_store,
)
from auto_check.app.flow_tool import FlowChainRunResult, FlowChainStepResult, FlowDefinition
from auto_check.app.pbc_import import TableColumn
from auto_check.app.repositories import DEFAULT_RECONCILE_TABLES
from auto_check.app.server import (
    ApiRouter,
    AutoCheckRequestHandler,
    RunJob,
    _connection_error_message,
    _runtime_error_message,
    build_display_details,
    previous_month_end,
)
from auto_check.app.reconcile_schema import ReconcileSchemaSettings, ReconcileSourceRef, ReconcileTableSchema
from auto_check.db_validation.metadata import TableFieldCatalog
from auto_check.db_validation.models import DbValidationRunResult, ValidationResultRow
from auto_check.engine.models import DifferenceDetail, ReconcileResult, ValuationMatch, ValuationRow
from auto_check.engine.reconcile import NoSourceReportData
from mysql_config_test_support import MemoryApplicationDatabase


def db_path_for_config(config_path):
    return Path(config_path).with_name("auto-check.db")


@pytest.fixture(autouse=True)
def shared_application_database(monkeypatch):
    """Keep legacy router tests on one MySQL-contract store per test."""
    database = MemoryApplicationDatabase()
    original_router_init = ApiRouter.__init__
    original_load_store = load_store
    original_save_store = save_store

    def router_init(self, *args, **kwargs):
        kwargs.setdefault("application_database", database)
        original_router_init(self, *args, **kwargs)

    def test_load_store(path=None, *, database_override=None, database=None):
        selected = database or database_override or shared_database
        return original_load_store(path, database=selected)

    def test_save_store(store, path=None, *, database_override=None, database=None):
        selected = database or database_override or shared_database
        return original_save_store(store, path, database=selected)

    shared_database = database
    monkeypatch.setattr(ApiRouter, "__init__", router_init)
    monkeypatch.setattr(sys.modules[__name__], "load_store", test_load_store)
    monkeypatch.setattr(sys.modules[__name__], "save_store", test_save_store)
    monkeypatch.setattr(
        server_module.ApplicationDatabase,
        "from_config_path",
        staticmethod(lambda _path: database),
    )
    return database


class FakeRunner:
    def run(self, date):
        assert date == "2026-04-30"
        return [
            ReconcileResult(
                project_code="P1",
                project_name="Project",
                asset_total=Decimal("100"),
                liability_equity_total=Decimal("90"),
                received_trust_balance=Decimal("50"),
                difference=Decimal("10"),
                direction="资产大于负债及权益",
                difference_reason="实收本金差异",
                match_status="已解释",
                details=[
                    DifferenceDetail(
                        kind="received_trust",
                        data={
                            "c1000_balance": "50",
                            "fa_4001_balance": "500",
                            "received_trust_difference": "450",
                        },
                    )
                ],
            )
        ]


class DisconnectingWriter:
    def write(self, _data):
        raise ConnectionAbortedError(10053, "client disconnected")


class CountingHistoryStore:
    def __init__(self, count=0):
        self.count = count
        self.list_called = False
        self.count_called = False

    def list_runs(self):
        self.list_called = True
        raise AssertionError("system info should not load full history payloads")

    def get_run(self, run_id):
        return None

    def save_run(self, run):
        pass

    def delete_run(self, run_id):
        return False

    def count_runs(self):
        self.count_called = True
        return self.count


def test_send_json_ignores_client_disconnect():
    handler = object.__new__(AutoCheckRequestHandler)
    handler.wfile = DisconnectingWriter()
    handler.send_response = lambda status: None
    handler.send_header = lambda name, value: None
    handler.end_headers = lambda: None

    handler._send_json(200, {"ok": True})


class FakeConnectionTester:
    def __call__(self, config):
        return {
            "dws": {"ok": True, "message": "连接成功"},
            "business": {"ok": True, "message": "连接成功"},
        }


def _complete_reconcile_schema_payload():
    return {
        "version": 1,
        "strict": True,
        "tables": {
            logical_key: {
                "source_ref": {
                    "id": table_schema.source_ref.id or "dws",
                    "name": table_schema.source_ref.name or table_schema.source_ref.id or "dws",
                    "match_by": table_schema.source_ref.match_by or "id_then_name",
                },
                "table": table_schema.table,
                "display_name": table_schema.display_name,
                "fields": dict(table_schema.fields),
                **({"optional_fields": dict(table_schema.optional_fields)} if table_schema.optional_fields else {}),
            }
            for logical_key, table_schema in DEFAULT_RECONCILE_TABLES.items()
        },
    }


def _all_reconcile_schema_columns(data_source, table):
    column_names = {
        physical_name
        for table_schema in DEFAULT_RECONCILE_TABLES.values()
        for physical_name in [*table_schema.fields.values(), *table_schema.optional_fields.values()]
    }
    return [TableColumn(column_name, "") for column_name in sorted(column_names)]


def _reconcile_schema_payload_to_yaml(payload):
    def scalar(value):
        if isinstance(value, bool):
            return "true" if value else "false"
        return str(value)

    def emit_mapping(mapping, indent=0):
        lines = []
        prefix = " " * indent
        for key, value in mapping.items():
            if isinstance(value, dict):
                lines.append(f"{prefix}{key}:")
                lines.extend(emit_mapping(value, indent + 2))
            else:
                lines.append(f"{prefix}{key}: {scalar(value)}")
        return lines

    return "\n".join(emit_mapping(payload))


class SlowRunner:
    def run(self, date):
        time.sleep(0.3)
        return []


class NoSourceRunner:
    def run(self, date):
        raise NoSourceReportData(date)


class FailingRunner:
    def __init__(self, message: str = "unsafe table identifier: zf_detail_2024ǮǮǮ"):
        self.message = message

    def run(self, date):
        raise ValueError(self.message)


class FakePbcImporter:
    def __init__(self):
        self.calls = []

    def __call__(self, **kwargs):
        self.calls.append(kwargs)
        kwargs["log"]("wrote 2 rows", 90, "write database")
        return 2


class SlowPbcImporter:
    def __init__(self):
        self.calls = []
        self.started = threading.Event()
        self.release = threading.Event()

    def __call__(self, **kwargs):
        self.calls.append(kwargs)
        self.started.set()
        self.release.wait(timeout=5)
        return 1


class FakeFlowChainExecutor:
    def __init__(self):
        self.calls = []

    def __call__(self, **kwargs):
        self.calls.append(kwargs)
        chain = kwargs["chain"]
        context = kwargs["context"]
        kwargs["log"]("流程A执行结束", 100, "流程A")
        return FlowChainRunResult(
            chain_id=chain.id,
            chain_name=chain.name,
            trigger_type=context.trigger_type,
            status="completed",
            steps=[
                FlowChainStepResult(
                    flow_id=step.flow_id,
                    flow_name=step.name,
                    status="completed",
                    sp_task_id=658149,
                    begin_time="2026-06-11 16:35:10",
                    end_time="2026-06-11 16:36:10",
                    message="执行结束",
                )
                for step in chain.steps
            ],
        )


class FakeFailingFlowChainExecutor:
    def __init__(self, error_message="流程B超时：等待流程执行结束超时"):
        self.calls = []
        self.error_message = error_message

    def __call__(self, **kwargs):
        self.calls.append(kwargs)
        kwargs["log"]("流程A执行结束", 50, "流程A")
        kwargs["log"](f"流程B执行失败: {self.error_message}", 50, "流程B")
        raise RuntimeError(self.error_message)


class BlockingFlowChainExecutor:
    def __init__(self):
        self.calls = []

    def __call__(self, **kwargs):
        self.calls.append(kwargs)
        kwargs["log"]("流程A等待结束", 10, "流程A")
        cancel_event = kwargs["cancel_event"]
        for _ in range(100):
            if cancel_event.is_set():
                raise RuntimeError("流程执行已取消")
            time.sleep(0.01)
        raise RuntimeError("测试流程未收到取消请求")


class FakeDbValidationExecutor:
    def __init__(self):
        self.calls = []

    def __call__(self, **kwargs):
        self.calls.append(kwargs)
        kwargs["log"]("checked ZG01", 88, "ZG01")
        output_path = kwargs["output_dir"] / "result.xlsx"
        output_path.write_bytes(b"xlsx")
        row = ValidationResultRow(
            data_date=kwargs["report_date"].isoformat(),
            org_code="D1003632000013",
            org_name="江苏省国际信托有限责任公司",
            manager_org="南京",
            detail="产品代码_产品名称:P1_A?",
            form="资管产品基本信息校验",
            value1="产品名称:A?",
            value2="",
            mark="20260531-D1003632000013-ZG01-Zg01_Rule6",
            rule="Zg01_Rule6:产品名称长度小于等于5个字，有特殊符号，需核实",
            error="产品名称过于简单，需核实",
        )
        return DbValidationRunResult(
            report_date=kwargs["report_date"].isoformat(),
            error_count=1,
            excel_path=output_path,
            rows=[row],
            warnings=["ZG02 当期表无数据"],
        )


class FakeDbValidationFieldMappingLoader:
    def __init__(self):
        self.calls = []

    def __call__(self, **kwargs):
        self.calls.append(kwargs)
        return TableFieldCatalog(
            {
                "zgxgzh_baseinfo_zg01_26": {
                    "产品代码": "projcode",
                    "产品名称": "projname",
                }
            },
            unmapped_field_count=1,
        )


def fake_table_columns(data_source, table):
    assert table.parts == ("dws", "public_information_th")
    return [
        TableColumn("info_type_name", "Info Type Name"),
        TableColumn("product_code", "Product Code"),
        TableColumn("issuer_name", "Issuer Name"),
    ]


def fake_any_pbc_table_columns(data_source, table):
    return [
        TableColumn("product_code", "Product Code"),
        TableColumn("product_name", "Product Name"),
        TableColumn("drop_me", "Drop Me"),
    ]


def _zip_bytes(files: dict[str, str]) -> bytes:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content.encode("utf-8-sig"))
    return buffer.getvalue()


def test_get_config_returns_defaults(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle("GET", "/api/config", None)

    assert status == 200
    assert payload["dws"]["db_type"] == "postgresql"
    assert payload["business"]["db_type"] == "mysql"
    assert payload["default_run_date"] == previous_month_end()


def test_api_router_passes_one_shared_application_database_to_config_handlers(
    monkeypatch, tmp_path
):
    application_database = object()
    calls = []
    store = ConfigStore()

    def fake_load_store(path, *, database):
        calls.append(("load", Path(path), database))
        return store

    def fake_save_store(saved_store, path, *, database):
        calls.append(("save", Path(path), database))
        assert saved_store is store

    monkeypatch.setattr(server_module, "load_store", fake_load_store)
    monkeypatch.setattr(server_module, "save_store", fake_save_store)
    config_path = tmp_path / "config.json"

    router = ApiRouter(
        config_path=config_path,
        application_database=application_database,
        history_path=tmp_path / "history.json",
    )
    assert router.handle("GET", "/api/settings/defaults", None)[0] == 200
    assert router.handle("POST", "/api/settings/defaults", {"page_size": 20})[0] == 200

    assert calls == [
        ("load", config_path, application_database),
        ("load", config_path, application_database),
        ("save", config_path, application_database),
    ]


def test_run_server_builds_validates_and_closes_application_database_before_services(
    monkeypatch, tmp_path
):
    events = []
    config_path = tmp_path / "config.json"

    class FakeApplicationDatabase:
        def test_connection(self):
            events.append("test_connection")

        def validate_schema(self):
            events.append("validate_schema")

        def close(self):
            events.append("close")

    application_database = FakeApplicationDatabase()

    def from_config_path(path):
        events.append(("from_config_path", Path(path)))
        return application_database

    class FakeServer:
        server_address = ("127.0.0.1", 8765)

        def serve_forever(self):
            events.append("serve_forever")

    def build_http_server(_address, _handler):
        events.append("http_server")
        return FakeServer()

    report_navigation_service = object()

    class FakeRouter:
        def __init__(self, *, config_path, application_database, **_kwargs):
            events.append(("router", application_database))
            self.config_path = Path(config_path)
            self.report_navigation = report_navigation_service

    class FakeReportNavigationScheduler:
        def __init__(self, service):
            events.append(("scheduler", service))

        def start(self):
            events.append("scheduler_start")

        def stop(self):
            events.append("scheduler_stop")

    class FakeAuthManager:
        def __init__(self, path, *, database):
            events.append(("auth", Path(path), database))

    monkeypatch.setattr(server_module, "ThreadingHTTPServer", build_http_server)
    monkeypatch.setattr(server_module, "ApiRouter", FakeRouter)
    monkeypatch.setattr(server_module, "AuthManager", FakeAuthManager)
    monkeypatch.setattr(server_module, "ReportNavigationScheduler", FakeReportNavigationScheduler)
    monkeypatch.setattr(server_module, "_is_tcp_port_active", lambda _host, _port: False)
    monkeypatch.setattr(
        server_module.ApplicationDatabase,
        "from_config_path",
        staticmethod(from_config_path),
    )

    server = server_module.run_server(
        host="127.0.0.1",
        port=8765,
        open_browser=False,
        config_path=config_path,
    )

    assert server is not None
    assert events == [
        ("from_config_path", config_path),
        "test_connection",
        "validate_schema",
        "http_server",
        ("router", application_database),
        ("auth", config_path, application_database),
        ("scheduler", report_navigation_service),
        "scheduler_start",
        "serve_forever",
        "scheduler_stop",
        "close",
    ]


def test_previous_month_end_uses_beijing_today(monkeypatch):
    monkeypatch.setattr("auto_check.app.server.beijing_today", lambda: date(2026, 1, 1))

    assert previous_month_end() == "2025-12-31"


def test_run_job_uses_beijing_timestamp_and_log_time(monkeypatch):
    monkeypatch.setattr("auto_check.app.server.beijing_timestamp", lambda: "2026-06-04T08:15:20")
    monkeypatch.setattr("auto_check.app.server.beijing_time_text", lambda: "08:15:20")
    job = RunJob(date="2026-04-30", max_combination_rows=50, current_user={"username": "admin"})

    job.start()
    job.log("读取数据", 20, "读取数据")

    payload = job.to_payload()
    assert payload["started_at"] == "2026-06-04T08:15:20"
    assert payload["logs"][0]["time"] == "08:15:20"


def test_run_server_opens_existing_url_when_port_is_busy(monkeypatch, tmp_path):
    opened_urls = []

    def raise_port_busy(_address, _handler):
        raise OSError(errno.EADDRINUSE, "Address already in use")

    monkeypatch.setattr(server_module, "ThreadingHTTPServer", raise_port_busy)
    monkeypatch.setattr(server_module.webbrowser, "open", lambda url: opened_urls.append(url))

    server = server_module.run_server(
        host="127.0.0.1",
        port=8765,
        open_browser=True,
        config_path=tmp_path / "config.json",
    )

    assert server is None
    assert opened_urls == ["http://127.0.0.1:8765"]
    assert not db_path_for_config(tmp_path / "config.json").exists()


def test_run_server_probes_port_before_creating_local_services(monkeypatch, tmp_path):
    opened_urls = []
    listener = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    listener.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    listener.bind(("127.0.0.1", 0))
    listener.listen(1)
    port = listener.getsockname()[1]

    def fail_if_server_created(_address, _handler):
        raise AssertionError("run_server should not create a second server on an active port")

    monkeypatch.setattr(server_module, "ThreadingHTTPServer", fail_if_server_created)
    monkeypatch.setattr(server_module.webbrowser, "open", lambda url: opened_urls.append(url))

    try:
        server = server_module.run_server(
            host="127.0.0.1",
            port=port,
            open_browser=True,
            config_path=tmp_path / "config.json",
        )
    finally:
        listener.close()

    assert server is None
    assert opened_urls == [f"http://127.0.0.1:{port}"]
    assert not db_path_for_config(tmp_path / "config.json").exists()


def test_post_config_saves_payload(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )
    status, payload = router.handle(
        "POST",
        "/api/config",
        {
            "dws": {
                "db_type": "postgresql",
                "host": "localhost",
                "port": 5432,
                "database": "dwsdb",
                "schema": "dws",
                "username": "u",
                "password": "",
            },
            "business": {
                "db_type": "mysql",
                "host": "localhost",
                "port": 3306,
                "database": "bizdb",
                "schema": "",
                "username": "u2",
                "password": "",
            },
        },
    )

    assert status == 200
    assert payload["ok"] is True

    _, loaded = router.handle("GET", "/api/config", None)
    assert loaded["business"]["database"] == "bizdb"


def test_default_settings_api_persists_across_router_instances(tmp_path):
    config_path = tmp_path / "config.json"
    router = ApiRouter(
        config_path=config_path,
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle(
        "POST",
        "/api/settings/defaults",
        {
            "session_expire_hours": 12,
            "page_size": 20,
            "combination_limit": 50,
            "auto_refresh_home": True,
            "visual_effects": False,
            "theme": "space-tech",
            "dark_mode": True,
        },
    )

    assert status == 200
    assert payload["settings"]["session_expire_hours"] == 12

    restarted_router = ApiRouter(
        config_path=config_path,
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )
    status, loaded = restarted_router.handle("GET", "/api/settings/defaults", None)

    assert status == 200
    assert loaded["settings"] == {
        "session_expire_hours": 12,
        "page_size": 20,
        "combination_limit": 50,
        "auto_refresh_home": True,
        "visual_effects": False,
        "theme": "space-tech",
        "dark_mode": True,
    }


def test_interface_settings_api_defaults_upserts_and_isolates_current_users(
    tmp_path, shared_application_database
):
    router = ApiRouter(config_path=tmp_path / "config.json")
    first_user = {"id": "user-1", "role": "user"}
    second_user = {"id": "user-2", "role": "admin"}

    assert router.handle(
        "GET", "/api/settings/interface", None, current_user=first_user
    ) == (
        200,
        {
            "settings": {
                "radius_px": 4,
                "line_chart_style": "straight",
            }
        },
    )

    assert router.handle(
        "POST",
        "/api/settings/interface",
        {
            "radius_px": 7,
            "line_chart_style": "smooth",
            "vitality_theme_color": "#FFFFFF",
            "calm_theme_color": "#000000",
            "user_id": "user-2",
        },
        current_user=first_user,
    ) == (
        200,
        {
            "settings": {
                "radius_px": 7,
                "line_chart_style": "smooth",
            }
        },
    )
    assert router.handle(
        "POST",
        "/api/settings/interface",
        {"radius_px": 9, "line_chart_style": "straight"},
        current_user=first_user,
    ) == (
        200,
        {
            "settings": {
                "radius_px": 9,
                "line_chart_style": "straight",
            }
        },
    )

    assert router.handle(
        "GET", "/api/settings/interface", None, current_user=first_user
    ) == (
        200,
        {
            "settings": {
                "radius_px": 9,
                "line_chart_style": "straight",
            }
        },
    )
    assert router.handle(
        "GET", "/api/settings/interface", None, current_user=second_user
    ) == (
        200,
        {
            "settings": {
                "radius_px": 4,
                "line_chart_style": "straight",
            }
        },
    )
    assert [
        (row["user_id"], row["radius_px"])
        for row in shared_application_database.connection.tables["user_interface_preferences"]
    ] == [("user-1", 9)]
    assert shared_application_database.connection.tables["user_interface_preferences"][0][
        "vitality_theme_color"
    ] is None
    assert shared_application_database.connection.tables["user_interface_preferences"][0][
        "calm_theme_color"
    ] is None


@pytest.mark.parametrize(("radius_px", "line_chart_style"), [(1, "straight"), (15, "smooth")])
def test_interface_settings_api_accepts_boundary_values(
    tmp_path, radius_px, line_chart_style
):
    router = ApiRouter(config_path=tmp_path / "config.json")
    current_user = {"id": "user-1", "role": "user"}

    assert router.handle(
        "POST",
        "/api/settings/interface",
        {
            "radius_px": radius_px,
            "line_chart_style": line_chart_style,
        },
        current_user=current_user,
    ) == (
        200,
        {
            "settings": {
                "radius_px": radius_px,
                "line_chart_style": line_chart_style,
            }
        },
    )
    assert router.handle(
        "GET", "/api/settings/interface", None, current_user=current_user
    ) == (
        200,
        {
            "settings": {
                "radius_px": radius_px,
                "line_chart_style": line_chart_style,
            }
        },
    )


@pytest.mark.parametrize(
    ("body", "error"),
    [
        pytest.param({"radius_px": 0, "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="below-minimum"),
        pytest.param({"radius_px": 16, "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="above-maximum"),
        pytest.param({"radius_px": 1.5, "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="float"),
        pytest.param({"radius_px": True, "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="radius-true"),
        pytest.param({"radius_px": False, "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="radius-false"),
        pytest.param({"radius_px": None, "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="radius-null"),
        pytest.param({"line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="radius-missing"),
        pytest.param({"radius_px": "4", "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="numeric-string"),
        pytest.param({"radius_px": "bad", "line_chart_style": "straight"}, "radius_px must be an integer between 1 and 15", id="radius-string"),
        pytest.param({"radius_px": 4, "line_chart_style": "curve"}, "line_chart_style must be one of: smooth, straight", id="line-style-curve"),
        pytest.param({"radius_px": 4, "line_chart_style": "STRAIGHT"}, "line_chart_style must be one of: smooth, straight", id="line-style-uppercase"),
        pytest.param({"radius_px": 4, "line_chart_style": 1}, "line_chart_style must be one of: smooth, straight", id="line-style-integer"),
        pytest.param({"radius_px": 4, "line_chart_style": None}, "line_chart_style must be one of: smooth, straight", id="line-style-null"),
        pytest.param({"radius_px": 4}, "line_chart_style must be one of: smooth, straight", id="line-style-missing"),
        pytest.param([1], "radius_px must be an integer between 1 and 15", id="body-list"),
        pytest.param("bad", "radius_px must be an integer between 1 and 15", id="body-string"),
        pytest.param(1, "radius_px must be an integer between 1 and 15", id="body-integer"),
        pytest.param(True, "radius_px must be an integer between 1 and 15", id="body-boolean"),
    ],
)
def test_interface_settings_api_rejects_invalid_values_without_writing(
    tmp_path, shared_application_database, body, error
):
    router = ApiRouter(config_path=tmp_path / "config.json")

    status, payload = router.handle(
        "POST",
        "/api/settings/interface",
        body,
        current_user={"id": "user-1", "role": "user"},
    )

    assert status == 400
    assert payload == {"error": error}
    assert shared_application_database.connection.tables["user_interface_preferences"] == []


@pytest.mark.parametrize("current_user", [None, {}, {"id": ""}])
def test_interface_settings_api_requires_current_user_id(tmp_path, current_user):
    router = ApiRouter(config_path=tmp_path / "config.json")

    assert router.handle(
        "GET", "/api/settings/interface", None, current_user=current_user
    ) == (401, {"error": "login required"})


def test_theme_colors_api_allows_anonymous_defaults(tmp_path):
    router = ApiRouter(config_path=tmp_path / "config.json")

    assert router.handle(
        "GET", "/api/settings/interface/theme-colors", None
    ) == (
        200,
        {
            "colors": {
                "system": {"vitality": "#3466D9", "calm": "#355F63"},
                "personal": {"vitality": None, "calm": None},
                "effective": {"vitality": "#3466D9", "calm": "#355F63"},
            },
            "capabilities": {"can_manage_system_theme_colors": False},
        },
    )


def test_theme_colors_api_resolves_personal_overrides_per_field_and_user(
    tmp_path, shared_application_database
):
    router = ApiRouter(config_path=tmp_path / "config.json")
    shared_application_database.connection.tables["system_interface_preferences"].append(
        {
            "id": 1,
            "vitality_theme_color": "#123456",
            "calm_theme_color": "#654321",
            "updated_by": "admin-1",
        }
    )
    shared_application_database.connection.tables["user_interface_preferences"].extend(
        [
            {
                "user_id": "user-1",
                "radius_px": 4,
                "line_chart_style": "straight",
                "vitality_theme_color": "#ABCDEF",
                "calm_theme_color": None,
            },
            {
                "user_id": "user-2",
                "radius_px": 4,
                "line_chart_style": "straight",
                "vitality_theme_color": None,
                "calm_theme_color": "#FEDCBA",
            },
        ]
    )

    status, first = router.handle(
        "GET",
        "/api/settings/interface/theme-colors",
        None,
        current_user={"id": "user-1", "role": "user"},
    )
    assert status == 200
    assert first == {
        "colors": {
            "system": {"vitality": "#123456", "calm": "#654321"},
            "personal": {"vitality": "#ABCDEF", "calm": None},
            "effective": {"vitality": "#ABCDEF", "calm": "#654321"},
        },
        "capabilities": {"can_manage_system_theme_colors": False},
    }

    status, second = router.handle(
        "GET",
        "/api/settings/interface/theme-colors",
        None,
        current_user={"id": "user-2", "role": "admin"},
    )
    assert status == 200
    assert second["colors"]["personal"] == {"vitality": None, "calm": "#FEDCBA"}
    assert second["colors"]["effective"] == {
        "vitality": "#123456",
        "calm": "#FEDCBA",
    }
    assert second["capabilities"] == {"can_manage_system_theme_colors": True}


def test_theme_colors_api_admin_save_normalizes_both_colors_and_records_updater(
    tmp_path, shared_application_database
):
    router = ApiRouter(config_path=tmp_path / "config.json")
    admin = {"id": "admin-1", "role": "admin"}

    status, payload = router.handle(
        "POST",
        "/api/settings/interface/theme-colors",
        {"vitality_theme_color": "#abcdef", "calm_theme_color": "#102030"},
        current_user=admin,
    )

    assert status == 200
    assert payload == {
        "colors": {
            "system": {"vitality": "#ABCDEF", "calm": "#102030"},
            "personal": {"vitality": None, "calm": None},
            "effective": {"vitality": "#ABCDEF", "calm": "#102030"},
        },
        "capabilities": {"can_manage_system_theme_colors": True},
    }
    assert shared_application_database.connection.tables["system_interface_preferences"] == [
        {
            "id": 1,
            "vitality_theme_color": "#ABCDEF",
            "calm_theme_color": "#102030",
            "updated_by": "admin-1",
            "updated_at": shared_application_database.connection.tables[
                "system_interface_preferences"
            ][0]["updated_at"],
        }
    ]


@pytest.mark.parametrize(
    ("body", "error"),
    [
        (
            {"vitality_theme_color": "bad", "calm_theme_color": "#102030"},
            "vitality_theme_color must be a #RRGGBB string",
        ),
        (
            {"vitality_theme_color": "#ABCDEF", "calm_theme_color": "#12345G"},
            "calm_theme_color must be a #RRGGBB string",
        ),
        (
            {"vitality_theme_color": "#ABCDEF"},
            "calm_theme_color must be a #RRGGBB string",
        ),
    ],
)
def test_theme_colors_api_rejects_invalid_pair_atomically(
    tmp_path, shared_application_database, body, error
):
    router = ApiRouter(config_path=tmp_path / "config.json")
    shared_application_database.connection.tables["system_interface_preferences"].append(
        {
            "id": 1,
            "vitality_theme_color": "#3466D9",
            "calm_theme_color": "#355F63",
            "updated_by": "admin-old",
        }
    )

    status, payload = router.handle(
        "POST",
        "/api/settings/interface/theme-colors",
        body,
        current_user={"id": "admin-1", "role": "admin"},
    )

    assert status == 400
    assert payload == {"error": error}
    assert shared_application_database.connection.tables["system_interface_preferences"] == [
        {
            "id": 1,
            "vitality_theme_color": "#3466D9",
            "calm_theme_color": "#355F63",
            "updated_by": "admin-old",
        }
    ]


@pytest.mark.parametrize(
    ("current_user", "expected"),
    [
        (None, (401, {"error": "login required"})),
        ({"id": "user-1", "role": "user"}, (403, {"error": "admin role required"})),
    ],
)
def test_theme_colors_api_write_requires_admin(tmp_path, current_user, expected):
    router = ApiRouter(config_path=tmp_path / "config.json")

    assert router.handle(
        "POST",
        "/api/settings/interface/theme-colors",
        {"vitality_theme_color": "#ABCDEF", "calm_theme_color": "#102030"},
        current_user=current_user,
    ) == expected


def test_reconcile_schema_settings_api_saves_and_initializes_from_yaml_template(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
                DataSourceEntry(
                    id="yaml-dws",
                    name="yaml",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "yaml", "u", "p"),
                ),
            ],
        ),
        config_path,
    )
    def schema_columns(data_source, table):
        return _all_reconcile_schema_columns(data_source, table) + [
            TableColumn("sys_proj", "项目代码"),
            TableColumn("sys_date", "估值日期"),
            TableColumn("sys_account", "科目代码"),
            TableColumn("sys_name", "科目名称"),
            TableColumn("sys_amt", "市值"),
            TableColumn("yaml_proj", "项目代码"),
            TableColumn("yaml_date", "估值日期"),
            TableColumn("yaml_account", "科目代码"),
            TableColumn("yaml_name", "科目名称"),
            TableColumn("yaml_amt", "市值"),
            TableColumn("changed_amt", "市值"),
        ]
    router = ApiRouter(config_path=config_path, pbc_table_column_loader=schema_columns)
    schema_payload = _complete_reconcile_schema_payload()
    for table_config in schema_payload["tables"].values():
        table_config["source_ref"] = {"id": "source-dws", "name": "DWS", "match_by": "id_then_name"}
    schema_payload["tables"]["fa_valuation"] = {
        "source_ref": {"id": "source-dws", "name": "DWS", "match_by": "id_then_name"},
        "table": "system.fa_valuation",
        "display_name": "FA valuation",
        "fields": {
            "project_code": "sys_proj",
            "valuation_date": "sys_date",
            "account_code": "sys_account",
            "account_name": "sys_name",
            "market_value": "sys_amt",
        },
    }

    status, payload = router.handle("POST", "/api/settings/reconcile-schema", schema_payload)

    assert status == 200
    assert payload["schema"]["strict"] is True
    assert payload["schema"]["tables"]["fa_valuation"]["table"] == "system.fa_valuation"
    loaded = load_store(config_path)
    assert loaded.reconcile_schema.tables["fa_valuation"].fields["market_value"] == "sys_amt"

    status, payload = router.handle("GET", "/api/settings/reconcile-schema", None)

    assert status == 200
    assert payload["schema"]["tables"]["fa_valuation"]["fields"]["market_value"] == "sys_amt"

    yaml_schema_payload = _complete_reconcile_schema_payload()
    for table_config in yaml_schema_payload["tables"].values():
        table_config["source_ref"] = {"id": "yaml-dws", "name": "yaml", "match_by": "id_then_name"}
    yaml_schema_payload["tables"]["fa_valuation"] = {
        "source_ref": {"id": "yaml-dws", "name": "yaml", "match_by": "id_then_name"},
        "table": "yaml.fa_valuation",
        "display_name": "FA valuation YAML",
        "fields": {
            "project_code": "yaml_proj",
            "valuation_date": "yaml_date",
            "account_code": "yaml_account",
            "account_name": "yaml_name",
            "market_value": "yaml_amt",
        },
    }
    config_path.with_name("reconcile-schema.yaml").write_text(
        _reconcile_schema_payload_to_yaml({"reconcile_schema": yaml_schema_payload}),
        encoding="utf-8",
    )

    status, payload = router.handle("POST", "/api/settings/reconcile-schema/init-from-file", {})

    assert status == 200
    assert payload["schema"]["strict"] is True
    assert payload["schema"]["tables"]["fa_valuation"]["table"] == "yaml.fa_valuation"
    assert payload["schema"]["tables"]["fa_valuation"]["fields"]["market_value"] == "yaml_amt"
    loaded = load_store(config_path)
    assert loaded.reconcile_schema.tables["fa_valuation"].table == "yaml.fa_valuation"

    config_path.with_name("reconcile-schema.yaml").write_text(
        """
reconcile_schema:
  version: 1
  tables:
    fa_valuation:
      table: changed.fa_valuation
      fields:
        market_value: changed_amt
""".strip(),
        encoding="utf-8",
    )

    status, payload = router.handle("GET", "/api/settings/reconcile-schema", None)

    assert status == 200
    assert payload["schema"]["tables"]["fa_valuation"]["table"] == "yaml.fa_valuation"
    assert payload["schema"]["tables"]["fa_valuation"]["fields"]["market_value"] == "yaml_amt"


def test_reconcile_schema_save_validates_table_and_field_existence(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )

    def schema_columns(data_source, table):
        assert table.parts == ("system", "fa_valuation")
        return [
            TableColumn("sys_proj", "项目代码"),
            TableColumn("sys_date", "估值日期"),
            TableColumn("sys_account", "科目代码"),
            TableColumn("sys_name", "科目名称"),
        ]

    router = ApiRouter(config_path=config_path, pbc_table_column_loader=schema_columns)

    status, payload = router.handle(
        "POST",
        "/api/settings/reconcile-schema",
        {
            "version": 1,
            "strict": True,
            "tables": {
                "fa_valuation": {
                    "source_ref": {"id": "source-dws", "name": "DWS", "match_by": "id_then_name"},
                    "table": "system.fa_valuation",
                    "display_name": "FA 估值表",
                    "fields": {
                        "project_code": "sys_proj",
                        "valuation_date": "sys_date",
                        "account_code": "sys_account",
                        "account_name": "sys_name",
                        "market_value": "sys_amt_missing",
                    },
                }
            },
        },
    )

    assert status == 400
    assert "表字段配置校验失败" in payload["error"]
    assert "FA 估值表" in payload["error"]
    assert "sys_amt_missing" in payload["error"]


def test_reconcile_schema_save_rejects_missing_required_business_fields(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )

    def schema_columns(data_source, table):
        return [
            TableColumn("c_projcode", "项目代码"),
            TableColumn("d_cldate", "截止日期"),
            TableColumn("c_pactid", "合同编号"),
            TableColumn("f_acbalance", "投资余额"),
            TableColumn("d_bdate", "合同开始日"),
        ]

    router = ApiRouter(config_path=config_path, pbc_table_column_loader=schema_columns)

    status, payload = router.handle(
        "POST",
        "/api/settings/reconcile-schema",
        {
            "version": 1,
            "strict": True,
            "tables": {
                "am_project_invest": {
                    "source_ref": {"id": "source-dws", "name": "DWS", "match_by": "id_then_name"},
                    "table": "am_projinvest_dws",
                    "display_name": "AM 项目投资",
                    "fields": {
                        "project_code": "c_projcode",
                        "close_date": "d_cldate",
                        "pact_id": "c_pactid",
                        "invest_balance": "f_acbalance",
                    },
                }
            },
        },
    )

    assert status == 400
    assert "表字段配置校验失败" in payload["error"]
    assert "AM 项目投资" in payload["error"]
    assert "contract_start_date" in payload["error"]
    assert "合同开始日" in payload["error"]
    assert "物理表 am_projinvest_dws" in payload["error"]


def test_reconcile_schema_save_rejects_missing_logical_tables(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="dws",
                    name="DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
                DataSourceEntry(
                    id="business",
                    name="报表库",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "reportdb", "report", "u", "p"),
                ),
            ],
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path, pbc_table_column_loader=_all_reconcile_schema_columns)
    schema_payload = _complete_reconcile_schema_payload()
    schema_payload["tables"].pop("pledge_back")

    status, payload = router.handle("POST", "/api/settings/reconcile-schema", schema_payload)

    assert status == 400
    assert "表字段配置校验失败" in payload["error"]
    assert "缺少逻辑表配置" in payload["error"]
    assert "回购质押明细表" in payload["error"]
    assert "pledge_back" in payload["error"]


def test_reconcile_schema_columns_api_returns_table_columns_for_selected_source(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path, pbc_table_column_loader=fake_table_columns)

    status, payload = router.handle(
        "POST",
        "/api/settings/reconcile-schema/columns",
        {"source_id": "source-dws", "table": "dws.public_information_th"},
    )

    assert status == 200
    assert payload["columns"] == [
        {"name": "info_type_name", "comment": "Info Type Name"},
        {"name": "product_code", "comment": "Product Code"},
        {"name": "issuer_name", "comment": "Issuer Name"},
    ]

    missing_status, missing_payload = router.handle(
        "POST",
        "/api/settings/reconcile-schema/columns",
        {"source_id": "missing-source", "table": "dws.public_information_th"},
    )

    assert missing_status == 400
    assert "数据源" in missing_payload["error"]


def test_api_router_uses_mysql_history_store_by_default(tmp_path, shared_application_database):
    config_path = tmp_path / "config.json"
    router = ApiRouter(config_path=config_path)

    router.history_store.save_run(
        {
            "id": "run-1",
            "run_date": "2026-06-01",
            "run_at": "2026-06-01 10:00:00",
            "config_fingerprint": "abc",
            "results": [],
        }
    )

    status, payload = router.handle("GET", "/api/history", None)

    assert status == 200
    assert [entry["id"] for entry in payload["history"]] == ["run-1"]
    assert [row["id"] for row in shared_application_database.connection.tables["run_headers"]] == ["run-1"]
    assert not db_path_for_config(config_path).exists()
    assert not config_path.with_name("history.json").exists()


def test_system_info_api_uses_lightweight_history_count(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-one",
                    name="one",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "db", "dws", "u", "p"),
                    is_default=True,
                ),
                DataSourceEntry(
                    id="source-two",
                    name="two",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "db", "dm", "u", "p"),
                ),
            ]
        ),
        config_path,
    )
    history_store = CountingHistoryStore(count=37)
    router = ApiRouter(config_path=config_path, history_store=history_store)

    status, payload = router.handle("GET", "/api/system-info", None)

    assert status == 200
    assert payload["history_run_count"] == 37
    assert payload["config_count"] == 2
    assert "settings" in payload
    assert history_store.count_called is True
    assert history_store.list_called is False


def test_pbc_import_settings_api_returns_sources_and_saved_tables(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
                DataSourceEntry(
                    id="source-business",
                    name="local - 报表库",
                    config=DataSourceConfig("mysql", "localhost", 3306, "bizdb", "", "u2", "p2"),
                ),
            ],
            pbc_import_tool=PbcImportToolSettings(
                recent_tables=["dws.aainfo"],
                last_config_name="local - DWS",
                last_source="dws",
            ),
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path)

    status, payload = router.handle("GET", "/api/tools/pbc-import/settings", None)

    assert status == 200
    assert payload["settings"]["recent_tables"] == ["dws.aainfo"]
    assert payload["settings"]["last_config_name"] == "local - DWS"
    assert payload["data_sources"] == [
        {"config_name": "local - DWS", "source": "dws", "label": "local - DWS", "db_type": "postgresql", "is_default": True},
        {"config_name": "local - 报表库", "source": "business", "label": "local - 报表库", "db_type": "mysql", "is_default": False},
    ]


def test_flow_tool_settings_round_trip_through_api(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("mysql", "192.168.107.81", 3306, "reg-report-analysis", "", "u", "p"),
                    is_default=True,
                )
            ],
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path)

    status, payload = router.handle(
        "POST",
        "/api/tools/flow/settings",
        {
            "source_id": "source-report",
            "execute_url": "http://192.168.107.81/assmag/spiderFlow/spider/testRun",
            "flow_table": "sp_flow",
            "task_table": "sp_task",
            "poll_interval_seconds": 5,
            "step_timeout_minutes": 60,
            "chains": [
                {
                    "id": "chain-zgxg-1",
                    "name": "资管新规1",
                    "enabled": True,
                    "schedule_enabled": True,
                    "schedule_cron": "0 7 * * *",
                    "steps": [
                        {"flow_id": "flow-a", "name": "流程A"},
                        {"flow_id": "flow-b", "name": "流程B"},
                    ],
                }
            ],
        },
    )

    assert status == 200
    assert payload["settings"]["chains"][0]["name"] == "资管新规1"
    status, payload = router.handle("GET", "/api/tools/flow/settings", None)
    assert status == 200
    assert payload["settings"]["source_id"] == "source-report"
    assert "schedule_enabled" not in payload["settings"]["chains"][0]
    assert "schedule_cron" not in payload["settings"]["chains"][0]
    assert "schedule_time" not in payload["settings"]["chains"][0]
    assert payload["settings"]["chains"][0]["steps"][1]["flow_id"] == "flow-b"


def test_flow_tool_settings_rejects_execute_url_with_fixed_flow_id(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("postgresql", "127.0.0.1", 5432, "auto_check_test", "reg-report-analysis", "u", "p"),
                    is_default=True,
                )
            ],
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path)

    status, payload = router.handle(
        "POST",
        "/api/tools/flow/settings",
        {
            "source_id": "source-report",
            "execute_url": "http://127.0.0.1/assmag/spiderFlow/spider/testRun?id=flow-a",
            "flow_table": "sp_flow",
            "task_table": "sp_task",
        },
    )

    assert status == 400
    assert "不要包含 id 参数" in payload["error"]


def test_flow_definitions_error_mentions_source_and_flow_table(monkeypatch, tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("postgresql", "127.0.0.1", 5432, "auto_check_test", "reg-report-analysis", "u", "p"),
                    is_default=True,
                )
            ],
            flow_tool=FlowToolSettings(source_id="source-report", flow_table="sp_flow", task_table="sp_task"),
        ),
        config_path,
    )

    class BrokenFlowGateway:
        def __init__(self, *_args, **_kwargs):
            pass

        def list_flows(self, _keyword):
            raise RuntimeError('relation "sp_flow" does not exist')

    monkeypatch.setattr("auto_check.app.server.DatabaseFlowGateway", BrokenFlowGateway)
    router = ApiRouter(config_path=config_path)

    status, payload = router.handle("GET", "/api/tools/flow/definitions", None)

    assert status == 500
    assert "流程表读取失败" in payload["error"]
    assert "申报平台库" in payload["error"]
    assert "sp_flow" in payload["error"]
    assert "数据表不存在" in payload["error"]


def test_flow_definitions_response_marks_truncated_and_passes_keyword(monkeypatch, tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("postgresql", "127.0.0.1", 5432, "auto_check_test", "reg-report-analysis", "u", "p"),
                    is_default=True,
                )
            ],
            flow_tool=FlowToolSettings(source_id="source-report", flow_table="sp_flow", task_table="sp_task"),
        ),
        config_path,
    )
    seen_keywords = []

    class ManyFlowGateway:
        def __init__(self, *_args, **_kwargs):
            pass

        def list_flows(self, keyword):
            seen_keywords.append(keyword)
            return [FlowDefinition(id=f"flow-{index}", name=f"流程{index}", enabled="1") for index in range(500)]

    monkeypatch.setattr("auto_check.app.server.DatabaseFlowGateway", ManyFlowGateway)
    router = ApiRouter(config_path=config_path)
    router._query_string = "keyword=flow-999"

    status, payload = router.handle("GET", "/api/tools/flow/definitions", None)

    assert status == 200
    assert seen_keywords == ["flow-999"]
    assert payload["limit"] == 500
    assert payload["truncated"] is True
    assert len(payload["flows"]) == 500


def test_flow_tool_manual_start_saves_history_with_trigger_type(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("mysql", "192.168.107.81", 3306, "reg-report-analysis", "", "u", "p"),
                    is_default=True,
                )
            ],
            flow_tool=FlowToolSettings(
                source_id="source-report",
                execute_url="http://192.168.107.81/assmag/spiderFlow/spider/testRun",
                chains=[
                    FlowChainConfig(
                        id="chain-zgxg-1",
                        name="资管新规1",
                        steps=[FlowChainStep(flow_id="flow-a", name="流程A")],
                    )
                ],
            ),
        ),
        config_path,
    )
    executor = FakeFlowChainExecutor()
    router = ApiRouter(config_path=config_path, flow_chain_executor=executor)

    status, payload = router.handle("POST", "/api/tools/flow/start", {"chain_id": "chain-zgxg-1"})

    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/tools/flow/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)
    assert status_payload["job"]["chain_name"] == "资管新规1"
    assert status_payload["job"]["trigger_type"] == "manual"
    assert status_payload["job"]["steps"][0]["sp_task_id"] == 658149
    assert executor.calls[0]["context"].trigger_type == "manual"

    status, history_payload = router.handle("GET", "/api/tools/flow/history", None)

    assert status == 200
    assert history_payload["history"][0]["chain_name"] == "资管新规1"
    assert history_payload["history"][0]["trigger_type"] == "manual"
    assert history_payload["history"][0]["steps"][0]["end_time"] == "2026-06-11 16:36:10"


def test_flow_chain_status_returns_active_job_payload(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("mysql", "192.168.107.81", 3306, "reg-report-analysis", "", "u", "p"),
                    is_default=True,
                )
            ],
            flow_tool=FlowToolSettings(
                source_id="source-report",
                execute_url="http://192.168.107.81/assmag/spiderFlow/spider/testRun",
                chains=[
                    FlowChainConfig(
                        id="chain-zgxg-1",
                        name="资管新规1",
                        steps=[FlowChainStep(flow_id="flow-a", name="流程A")],
                    )
                ],
            ),
        ),
        config_path,
    )
    executor = FakeFlowChainExecutor()
    router = ApiRouter(config_path=config_path, flow_chain_executor=executor)

    status, payload = router.handle("POST", "/api/tools/flow/start", {"chain_id": "chain-zgxg-1"})
    assert status == 200
    job_id = payload["job_id"]

    for _ in range(20):
        status, status_payload = router.handle("GET", "/api/flow-chain/status", None)
        if status == 200:
            assert status_payload["job"]["id"] == job_id
            assert status_payload["job"]["chain_name"] == "资管新规1"
            break
        time.sleep(0.05)
    else:
        raise AssertionError("active flow chain job not found")

    for _ in range(40):
        status, status_payload = router.handle("GET", "/api/flow-chain/status", None)
        if status == 200 and status_payload["job"] is None:
            break
        time.sleep(0.05)
    else:
        raise AssertionError("flow chain job did not finish")


def test_flow_chain_cancel_accepts_job_id_and_keeps_cancelled_status(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("mysql", "192.168.107.81", 3306, "reg-report-analysis", "", "u", "p"),
                    is_default=True,
                )
            ],
            flow_tool=FlowToolSettings(
                source_id="source-report",
                execute_url="http://192.168.107.81/assmag/spiderFlow/spider/testRun",
                chains=[
                    FlowChainConfig(
                        id="chain-zgxg-1",
                        name="资管新规1",
                        steps=[FlowChainStep(flow_id="flow-a", name="流程A")],
                    )
                ],
            ),
        ),
        config_path,
    )
    executor = BlockingFlowChainExecutor()
    router = ApiRouter(config_path=config_path, flow_chain_executor=executor)

    status, payload = router.handle("POST", "/api/tools/flow/start", {"chain_id": "chain-zgxg-1"})
    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        if executor.calls:
            break
        time.sleep(0.02)
    else:
        raise AssertionError("flow chain executor did not start")

    status, cancel_payload = router.handle("POST", "/api/tools/flow/cancel", {"job_id": job_id})

    assert status == 200
    assert cancel_payload["ok"] is True
    assert cancel_payload["job"]["status"] == "cancelled"
    assert cancel_payload["job"]["step"] == "已取消"

    for _ in range(40):
        status, status_payload = router.handle("GET", f"/api/tools/flow/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["finished_at"]:
            break
        time.sleep(0.02)
    assert status_payload["job"]["status"] == "cancelled"
    assert status_payload["job"]["error"] == "流程执行已取消"


def test_flow_chain_failure_saves_history_with_error_and_logs(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("mysql", "192.168.107.81", 3306, "reg-report-analysis", "", "u", "p"),
                    is_default=True,
                )
            ],
            flow_tool=FlowToolSettings(
                source_id="source-report",
                execute_url="http://192.168.107.81/assmag/spiderFlow/spider/testRun",
                chains=[
                    FlowChainConfig(
                        id="chain-zgxg-1",
                        name="资管新规1",
                        steps=[FlowChainStep(flow_id="flow-a", name="流程A"), FlowChainStep(flow_id="flow-b", name="流程B")],
                    )
                ],
            ),
        ),
        config_path,
    )
    executor = FakeFailingFlowChainExecutor(error_message="流程B超时：等待流程执行结束超时")
    router = ApiRouter(config_path=config_path, flow_chain_executor=executor)

    status, payload = router.handle("POST", "/api/tools/flow/start", {"chain_id": "chain-zgxg-1"})
    assert status == 200
    job_id = payload["job_id"]

    for _ in range(40):
        status, status_payload = router.handle("GET", f"/api/tools/flow/status/{job_id}", None)
        if status == 200 and status_payload["job"]["status"] == "failed":
            break
        time.sleep(0.05)
    else:
        raise AssertionError("flow chain job did not fail")

    assert status_payload["job"]["error"] == "流程B超时：等待流程执行结束超时"
    assert len(status_payload["job"]["logs"]) > 0

    status, history_payload = router.handle("GET", "/api/tools/flow/history", None)
    assert status == 200
    assert len(history_payload["history"]) == 1
    entry = history_payload["history"][0]
    assert entry["chain_name"] == "资管新规1"
    assert entry["status"] == "failed"
    assert entry["error"] == "流程B超时：等待流程执行结束超时"
    assert len(entry["logs"]) > 0
    assert any("流程B执行失败" in log.get("message", "") for log in entry["logs"])


def test_flow_chain_success_persists_logs_to_mysql_storage(tmp_path, shared_application_database):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-report",
                    name="申报平台库",
                    config=DataSourceConfig("mysql", "192.168.107.81", 3306, "reg-report-analysis", "", "u", "p"),
                    is_default=True,
                )
            ],
            flow_tool=FlowToolSettings(
                source_id="source-report",
                execute_url="http://192.168.107.81/assmag/spiderFlow/spider/testRun",
                chains=[
                    FlowChainConfig(
                        id="chain-zgxg-1",
                        name="资管新规1",
                        steps=[FlowChainStep(flow_id="flow-a", name="流程A")],
                    )
                ],
            ),
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path, flow_chain_executor=FakeFlowChainExecutor())

    status, payload = router.handle("POST", "/api/tools/flow/start", {"chain_id": "chain-zgxg-1"})
    assert status == 200
    job_id = payload["job_id"]

    for _ in range(40):
        status, status_payload = router.handle("GET", f"/api/tools/flow/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)
    else:
        raise AssertionError("flow chain job did not complete")

    logs = [
        (row["message"], row["progress"], row["step"])
        for row in shared_application_database.connection.tables["flow_chain_run_logs"]
        if row["run_id"] == job_id
    ]

    assert logs
    assert any("流程A执行结束" in message for message, _progress, _step in logs)


def test_configs_api_returns_single_data_sources(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="DWS 数据源",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
                DataSourceEntry(
                    id="source-business",
                    name="报表库数据源",
                    config=DataSourceConfig("mysql", "localhost", 3306, "bizdb", "", "u2", "p2"),
                ),
            ],
            reconcile_data_sources=ReconcileDataSourceSettings(
                dws_source_id="source-dws",
                business_source_id="source-business",
            ),
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path)

    status, payload = router.handle("GET", "/api/configs", None)

    assert status == 200
    assert payload["default_source_id"] == "source-dws"
    assert payload["data_sources"][0]["id"] == "source-dws"
    assert payload["data_sources"][0]["name"] == "DWS 数据源"
    assert payload["data_sources"][0]["password_set"] is True
    assert "dws" not in payload["data_sources"][0]
    assert "business" not in payload["data_sources"][0]


def test_referenced_data_source_cannot_be_deleted(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="DWS 数据源",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
                DataSourceEntry(
                    id="source-business",
                    name="报表库数据源",
                    config=DataSourceConfig("mysql", "localhost", 3306, "bizdb", "", "u2", "p2"),
                ),
            ],
            reconcile_data_sources=ReconcileDataSourceSettings(
                dws_source_id="source-dws",
                business_source_id="source-business",
            ),
            db_validation=DbValidationSettings(
                detail=DbValidationDatasetSettings(source_id="source-dws"),
                field_mapping_source_id="source-dws",
            ),
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path)

    status, payload = router.handle("DELETE", "/api/configs", {"id": "source-dws"})

    assert status == 400
    assert "正在被使用" in payload["error"]


def test_pbc_import_upload_saves_zip_and_returns_detected_columns(tmp_path):
    router = ApiRouter(config_path=tmp_path / "config.json")

    status, payload = router.handle_pbc_import_upload(
        "pbc.zip",
        _zip_bytes({"fund.csv": "Product Code,Product Name,Drop Me\nP1,Product One,X\n"}),
    )

    assert status == 200
    assert payload["upload_id"]
    assert payload["columns"] == ["Product Code", "Product Name", "Drop Me"]
    assert payload["files"][0]["name"] == "fund.csv"


def test_pbc_import_upload_accepts_single_csv_and_returns_single_file(tmp_path):
    router = ApiRouter(config_path=tmp_path / "config.json")

    status, payload = router.handle_pbc_import_upload(
        "public_information.csv",
        "Product Code,Product Name\nP1,Product One\n".encode("utf-8-sig"),
    )

    assert status == 200
    assert payload["upload_id"]
    assert payload["filename"] == "public_information.csv"
    assert payload["upload_ext"] == ".csv"
    assert payload["columns"] == ["Product Code", "Product Name"]
    assert payload["files"][0] == {
        "name": "public_information.csv",
        "file_type": "csv",
        "columns": ["Product Code", "Product Name"],
        "header_row": 1,
        "data_start_row": 2,
        "detection": "default",
        "matched_columns": [],
    }


def test_pbc_import_start_runs_background_job_and_saves_recent_table(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )
    importer = FakePbcImporter()
    router = ApiRouter(config_path=config_path, pbc_import_executor=importer, pbc_table_column_loader=fake_any_pbc_table_columns)
    _, upload_payload = router.handle_pbc_import_upload(
        "pbc.zip",
        _zip_bytes({"fund.csv": "Product Code,Product Name,Drop Me\nP1,Product One,X\nP2,Product Two,Y\n"}),
    )

    status, payload = router.handle(
        "POST",
        "/api/tools/pbc-import/start",
        {
            "upload_id": upload_payload["upload_id"],
            "config_name": "local - DWS",
            "source": "dws",
            "target_table": "dws.aainfo",
            "columns": ["Product Code", "Product Name", "Drop Me"],
            "drop_columns": ["Drop Me"],
            "column_order": ["Product Name", "Product Code"],
            "column_mappings": [
                {"source_column": "Product Name", "target_column": "product_name", "target_comment": "Product Name"},
                {"source_column": "Product Code", "target_column": "product_code", "target_comment": "Product Code"},
            ],
            "mode": "append",
        },
    )

    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/tools/pbc-import/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)
    job = status_payload["job"]
    assert job["rows_imported"] == 2
    assert importer.calls[0]["table"].parts == ("dws", "aainfo")
    assert importer.calls[0]["column_order"] == ["Product Name", "Product Code"]
    assert [(mapping.source_column, mapping.target_column) for mapping in importer.calls[0]["column_mappings"]] == [
        ("Product Name", "product_name"),
        ("Product Code", "product_code"),
    ]

    _, settings_payload = router.handle("GET", "/api/tools/pbc-import/settings", None)
    assert settings_payload["settings"]["recent_tables"] == ["dws.aainfo"]


def test_pbc_import_start_accepts_multiple_uploaded_files(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )
    importer = FakePbcImporter()
    router = ApiRouter(config_path=config_path, pbc_import_executor=importer, pbc_table_column_loader=fake_any_pbc_table_columns)
    _, archive_payload = router.handle_pbc_import_upload(
        "pbc.zip",
        _zip_bytes({"fund.csv": "Product Code,Product Name\nP1,Product One\n"}),
    )
    _, csv_payload = router.handle_pbc_import_upload(
        "single.csv",
        "Product Code,Product Name\nP2,Product Two\n".encode("utf-8-sig"),
    )

    status, payload = router.handle(
        "POST",
        "/api/tools/pbc-import/start",
        {
            "upload_ids": [archive_payload["upload_id"], csv_payload["upload_id"]],
            "config_name": "local - DWS",
            "source": "dws",
            "target_table": "dws.aainfo",
            "columns": ["Product Code", "Product Name"],
            "column_mappings": [
                {"source_column": "Product Code", "target_column": "product_code", "target_comment": "Product Code"},
            ],
        },
    )

    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/tools/pbc-import/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)
    assert len(importer.calls[0]["zip_path"]) == 2
    assert {path.name for path in importer.calls[0]["zip_path"]} == {
        f"{archive_payload['upload_id']}.zip",
        f"{csv_payload['upload_id']}.csv",
    }


def test_pbc_import_start_rejects_same_table_while_job_is_running(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )
    importer = SlowPbcImporter()
    router = ApiRouter(config_path=config_path, pbc_import_executor=importer, pbc_table_column_loader=fake_any_pbc_table_columns)
    _, first_upload = router.handle_pbc_import_upload(
        "first.csv",
        "Product Code,Product Name\nP1,Product One\n".encode("utf-8-sig"),
    )
    _, other_table_upload = router.handle_pbc_import_upload(
        "other.csv",
        "Product Code,Product Name\nP3,Product Three\n".encode("utf-8-sig"),
    )
    _, second_upload = router.handle_pbc_import_upload(
        "second.csv",
        "Product Code,Product Name\nP2,Product Two\n".encode("utf-8-sig"),
    )
    request = {
        "config_name": "local - DWS",
        "source": "dws",
        "target_table": "dws.aainfo",
        "columns": ["Product Code", "Product Name"],
        "column_mappings": [
            {"source_column": "Product Code", "target_column": "product_code", "target_comment": "Product Code"},
        ],
    }

    status, payload = router.handle("POST", "/api/tools/pbc-import/start", {**request, "upload_id": first_upload["upload_id"]})
    assert status == 200
    first_job_id = payload["job_id"]
    assert importer.started.wait(timeout=2)

    status, payload = router.handle(
        "POST",
        "/api/tools/pbc-import/start",
        {**request, "target_table": "dws.other_table", "upload_id": other_table_upload["upload_id"]},
    )
    assert status == 200

    status, payload = router.handle("POST", "/api/tools/pbc-import/start", {**request, "upload_id": second_upload["upload_id"]})
    assert status == 409
    assert "正在导入" in payload["error"]
    assert "等待上一个任务完成" in payload["error"]

    importer.release.set()
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/tools/pbc-import/status/{first_job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)

    status, payload = router.handle("POST", "/api/tools/pbc-import/start", {**request, "upload_id": second_upload["upload_id"]})
    assert status == 200


def test_pbc_import_columns_api_maps_source_columns_to_target_comments(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path, pbc_table_column_loader=fake_table_columns)

    status, payload = router.handle(
        "POST",
        "/api/tools/pbc-import/columns",
        {
            "config_name": "local - DWS",
            "source": "dws",
            "target_table": "dws.public_information_th",
            "source_columns": ["Info Type Name", "Product Code", "Not Existing"],
        },
    )

    assert status == 200
    assert payload["table_columns"][0] == {"name": "info_type_name", "comment": "Info Type Name"}
    assert [(item["source_column"], item["target_column"]) for item in payload["mappings"]] == [
        ("Info Type Name", "info_type_name"),
        ("Product Code", "product_code"),
        ("Not Existing", ""),
    ]


def test_pbc_import_columns_api_detects_header_rows_from_uploaded_files(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path, pbc_table_column_loader=fake_any_pbc_table_columns)
    _, upload_payload = router.handle_pbc_import_upload(
        "single.csv",
        "产品公开信息报表\nProduct Code,Product Name\nP1,Product One\n".encode("utf-8-sig"),
    )

    status, payload = router.handle(
        "POST",
        "/api/tools/pbc-import/columns",
        {
            "config_name": "local - DWS",
            "source": "dws",
            "target_table": "dws.aainfo",
            "source_columns": ["产品公开信息报表"],
            "upload_ids": [upload_payload["upload_id"]],
        },
    )

    assert status == 200
    assert payload["source_columns"] == ["Product Code", "Product Name"]
    assert [(item["source_column"], item["target_column"]) for item in payload["mappings"]] == [
        ("Product Code", "product_code"),
        ("Product Name", "product_name"),
    ]
    file_info = payload["upload_inspections"][0]["files"][0]
    assert file_info["header_row"] == 2
    assert file_info["data_start_row"] == 3
    assert file_info["detection"] == "smart"


def test_pbc_import_columns_api_returns_generic_headers_when_auto_mapping_has_no_matches(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="source-dws",
                    name="local - DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    is_default=True,
                ),
            ],
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path, pbc_table_column_loader=fake_any_pbc_table_columns)
    _, upload_payload = router.handle_pbc_import_upload(
        "single.csv",
        "Template Title\nExternal Key,External Label\nP1,Fund One\n".encode("utf-8-sig"),
    )

    status, payload = router.handle(
        "POST",
        "/api/tools/pbc-import/columns",
        {
            "config_name": "local - DWS",
            "source": "dws",
            "target_table": "dws.aainfo",
            "source_columns": ["Template Title"],
            "upload_ids": [upload_payload["upload_id"]],
        },
    )

    assert status == 200
    assert payload["source_columns"] == ["External Key", "External Label"]
    assert [(item["source_column"], item["target_column"]) for item in payload["mappings"]] == [
        ("External Key", ""),
        ("External Label", ""),
    ]
    file_info = payload["upload_inspections"][0]["files"][0]
    assert file_info["header_row"] == 2
    assert file_info["data_start_row"] == 3
    assert file_info["detection"] == "smart"


def test_db_validation_settings_api_returns_sources_and_persists_settings(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            configs=[
                NamedConfig(
                    name="local",
                    dws=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    business=DataSourceConfig("mysql", "localhost", 3306, "bizdb", "", "u2", "p2"),
                    is_default=True,
                ),
                NamedConfig(
                    name="metadata",
                    dws=DataSourceConfig("postgresql", "localhost", 5432, "metadb", "meta", "mu", "mp"),
                    business=DataSourceConfig("mysql", "localhost", 3306, "metabiz", "", "mbu", "mbp"),
                ),
            ],
            default_name="local",
            db_validation=DbValidationSettings(
                detail=DbValidationDatasetSettings(
                    config_name="local",
                    source="dws",
                    sys_manage_id="DETAIL_SYS",
                    classification_id="DETAIL_CLASS",
                ),
                public_info=DbValidationDatasetSettings(
                    config_name="metadata",
                    source="business",
                    sys_manage_id="PUBLIC_SYS",
                    classification_id="PUBLIC_CLASS",
                ),
                template=DbValidationDatasetSettings(
                    config_name="metadata",
                    source="dws",
                    sys_manage_id="TEMPLATE_SYS",
                    classification_id="TEMPLATE_CLASS",
                ),
                field_mapping_config_name="metadata",
                field_mapping_source="dws",
            ),
        ),
        config_path,
    )
    router = ApiRouter(config_path=config_path)

    status, payload = router.handle("GET", "/api/tools/db-validation/settings", None)

    assert status == 200
    assert payload["settings"]["detail"]["sys_manage_id"] == "DETAIL_SYS"
    assert payload["settings"]["detail"]["source_id"] == "legacy:local:dws"
    assert payload["settings"]["public_info"]["source_id"] == "legacy:metadata:business"
    assert payload["settings"]["field_mapping_source_id"] == "legacy:metadata:dws"
    assert payload["default_report_date"] == previous_month_end()
    assert payload["data_sources"][0]["id"] == "legacy:local:dws"
    assert payload["field_mapping"]["initialized"] is False

    status, payload = router.handle(
        "POST",
        "/api/tools/db-validation/settings",
        {
            "detail": {
                "source_id": "legacy:local:dws",
                "sys_manage_id": "D2",
                "classification_id": "C2",
            },
            "public_info": {
                "source_id": "legacy:metadata:business",
                "sys_manage_id": "P2",
                "classification_id": "PC2",
            },
            "template": {
                "source_id": "legacy:metadata:dws",
                "sys_manage_id": "T2",
                "classification_id": "TC2",
            },
            "field_mapping_source_id": "legacy:metadata:business",
        },
    )

    assert status == 200
    assert payload["settings"]["detail"]["classification_id"] == "C2"
    assert payload["settings"]["field_mapping_source_id"] == "legacy:metadata:business"
    assert payload["field_mapping"]["initialized"] is False
    restarted_router = ApiRouter(config_path=config_path)
    _, loaded = restarted_router.handle("GET", "/api/tools/db-validation/settings", None)
    assert loaded["settings"]["public_info"]["sys_manage_id"] == "P2"


def test_db_validation_start_runs_background_job_and_exposes_download(tmp_path, shared_application_database):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            configs=[
                NamedConfig(
                    name="local",
                    dws=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    business=DataSourceConfig("mysql", "localhost", 3306, "bizdb", "", "u2", "p2"),
                    is_default=True,
                ),
                NamedConfig(
                    name="metadata",
                    dws=DataSourceConfig("postgresql", "localhost", 5432, "metadb", "meta", "mu", "mp"),
                    business=DataSourceConfig("mysql", "localhost", 3306, "metabiz", "", "mbu", "mbp"),
                ),
            ],
            default_name="local",
            db_validation=DbValidationSettings(
                detail=DbValidationDatasetSettings(
                    config_name="local",
                    source="dws",
                    sys_manage_id="DETAIL_SYS",
                    classification_id="DETAIL_CLASS",
                ),
                public_info=DbValidationDatasetSettings(
                    config_name="metadata",
                    source="business",
                    sys_manage_id="PUBLIC_SYS",
                    classification_id="PUBLIC_CLASS",
                ),
                field_mapping_config_name="metadata",
                field_mapping_source="dws",
            ),
        ),
        config_path,
    )
    executor = FakeDbValidationExecutor()
    field_mapping_loader = FakeDbValidationFieldMappingLoader()
    router = ApiRouter(
        config_path=config_path,
        db_validation_executor=executor,
        db_validation_field_mapping_loader=field_mapping_loader,
    )

    status, payload = router.handle(
        "POST",
        "/api/tools/db-validation/start",
        {
            "report_date": "2026-05-31",
            "selected_tables": ["ZG01"],
            "enable_public_info_check": True,
            "enable_template_check": True,
        },
        current_user={"id": "u1", "username": "operator", "display_name": "张三"},
    )

    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/tools/db-validation/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)
    job = status_payload["job"]
    assert job["result"]["error_count"] == 1
    assert job["result"]["warnings"] == ["ZG02 当期表无数据"]
    assert job["download_url"] == f"/api/tools/db-validation/download/{job_id}"
    assert executor.calls[0]["data_source"].database == "dwdb"
    assert executor.calls[0]["metadata_source"].database == "metadb"
    assert executor.calls[0]["public_info_source"].database == "metabiz"
    assert executor.calls[0]["template_source"].database == "dwdb"
    assert executor.calls[0]["selected_tables"] == ["ZG01"]
    assert executor.calls[0]["enable_public_info_check"] is True
    assert executor.calls[0]["enable_template_check"] is True
    assert executor.calls[0]["detail_sys_manage_id"] == "DETAIL_SYS"
    assert executor.calls[0]["public_info_classification_id"] == "PUBLIC_CLASS"
    assert isinstance(executor.calls[0]["field_catalog"], TableFieldCatalog)
    download_path, download_name = router.get_db_validation_download(job_id)
    assert download_path.exists()
    assert download_name == "result.xlsx"
    for _ in range(20):
        history_status, history_payload = router.handle("GET", "/api/tools/db-validation/history", None)
        if history_payload["history"]:
            break
        time.sleep(0.05)
    assert history_status == 200
    history = history_payload["history"][0]
    assert history["id"] == job_id
    assert "T" not in history["run_at"]
    assert history["report_date"] == "2026-05-31"
    assert history["result_count"] == 1
    assert history["executor_id"] == "u1"
    assert history["executor_username"] == "operator"
    assert history["executor_name"] == "张三"
    assert history["enable_public_info_check"] is True
    assert history["enable_template_check"] is True
    assert history["download_url"] == f"/api/tools/db-validation/history/download/{job_id}"
    assert "rows" not in history
    history_path, history_name = router.get_db_validation_history_download(job_id)
    assert history_path == download_path
    assert history_name == "result.xlsx"

    warning_messages = [
        row["message"]
        for row in shared_application_database.connection.tables["db_validation_warnings"]
        if row["run_id"] == job_id
    ]
    result_rows = [
        row["payload_json"]
        for row in shared_application_database.connection.tables["db_validation_result_rows"]
        if row["run_id"] == job_id
    ]
    assert warning_messages == ["ZG02 当期表无数据"]
    assert len(result_rows) == 1
    assert "Zg01_Rule6" in result_rows[0]


def test_db_validation_history_api_sorts_by_execution_time_desc(tmp_path):
    router = ApiRouter(config_path=tmp_path / "config.json")
    router.db_validation_history_store.save_run(
        {
            "id": "newer-report-earlier-execution",
            "run_at": "2026-06-08 11:01:59",
            "run_date": "2026-06-30",
            "report_date": "2026-06-30",
            "excel_path": str(tmp_path / "earlier.xlsx"),
        }
    )
    router.db_validation_history_store.save_run(
        {
            "id": "same-day-latest",
            "run_at": "2026-06-08 14:27:15",
            "run_date": "2026-05-31",
            "report_date": "2026-05-31",
            "excel_path": str(tmp_path / "latest.xlsx"),
        }
    )
    router.db_validation_history_store.save_run(
        {
            "id": "same-day-middle",
            "run_at": "2026-06-08 14:03:50",
            "run_date": "2026-05-31",
            "report_date": "2026-05-31",
            "excel_path": str(tmp_path / "middle.xlsx"),
        }
    )

    status, payload = router.handle("GET", "/api/tools/db-validation/history", None)

    assert status == 200
    assert [entry["id"] for entry in payload["history"]] == [
        "same-day-latest",
        "same-day-middle",
        "newer-report-earlier-execution",
    ]


def test_db_validation_reuses_cached_field_mapping_and_manual_refreshes(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            configs=[
                NamedConfig(
                    name="local",
                    dws=DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p"),
                    business=DataSourceConfig("mysql", "localhost", 3306, "bizdb", "", "u2", "p2"),
                    is_default=True,
                ),
                NamedConfig(
                    name="metadata",
                    dws=DataSourceConfig("postgresql", "localhost", 5432, "metadb", "meta", "mu", "mp"),
                    business=DataSourceConfig("mysql", "localhost", 3306, "metabiz", "", "mbu", "mbp"),
                ),
            ],
            default_name="local",
            db_validation=DbValidationSettings(
                detail=DbValidationDatasetSettings(
                    config_name="local",
                    source="dws",
                    sys_manage_id="DETAIL_SYS;DETAIL_SYS_2",
                    classification_id="DETAIL_CLASS",
                ),
                field_mapping_config_name="metadata",
                field_mapping_source="dws",
                baseinfo_table="xt_reg_table_baseinfo",
                field_info_table="xt_reg_table_field_info",
            ),
        ),
        config_path,
    )
    executor = FakeDbValidationExecutor()
    field_mapping_loader = FakeDbValidationFieldMappingLoader()
    router = ApiRouter(
        config_path=config_path,
        db_validation_executor=executor,
        db_validation_field_mapping_loader=field_mapping_loader,
    )

    for _ in range(2):
        status, payload = router.handle(
            "POST",
            "/api/tools/db-validation/start",
            {"report_date": "2026-05-31", "selected_tables": ["ZG01"]},
        )
        assert status == 200
        job_id = payload["job_id"]
        for _ in range(20):
            status, status_payload = router.handle("GET", f"/api/tools/db-validation/status/{job_id}", None)
            assert status == 200
            if status_payload["job"]["status"] == "completed":
                break
            time.sleep(0.05)
        assert status_payload["job"]["status"] == "completed"

    assert len(field_mapping_loader.calls) == 1
    assert executor.calls[0]["field_catalog"] is executor.calls[1]["field_catalog"]
    assert field_mapping_loader.calls[0]["metadata_source"].database == "metadb"
    assert field_mapping_loader.calls[0]["sys_manage_id"] == "DETAIL_SYS;DETAIL_SYS_2"

    status, payload = router.handle("POST", "/api/tools/db-validation/field-mapping/refresh", {})

    assert status == 200
    assert payload["field_mapping"]["initialized"] is True
    assert payload["field_mapping"]["field_count"] == 2
    assert payload["field_mapping"]["unmapped_field_count"] == 1
    assert payload["field_mapping"]["refresh_source"] == "manual"
    assert len(field_mapping_loader.calls) == 2

    status, payload = router.handle(
        "POST",
        "/api/tools/db-validation/start",
        {"report_date": "2026-05-31", "selected_tables": ["ZG01"]},
    )
    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/tools/db-validation/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)

    assert status_payload["job"]["status"] == "completed"
    assert len(field_mapping_loader.calls) == 2
    assert executor.calls[2]["field_catalog"] is not executor.calls[0]["field_catalog"]


def test_db_validation_field_mapping_refresh_returns_readable_status_on_failure(tmp_path):
    class FailingFieldMappingLoader:
        def __call__(self, **kwargs):
            raise RuntimeError("baseinfo 表 xt_reg_table_baseinfo 不存在")

    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            configs=[
                NamedConfig(
                    name="metadata",
                    dws=DataSourceConfig("postgresql", "localhost", 5432, "metadb", "meta", "mu", "mp"),
                    business=DataSourceConfig("postgresql", "localhost", 5432, "unused", "dws", "u", "p"),
                    is_default=True,
                )
            ],
            default_name="metadata",
            db_validation=DbValidationSettings(
                detail=DbValidationDatasetSettings(config_name="metadata", source="dws"),
                field_mapping_config_name="metadata",
                field_mapping_source="dws",
                baseinfo_table="xt_reg_table_baseinfo",
                field_info_table="xt_reg_table_field_info",
            ),
        ),
        config_path,
    )
    router = ApiRouter(
        config_path=config_path,
        db_validation_field_mapping_loader=FailingFieldMappingLoader(),
    )

    status, payload = router.handle("POST", "/api/tools/db-validation/field-mapping/refresh", {})

    assert status == 200
    assert payload["field_mapping"]["initialized"] is False
    assert "xt_reg_table_baseinfo" in payload["field_mapping"]["last_error"]


def test_db_validation_rules_document_download_is_valid_workbook(tmp_path):
    router = ApiRouter(config_path=tmp_path / "config.json")

    filename, payload = router.get_db_validation_rules_document()

    assert filename == "数据库校验规则说明.xlsx"
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    assert "规则明细" in workbook.sheetnames
    values = [cell for row in workbook["规则明细"].iter_rows(values_only=True) for cell in row if cell]
    assert "Zg06_Rule6" in values
    workbook.close()


def test_post_run_returns_serialized_results(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle(
        "POST",
        "/api/run",
        {"date": "2026-04-30"},
        current_user={"id": "u1", "username": "operator", "display_name": "张三"},
    )

    assert status == 200
    assert payload["results"][0]["project_code"] == "P1"
    assert payload["results"][0]["difference"] == "10"
    details = payload["results"][0]["display_details"]
    assert [detail["title"] for detail in details] == ["最终判断结果"]
    assert details[0]["rows"][0] == {"label": "差异类型", "value": "实收本金差异"}
    assert {"label": "c1000 实收本金余额", "value": "50"} in details[0]["rows"]
    assert {"label": "FA 4001 科目余额", "value": "500"} in details[0]["rows"]
    assert "项目差异概览" not in [detail["title"] for detail in details]
    assert "估值表匹配过程" not in [detail["title"] for detail in details]
    assert payload["history"]["total_count"] == 1
    assert payload["history"]["added_count"] is None
    assert payload["history"]["removed_count"] is None
    assert payload["history"]["executor_id"] == "u1"
    assert payload["history"]["executor_username"] == "operator"
    assert payload["history"]["executor_name"] == "张三"


def test_post_run_with_no_source_report_data_returns_notice_without_history(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: NoSourceRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle("POST", "/api/run", {"date": "2026-04-30"})

    assert status == 200
    assert payload["results"] == []
    assert payload["history"]["no_source_data"] is True
    assert payload["history"]["run_date"] == "2026-04-30"
    assert "报表对应日期无数据" in payload["history"]["message"]
    _, history_payload = router.handle("GET", "/api/history", None)
    assert history_payload["history"] == []


def test_run_history_uses_reconcile_business_data_source_name(tmp_path):
    config_path = tmp_path / "config.json"
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(
                    id="default-dws",
                    name="默认 DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "default_dws", "dws", "u", "p"),
                    is_default=True,
                ),
                DataSourceEntry(
                    id="selected-dws",
                    name="对账 DWS",
                    config=DataSourceConfig("postgresql", "localhost", 5432, "selected_dws", "dws", "u", "p"),
                ),
                DataSourceEntry(
                    id="selected-business",
                    name="对账报表库",
                    config=DataSourceConfig("mysql", "localhost", 3306, "selected_biz", "", "u", "p"),
                ),
            ],
            reconcile_data_sources=ReconcileDataSourceSettings(
                dws_source_id="selected-dws",
                business_source_id="selected-business",
            ),
        ),
        config_path,
    )
    router = ApiRouter(
        config_path=config_path,
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle("POST", "/api/run", {"date": "2026-04-30"})

    assert status == 200
    assert payload["history"]["config_name"] == "对账报表库"
    assert payload["history"]["dws_source_name"] == "对账 DWS"
    _, history_payload = router.handle("GET", "/api/history", None)
    assert history_payload["history"][0]["config_name"] == "对账报表库"
    assert history_payload["history"][0]["dws_source_name"] == "对账 DWS"


def test_run_once_passes_reconcile_schema_and_source_lookup_to_repository(tmp_path, monkeypatch):
    config_path = tmp_path / "config.json"
    dws_source = DataSourceConfig("postgresql", "localhost", 5432, "dwdb", "dws", "u", "p")
    report_source = DataSourceConfig("mysql", "localhost", 3306, "reportdb", "", "u2", "p2")
    save_store(
        ConfigStore(
            data_sources=[
                DataSourceEntry(id="source-dws", name="DWS", config=dws_source, is_default=True),
                DataSourceEntry(id="source-report", name="report-db", config=report_source),
            ],
            reconcile_data_sources=ReconcileDataSourceSettings(
                dws_source_id="source-dws",
                business_source_id="source-report",
            ),
            reconcile_schema=ReconcileSchemaSettings(
                tables={
                    "fa_valuation": ReconcileTableSchema(
                        source_ref=ReconcileSourceRef(id="source-dws", name="DWS"),
                        table="custom.fa_valuation",
                        fields={"market_value": "market_amt"},
                    )
                }
            ),
        ),
        config_path,
    )
    captured = {}

    class FakeEngine:
        def __init__(self, repository, **kwargs):
            captured["schema"] = repository.schema
            captured["source_configs"] = repository._source_configs

        def run(self, date):
            return []

    monkeypatch.setattr(server_module, "ReconcileEngine", FakeEngine)
    router = ApiRouter(config_path=config_path, history_path=tmp_path / "history.json")

    router._run_once("2026-04-30", max_combination_rows=50)

    assert captured["schema"].tables["fa_valuation"].table == "custom.fa_valuation"
    assert captured["schema"].tables["fa_valuation"].fields["market_value"] == "market_amt"
    assert captured["source_configs"]["source-dws"] == dws_source
    assert captured["source_configs"]["source-report"] == report_source
    assert captured["source_configs"]["report-db"] == report_source


def test_run_start_status_returns_background_job_result(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle(
        "POST",
        "/api/run/start",
        {"date": "2026-04-30", "max_combination_rows": 50},
        current_user={"id": "u1", "username": "operator", "display_name": "张三"},
    )

    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/run/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)
    job = status_payload["job"]
    assert job["status"] == "completed"
    assert job["results"][0]["project_code"] == "P1"
    assert job["history"]["total_count"] == 1
    assert job["history"]["executor_name"] == "张三"
    assert any("后台任务已启动" in log["message"] for log in job["logs"])


def test_run_start_status_returns_no_source_notice_without_history(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: NoSourceRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle("POST", "/api/run/start", {"date": "2026-04-30"})

    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/run/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)
    job = status_payload["job"]
    assert job["status"] == "completed"
    assert job["results"] == []
    assert job["history"]["no_source_data"] is True
    assert any("报表对应日期无数据" in log["message"] for log in job["logs"])
    assert sum("报表对应日期无数据" in log["message"] for log in job["logs"]) == 1
    _, history_payload = router.handle("GET", "/api/history", None)
    assert history_payload["history"] == []


def test_run_start_status_exposes_reconcile_runtime_error_summary(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: FailingRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle("POST", "/api/run/start", {"date": "2026-04-30"})

    assert status == 200
    job_id = payload["job_id"]
    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/run/status/{job_id}", None)
        assert status == 200
        if status_payload["job"]["status"] == "failed":
            break
        time.sleep(0.05)
    job = status_payload["job"]
    assert job["status"] == "failed"
    assert "unsafe table identifier: zf_detail_2024ǮǮǮ" in job["error"]
    assert any("unsafe table identifier: zf_detail_2024ǮǮǮ" in log["message"] for log in job["logs"])


def test_run_start_rejects_when_any_run_job_is_active(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: SlowRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle(
        "POST",
        "/api/run/start",
        {"date": "2026-04-30"},
        current_user={"id": "u1", "username": "operator", "display_name": "张三"},
    )
    assert status == 200
    first_job_id = payload["job_id"]

    status, payload = router.handle(
        "POST",
        "/api/run/start",
        {"date": "2026-04-30"},
        current_user={"id": "u2", "username": "auditor", "display_name": "李四"},
    )
    assert status == 409
    assert "正在执行" in payload["error"]
    assert "请等待当前任务完成" in payload["error"]
    assert payload["active_job"]["id"] == first_job_id
    assert payload["active_job"]["executor"]["display_name"] == "张三"

    for _ in range(20):
        status, status_payload = router.handle("GET", f"/api/run/status/{first_job_id}", None)
        if status == 200 and status_payload["job"]["status"] == "completed":
            break
        time.sleep(0.05)

    status, payload = router.handle(
        "POST",
        "/api/run/start",
        {"date": "2026-04-30"},
        current_user={"id": "u2", "username": "auditor", "display_name": "李四"},
    )
    assert status == 200


def test_run_cancel_endpoint_marks_running_job_cancelled_immediately(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: SlowRunner(),
        connection_tester=FakeConnectionTester(),
    )

    _, payload = router.handle("POST", "/api/run/start", {"date": "2026-04-30"})
    status, cancel_payload = router.handle("POST", "/api/run/cancel", {"job_id": payload["job_id"]})

    assert status == 200
    assert cancel_payload["ok"] is True
    assert cancel_payload["job"]["status"] == "cancelled"
    assert cancel_payload["job"]["step"] == "已终止"
    assert any("停止执行" in log["message"] for log in cancel_payload["job"]["logs"])


def test_history_api_lists_detail_and_deletes_entries(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        history_path=tmp_path / "history.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )
    _, run_payload = router.handle(
        "POST",
        "/api/run",
        {"date": "2026-04-30"},
        current_user={"id": "u1", "username": "operator", "display_name": "张三"},
    )
    history_id = run_payload["history"]["id"]

    status, list_payload = router.handle("GET", "/api/history", None)
    assert status == 200
    assert list_payload["history"][0]["id"] == history_id
    assert list_payload["history"][0]["executor_name"] == "张三"
    assert "results" not in list_payload["history"][0]

    status, detail_payload = router.handle("GET", f"/api/history/{history_id}", None)
    assert status == 200
    assert detail_payload["history"]["results"][0]["project_code"] == "P1"
    assert detail_payload["history"]["added_results"] == []
    assert detail_payload["history"]["removed_results"] == []

    status, delete_payload = router.handle(
        "DELETE",
        "/api/history",
        {"id": history_id},
        current_user={"id": "u2", "username": "normal", "role": "user"},
    )
    assert status == 403
    assert delete_payload["error"] == "admin role required"
    _, list_payload = router.handle("GET", "/api/history", None)
    assert list_payload["history"][0]["id"] == history_id

    status, delete_payload = router.handle(
        "DELETE",
        "/api/history",
        {"id": history_id},
        current_user={"id": "admin", "username": "admin", "role": "admin"},
    )
    assert status == 200
    assert delete_payload["ok"] is True
    _, list_payload = router.handle("GET", "/api/history", None)
    assert list_payload["history"] == []


def test_display_details_show_final_asset_gap_and_am_judgement_only():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("900"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("-100"),
        direction="资产小于负债及权益",
        difference_reason="资产缺失",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="asset_gap",
                data={
                    "reason": "资产缺失",
                    "valuation_asset_total": "1000",
                    "asset_gap": "100",
                    "match_type": "single",
                    "match_total": "100",
                },
            ),
            DifferenceDetail(
                kind="fa_am",
                data={
                    "fa_account_code": "1101.02.15.01.244733",
                    "fa_account_name": "G26资控1",
                    "fa_tail_code": "244733",
                    "am_asset_name": "G26资控1",
                    "am_stock_code": "244978",
                },
            ),
        ],
        valuation_match=ValuationMatch(
            match_type="single",
            rows=[ValuationRow("1101.02.15.01.244733", "G26资控1", Decimal("100"))],
        ),
    )

    details = build_display_details(result)

    assert [detail["title"] for detail in details] == ["最终判断结果", "具体差异明细", "标的代码核对"]
    assert details[0]["rows"][0] == {"label": "差异类型", "value": "资产缺失"}
    assert details[1]["table"]["headers"] == ["科目代码", "科目名称", "科目尾段代码", "金额"]
    assert {"label": "AM 标的代码", "value": "244978"} in details[2]["rows"]


def test_display_details_show_asset_type_specific_reason_from_asset_gap():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("900"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("-100"),
        direction="资产小于负债及权益",
        difference_reason="资产缺失",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="asset_gap",
                data={
                    "reason": "资产缺失",
                    "valuation_asset_total": "1000",
                    "asset_gap": "100",
                    "match_type": "single",
                    "match_total": "100",
                    "specific_reason": "债券资产缺失",
                },
            ),
        ],
        valuation_match=ValuationMatch(
            match_type="single",
            rows=[ValuationRow("1501.01.01.01.102381204", "23苏城投MTN004", Decimal("100"))],
        ),
    )

    details = build_display_details(result)

    assert {"label": "具体原因", "value": "债券资产缺失"} in details[0]["rows"]


def test_display_details_show_ambiguous_candidate_groups():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("950"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("50"),
        direction="资产大于负债及权益",
        difference_reason="负债及权益科目缺失 + 暂无法确定",
        match_status="候选不唯一",
        details=[
            DifferenceDetail(
                kind="liability_equity",
                data={
                    "specific_reason": "候选不唯一",
                    "match_type": "ambiguous_combination",
                    "match_total": "50",
                    "match_message": "候选不唯一",
                    "match_target": "50",
                    "account_scope": "非1开头科目",
                    "candidate_groups": [
                        {
                            "index": "候选组合1",
                            "total": "50",
                            "rows": [
                                {
                                    "account_code": "2209.01.01.01.A",
                                    "account_name": "应付管理费A",
                                    "account_tail": "A",
                                    "market_value": "20",
                                },
                                {
                                    "account_code": "2209.01.01.01.B",
                                    "account_name": "应付托管费B",
                                    "account_tail": "B",
                                    "market_value": "30",
                                },
                            ],
                        },
                        {
                            "index": "候选组合2",
                            "total": "50",
                            "rows": [
                                {
                                    "account_code": "2221.01.01.01.C",
                                    "account_name": "应交税费C",
                                    "account_tail": "C",
                                    "market_value": "10",
                                },
                                {
                                    "account_code": "2221.01.01.01.D",
                                    "account_name": "其他应付款D",
                                    "account_tail": "D",
                                    "market_value": "40",
                                },
                            ],
                        },
                    ],
                },
            ),
        ],
        valuation_match=ValuationMatch(match_type="ambiguous_combination", message="候选不唯一"),
    )

    details = build_display_details(result)

    assert [detail["title"] for detail in details] == ["最终判断结果", "候选组合明细"]
    assert details[0]["rows"][0] == {"label": "差异类型", "value": "负债及权益科目缺失 + 暂无法确定"}
    assert {"label": "具体原因", "value": "候选不唯一"} in details[0]["rows"]
    assert {"label": "命中方式", "value": "候选不唯一"} in details[0]["rows"]
    assert details[1]["table"]["headers"] == ["候选组合", "组内合计", "科目代码", "科目名称", "科目尾段", "金额"]
    assert details[1]["table"]["rows"] == [
        ["候选组合1", "50", "2209.01.01.01.A", "应付管理费A", "A", "20"],
        ["候选组合1", "50", "2209.01.01.01.B", "应付托管费B", "B", "30"],
        ["候选组合2", "50", "2221.01.01.01.C", "应交税费C", "C", "10"],
        ["候选组合2", "50", "2221.01.01.01.D", "其他应付款D", "D", "40"],
    ]


def test_display_details_show_asset_gap_ambiguous_candidate_groups():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("900"),
        liability_equity_total=Decimal("800"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("100"),
        direction="资产大于负债及权益",
        difference_reason="资产缺失 + 暂无法确定",
        match_status="候选不唯一",
        details=[
            DifferenceDetail(
                kind="asset_gap",
                data={
                    "reason": "资产缺失",
                    "zf_asset_total": "900",
                    "valuation_asset_total": "950",
                    "asset_gap": "50",
                    "specific_reason": "候选不唯一",
                    "match_type": "ambiguous_combination",
                    "match_total": "50",
                    "match_message": "候选不唯一",
                    "account_scope": "1开头末级科目",
                    "candidate_groups": [
                        {
                            "index": "候选组合1",
                            "total": "50",
                            "rows": [
                                {
                                    "account_code": "1001.01.01.01.0001",
                                    "account_name": "资产A",
                                    "account_tail": "0001",
                                    "market_value": "20",
                                },
                                {
                                    "account_code": "1002.01.01.01.0002",
                                    "account_name": "资产B",
                                    "account_tail": "0002",
                                    "market_value": "30",
                                },
                            ],
                        },
                        {
                            "index": "候选组合2",
                            "total": "50",
                            "rows": [
                                {
                                    "account_code": "1003.01.01.01.0003",
                                    "account_name": "资产C",
                                    "account_tail": "0003",
                                    "market_value": "10",
                                },
                                {
                                    "account_code": "1004.01.01.01.0004",
                                    "account_name": "资产D",
                                    "account_tail": "0004",
                                    "market_value": "40",
                                },
                            ],
                        },
                    ],
                },
            ),
        ],
        valuation_match=ValuationMatch(match_type="ambiguous_combination", message="候选不唯一"),
    )

    details = build_display_details(result)

    assert [detail["title"] for detail in details] == ["最终判断结果", "候选组合明细"]
    assert details[0]["rows"][0] == {"label": "差异类型", "value": "资产缺失 + 暂无法确定"}
    assert {"label": "具体原因", "value": "候选不唯一"} in details[0]["rows"]
    assert {"label": "命中方式", "value": "候选不唯一"} in details[0]["rows"]
    assert details[1]["table"]["rows"] == [
        ["候选组合1", "50", "1001.01.01.01.0001", "资产A", "0001", "20"],
        ["候选组合1", "50", "1002.01.01.01.0002", "资产B", "0002", "30"],
        ["候选组合2", "50", "1003.01.01.01.0003", "资产C", "0003", "10"],
        ["候选组合2", "50", "1004.01.01.01.0004", "资产D", "0004", "40"],
    ]


def test_display_details_show_asset_missing_refinement_table_and_specific_reason():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("900"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("-100"),
        direction="资产小于负债及权益",
        difference_reason="资产缺失",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="asset_gap",
                data={
                    "reason": "资产缺失",
                    "valuation_asset_total": "1000",
                    "asset_gap": "100",
                    "match_type": "single",
                    "match_total": "100",
                },
            ),
            DifferenceDetail(
                kind="asset_missing_refinement",
                data={
                    "specific_reason": "①债券缺失：23苏城投MTN004；原因：资负数据子系统-债务证券明细表无数据",
                    "rows": [
                        {
                            "index": "①",
                            "asset_type": "债券",
                            "asset_name": "23苏城投MTN004",
                            "fa_account_code": "1501.01.01.01.102381204",
                            "account_tail": "102381204",
                            "fa_market_value": "100",
                            "check_table": "currency_report_24.currency_detail_project_2_1_4",
                            "check_result": "资负数据子系统-债务证券明细表无数据",
                            "key_field": "",
                            "am_stock_code": "",
                            "pact_id": "",
                            "reason": "资负数据子系统-债务证券明细表无数据",
                        }
                    ],
                },
            ),
        ],
        valuation_match=ValuationMatch(
            match_type="single",
            rows=[ValuationRow("1501.01.01.01.102381204", "23苏城投MTN004", Decimal("100"))],
        ),
    )

    details = build_display_details(result)

    assert {"label": "具体原因", "value": "①债券缺失：23苏城投MTN004；原因：资负数据子系统-债务证券明细表无数据"} in details[0]["rows"]
    assert [detail["title"] for detail in details] == ["最终判断结果", "具体差异明细", "资产缺失细分"]
    assert details[2]["table"]["headers"] == ["序号", "资产类型", "资产名称", "FA科目编码", "科目尾段", "FA估值金额", "核查表", "核查结果", "关键字段", "AM标的代码", "AM合同代码", "原因"]
    assert details[2]["table"]["rows"][0] == [
        "①",
        "债券",
        "23苏城投MTN004",
        "1501.01.01.01.102381204",
        "102381204",
        "100",
        "currency_report_24.currency_detail_project_2_1_4",
        "资负数据子系统-债务证券明细表无数据",
        "",
        "",
        "",
        "资负数据子系统-债务证券明细表无数据",
    ]


def test_display_details_show_asset_duplicate_refinement_table_and_specific_reason():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("1100"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("100"),
        direction="资产大于负债及权益",
        difference_reason="资产重复",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="asset_gap",
                data={
                    "reason": "资产重复",
                    "valuation_asset_total": "1000",
                    "asset_gap": "100",
                    "match_type": "single",
                    "match_total": "100",
                },
            ),
            DifferenceDetail(
                kind="asset_duplicate_refinement",
                data={
                    "specific_reason": "①私募基金重复：某私募产品一号；原因：该资产在证券信息表中为私募产品但在AM中不为私募产品",
                    "rows": [
                        {
                            "index": "①",
                            "asset_type": "私募基金",
                            "asset_name": "某私募产品一号",
                            "fa_account_code": "1101.05.06.01.SM001",
                            "account_tail": "SM001",
                            "fa_market_value": "100",
                            "check_table": "am_pactasset_dws",
                            "check_result": "该资产在证券信息表中为私募产品但在AM中不为私募产品",
                            "key_field": "c_spv_type/c_assettype",
                            "am_spv_type": "10",
                            "am_asset_type": "31",
                            "reason": "该资产在证券信息表中为私募产品但在AM中不为私募产品",
                        }
                    ],
                },
            ),
        ],
        valuation_match=ValuationMatch(
            match_type="single",
            rows=[ValuationRow("1101.05.06.01.SM001", "某私募产品一号", Decimal("100"))],
        ),
    )

    details = build_display_details(result)

    assert {"label": "具体原因", "value": "①私募基金重复：某私募产品一号；原因：该资产在证券信息表中为私募产品但在AM中不为私募产品"} in details[0]["rows"]
    assert [detail["title"] for detail in details] == ["最终判断结果", "具体差异明细", "资产重复细分"]
    assert details[2]["table"]["headers"] == ["序号", "资产类型", "资产名称", "FA科目编码", "科目尾段", "FA估值金额", "核查表", "核查结果", "关键字段", "AM SPV类型", "AM资产类型", "原因"]
    assert details[2]["table"]["rows"][0] == [
        "①",
        "私募基金",
        "某私募产品一号",
        "1101.05.06.01.SM001",
        "SM001",
        "100",
        "am_pactasset_dws",
        "该资产在证券信息表中为私募产品但在AM中不为私募产品",
        "c_spv_type/c_assettype",
        "10",
        "31",
        "该资产在证券信息表中为私募产品但在AM中不为私募产品",
    ]


def test_display_details_show_asset_difference_refinement_table_and_specific_reason():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("900"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("-100"),
        direction="资产小于负债及权益",
        difference_reason="资产差异",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="asset_gap",
                data={
                    "reason": "资产缺失",
                    "valuation_asset_total": "1000",
                    "asset_gap": "100",
                    "match_type": "none",
                    "match_total": "0",
                },
            ),
            DifferenceDetail(
                kind="asset_difference_refinement",
                data={
                    "specific_reason": "①流动资金贷款一号贷款合同：FA科目余额与AM投融资余额有差异，差异值-100",
                    "market_total": "300",
                    "project_invest_total": "200",
                    "difference_total": "-100",
                    "asset_total_gap": "-100",
                    "basis": "贷款、财产权合同分别比较 AM投融资余额 - FA科目余额；逆回购比较存续回购业务表金额 - FA科目余额。",
                    "rows": [
                        {
                            "index": "①",
                            "asset_type": "贷款合同",
                            "asset_name": "流动资金贷款一号",
                            "account_code": "1501.04.05.01.DK20260531001",
                            "pact_id": "DK20260531001",
                            "market_value": "300",
                            "project_invest_balance": "200",
                            "difference": "-100",
                            "check_table": "dm.am_projinvest_zgxg_dm",
                            "reason": "流动资金贷款一号贷款合同：FA科目余额与AM投融资余额有差异，差异值-100",
                        }
                    ],
                },
            ),
        ],
        valuation_match=ValuationMatch(match_type="none", rows=[]),
    )

    details = build_display_details(result)

    assert {"label": "具体原因", "value": "①流动资金贷款一号贷款合同：FA科目余额与AM投融资余额有差异，差异值-100"} in details[0]["rows"]
    assert {"label": "资产合计差额(a0001-0004)", "value": "-100"} in details[0]["rows"]
    assert {"label": "资产差异金额合计", "value": "-100"} in details[0]["rows"]
    assert [detail["title"] for detail in details] == ["最终判断结果", "资产差异细分"]
    assert details[1]["table"]["headers"] == ["序号", "资产类型", "资产名称", "FA科目编码", "合同代码/证券代码", "FA科目余额", "DM证券余额/AM投融资余额/存续回购业务表金额", "差异值", "核查表", "原因"]
    assert details[1]["table"]["rows"][0] == [
        "①",
        "贷款合同",
        "流动资金贷款一号",
        "1501.04.05.01.DK20260531001",
        "DK20260531001",
        "300",
        "200",
        "-100",
        "dm.am_projinvest_zgxg_dm",
        "流动资金贷款一号贷款合同：FA科目余额与AM投融资余额有差异，差异值-100",
    ]


def test_display_details_show_bond_security_code_in_asset_difference_refinement():
    result = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("1015"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("15"),
        direction="资产大于负债及权益",
        difference_reason="资产差异",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="asset_difference_refinement",
                data={
                    "specific_reason": "①23苏城投MTN002债券：FA债券本金科目余额与DM证券余额有差异，债券代码ZQ002，差异值-15",
                    "market_total": "100",
                    "project_invest_total": "85",
                    "difference_total": "-15",
                    "asset_total_gap": "-15",
                    "basis": "债券比较 DM证券余额 - FA债券本金科目余额。",
                    "rows": [
                        {
                            "index": "①",
                            "asset_type": "债券",
                            "asset_name": "23苏城投MTN002",
                            "account_code": "1101.02.01.01.ZQ002",
                            "pact_id": "",
                            "security_code": "ZQ002",
                            "market_value": "100",
                            "project_invest_balance": "85",
                            "difference": "-15",
                            "check_table": "dm.fa_security_balance_zgxg_dm",
                            "reason": "23苏城投MTN002债券：FA债券本金科目余额与DM证券余额有差异，债券代码ZQ002，差异值-15",
                        }
                    ],
                },
            ),
        ],
        valuation_match=ValuationMatch(match_type="none", rows=[]),
    )

    details = build_display_details(result)

    assert {"label": "资产差异DM证券余额/AM投融资余额/存续回购业务表金额合计", "value": "85"} in details[0]["rows"]
    assert details[1]["table"]["rows"][0] == [
        "①",
        "债券",
        "23苏城投MTN002",
        "1101.02.01.01.ZQ002",
        "ZQ002",
        "100",
        "85",
        "-15",
        "dm.fa_security_balance_zgxg_dm",
        "23苏城投MTN002债券：FA债券本金科目余额与DM证券余额有差异，债券代码ZQ002，差异值-15",
    ]


def test_display_details_show_combined_asset_and_liability_difference_reason():
    result = ReconcileResult(
        project_code="P9",
        project_name="Project 9",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("850"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("150"),
        direction="资产大于负债及权益",
        difference_reason="资产差异 + 负债及权益科目差异",
        match_status="已解释",
        valuation_asset_total=Decimal("900"),
        details=[
            DifferenceDetail(
                kind="asset_gap",
                data={
                    "reason": "资产重复",
                    "zf_asset_total": "1000",
                    "valuation_asset_total": "900",
                    "asset_gap": "100",
                    "match_type": "none",
                    "match_total": "0",
                    "account_scope": "1开头末级科目",
                },
            ),
            DifferenceDetail(
                kind="asset_difference_refinement",
                data={
                    "specific_reason": "①逆回购：FA科目余额与存续回购业务表逆回购金额有差异，差异值100",
                    "market_total": "300",
                    "project_invest_total": "400",
                    "difference_total": "100",
                    "asset_total_gap": "100",
                    "remaining_difference": "50",
                    "basis": "差异合计与 a0001-估值表0004 相等。",
                    "rows": [],
                },
            ),
            DifferenceDetail(
                kind="liability_equity",
                data={
                    "match_type": "none",
                    "match_total": "0",
                    "match_target": "50",
                    "account_scope": "非1开头科目",
                    "specific_reason": (
                        "①逆回购：FA科目余额与存续回购业务表逆回购金额有差异，差异值100\n"
                        "②正回购：FA科目余额与存续回购业务表正回购金额有差异，差异值50"
                    ),
                    "rows": [],
                },
            ),
        ],
    )

    details = build_display_details(result)

    assert {"label": "差异类型", "value": "资产差异 + 负债及权益科目差异"} in details[0]["rows"]
    assert {
        "label": "具体原因",
        "value": (
            "①逆回购：FA科目余额与存续回购业务表逆回购金额有差异，差异值100\n"
            "②正回购：FA科目余额与存续回购业务表正回购金额有差异，差异值50"
        ),
    } in details[0]["rows"]
    assert {"label": "资产端解释后剩余差额", "value": "50"} in details[0]["rows"]


def test_display_details_show_am_missing_and_project_invest_sections():
    am_missing = ReconcileResult(
        project_code="P2",
        project_name="Project 2",
        asset_total=Decimal("900"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("-100"),
        direction="资产小于负债及权益",
        difference_reason="资产缺失",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="am_missing",
                data={
                    "fa_account_code": "1101.05.03.01.0002",
                    "fa_account_name": "Asset A",
                    "fa_tail_code": "0002",
                    "fa_market_value": "100",
                    "expected_account_level": "1101.05.03.01",
                    "specific_reason": "AM标的缺失",
                },
            )
        ],
    )
    invest_zero = ReconcileResult(
        project_code="P3",
        project_name="Project 3",
        asset_total=Decimal("900"),
        liability_equity_total=Decimal("1000"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("-100"),
        direction="资产小于负债及权益",
        difference_reason="资产缺失",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="project_invest_balance",
                data={
                    "fa_account_code": "1101.05.03.01.0002",
                    "fa_account_name": "Asset A",
                    "fa_tail_code": "0002",
                    "fa_market_value": "100",
                    "am_asset_name": "Asset A",
                    "am_stock_code": "0002",
                    "pact_id": "PACT1",
                    "project_invest_balance": "0",
                    "specific_reason": "合同投融资余额为0但FA科目余额不为0",
                },
            )
        ],
    )

    am_missing_details = build_display_details(am_missing)
    invest_details = build_display_details(invest_zero)

    assert [detail["title"] for detail in am_missing_details] == ["最终判断结果", "AM标的缺失"]
    assert {"label": "具体原因", "value": "AM标的缺失"} in am_missing_details[0]["rows"]
    assert {"label": "需复核四级科目", "value": "1101.05.03.01"} in am_missing_details[1]["rows"]
    assert [detail["title"] for detail in invest_details] == ["最终判断结果", "合同投融资余额核对"]
    assert {"label": "具体原因", "value": "合同投融资余额为0但FA科目余额不为0"} in invest_details[0]["rows"]
    assert {"label": "AM 合同代码", "value": "PACT1"} in invest_details[1]["rows"]
    assert {"label": "合同投融资余额", "value": "0"} in invest_details[1]["rows"]


def test_display_details_show_ta_received_trust_sections():
    total_mismatch = ReconcileResult(
        project_code="P4",
        project_name="Project 4",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("900"),
        received_trust_balance=Decimal("400"),
        difference=Decimal("100"),
        direction="资产大于负债及权益",
        difference_reason="实收本金差异",
        match_status="已解释",
        details=[
            DifferenceDetail(kind="received_trust", data={"c1000_balance": "400", "fa_4001_balance": "500", "received_trust_difference": "100"}),
            DifferenceDetail(kind="ta_total_mismatch", data={"dm_total": "480", "dws_total": "500", "difference": "-20", "specific_reason": "①实收本金差异：FA 4001与c1000存在差异，差异值100；原因：DM表TA份额余额错误"}),
        ],
    )
    blank_client_type = ReconcileResult(
        project_code="P5",
        project_name="Project 5",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("900"),
        received_trust_balance=Decimal("400"),
        difference=Decimal("100"),
        direction="资产大于负债及权益",
        difference_reason="实收本金差异",
        match_status="已解释",
        details=[
            DifferenceDetail(kind="received_trust", data={"c1000_balance": "400", "fa_4001_balance": "500", "received_trust_difference": "100"}),
            DifferenceDetail(
                kind="ta_blank_client_type",
                data={
                    "blank_client_type_total": "100",
                    "specific_reason": "①实收本金差异：FA 4001与c1000存在差异，差异值100；原因：dm.ta_pact_survamt_day_zgxg_dm表中客户类型为空导致实收信托有误",
                    "rows": [
                        {
                            "pact_id": "PACT1",
                            "client_name": "客户A",
                            "client_kind": "4",
                            "client_kind_index": "",
                            "spv_type": "SPV",
                            "ht_income": "30",
                            "share_amount": "70",
                            "amount": "100",
                        }
                    ],
                },
            ),
        ],
    )

    total_details = build_display_details(total_mismatch)
    blank_details = build_display_details(blank_client_type)

    assert [detail["title"] for detail in total_details] == ["最终判断结果", "TA汇总核对"]
    assert {"label": "具体原因", "value": "①实收本金差异：FA 4001与c1000存在差异，差异值100；原因：DM表TA份额余额错误"} in total_details[0]["rows"]
    assert {"label": "DM TA 份额余额+待结转收益", "value": "480"} in total_details[1]["rows"]
    assert [detail["title"] for detail in blank_details] == ["最终判断结果", "DM TA客户类型为空"]
    assert {"label": "具体原因", "value": "①实收本金差异：FA 4001与c1000存在差异，差异值100；原因：dm.ta_pact_survamt_day_zgxg_dm表中客户类型为空导致实收信托有误"} in blank_details[0]["rows"]
    assert blank_details[1]["table"]["headers"] == ["合同编号", "客户名称", "客户类型", "客户类型明细", "SPV类型", "待结转收益", "份额余额", "合计"]
    assert blank_details[1]["table"]["rows"][0][0] == "PACT1"


def test_display_details_show_received_trust_and_liability_equity_refinement_tables():
    result = ReconcileResult(
        project_code="P6",
        project_name="Project 6",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("950"),
        received_trust_balance=Decimal("400"),
        difference=Decimal("50"),
        direction="资产大于负债及权益",
        difference_reason="负债及权益科目重复",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="received_trust",
                data={
                    "c1000_balance": "400",
                    "fa_4001_balance": "500",
                    "received_trust_difference": "100",
                    "specific_reason": "①实收本金差异：FA 4001与c1000存在差异，差异值100",
                    "refinement_rows": [
                        {
                            "index": "①",
                            "type": "实收本金差异",
                            "fa_4001_balance": "500",
                            "c1000_balance": "400",
                            "difference": "100",
                            "check_table": "fa_accountbalance_dws/余额表c1000",
                            "check_result": "FA 4001与c1000存在差异",
                            "reason": "",
                        }
                    ],
                },
            ),
            DifferenceDetail(
                kind="liability_equity",
                data={
                    "specific_reason": "①实收本金差异：FA 4001与c1000存在差异，差异值100\n②负债及权益科目重复：其他收益",
                    "rows": [
                        {
                            "index": "②",
                            "account_type": "负债及权益科目",
                            "account_name": "其他收益",
                            "account_code": "4002",
                            "account_tail": "4002",
                            "market_value": "50",
                            "direction": "重复",
                            "check_result": "命中",
                            "reason": "",
                        }
                    ],
                    "match_type": "single",
                    "match_total": "50",
                },
            ),
        ],
        valuation_match=ValuationMatch(match_type="single", rows=[ValuationRow("4002", "其他收益", Decimal("50"))]),
    )

    details = build_display_details(result)

    assert [detail["title"] for detail in details] == ["最终判断结果", "具体差异明细", "实收本金细分", "负债及权益科目细分"]
    assert details[2]["table"]["headers"] == ["序号", "类型", "FA 4001科目余额", "c1000实收本金余额", "差异值", "核查表", "核查结果", "原因"]
    assert details[2]["table"]["rows"][0] == ["①", "实收本金差异", "500", "400", "100", "fa_accountbalance_dws/余额表c1000", "FA 4001与c1000存在差异", ""]
    assert details[3]["table"]["headers"] == ["序号", "科目类型", "科目名称", "FA科目编码", "科目尾段", "FA科目金额", "存续回购业务表金额", "差异方向", "核查结果", "原因"]
    assert details[3]["table"]["rows"][0] == ["②", "负债及权益科目", "其他收益", "4002", "4002", "50", "", "重复", "命中", ""]


def test_display_details_show_received_trust_residual_liability_equity_rows():
    result = ReconcileResult(
        project_code="P6",
        project_name="Project 6",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("950"),
        received_trust_balance=Decimal("400"),
        difference=Decimal("50"),
        direction="资产大于负债及权益",
        difference_reason="负债及权益科目重复",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="received_trust",
                data={
                    "c1000_balance": "400",
                    "fa_4001_balance": "500",
                    "received_trust_difference": "100",
                },
            ),
            DifferenceDetail(
                kind="liability_equity",
                data={
                    "match_type": "single",
                    "match_total": "50",
                    "match_target": "50",
                    "account_scope": "非1开头科目",
                    "main_difference": "50",
                    "received_trust_difference": "100",
                    "residual_difference": "-50",
                    "specific_reason": "①实收本金差异：FA 4001与c1000存在差异，差异值100\n②负债及权益科目重复：其他收益",
                },
            ),
        ],
    )

    details = build_display_details(result)

    assert details[0]["rows"][0] == {"label": "差异类型", "value": "负债及权益科目重复"}
    assert {"label": "具体原因", "value": "①实收本金差异：FA 4001与c1000存在差异，差异值100\n②负债及权益科目重复：其他收益"} in details[0]["rows"]
    assert {"label": "4001-c1000 差异", "value": "100"} in details[0]["rows"]
    assert {"label": "主差异", "value": "50"} in details[0]["rows"]
    assert {"label": "剩余差额", "value": "-50"} in details[0]["rows"]
    assert {"label": "核对范围", "value": "非1开头科目"} in details[0]["rows"]
    assert {"label": "命中金额", "value": "50"} in details[0]["rows"]


def test_display_details_show_positive_repo_specific_reason():
    result = ReconcileResult(
        project_code="P7",
        project_name="Project 7",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("950"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("50"),
        direction="资产大于负债及权益",
        difference_reason="负债及权益科目缺失",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="liability_equity",
                data={
                    "match_type": "single",
                    "match_total": "50",
                    "match_target": "50",
                    "account_scope": "非1开头科目",
                    "specific_reason": "①正回购缺失：卖出回购金融资产款；原因：正回购差异",
                },
            ),
        ],
    )

    details = build_display_details(result)

    assert details[0]["rows"][0] == {"label": "差异类型", "value": "负债及权益科目缺失"}
    assert {"label": "具体原因", "value": "①正回购缺失：卖出回购金融资产款；原因：正回购差异"} in details[0]["rows"]


def test_display_details_show_positive_repo_business_amount_difference_row():
    result = ReconcileResult(
        project_code="P8",
        project_name="Project 8",
        asset_total=Decimal("1000"),
        liability_equity_total=Decimal("950"),
        received_trust_balance=Decimal("500"),
        difference=Decimal("50"),
        direction="资产大于负债及权益",
        difference_reason="负债及权益科目差异",
        match_status="已解释",
        details=[
            DifferenceDetail(
                kind="liability_equity",
                data={
                    "match_type": "none",
                    "match_total": "0",
                    "match_target": "50",
                    "account_scope": "非1开头科目",
                    "specific_reason": "①正回购：FA科目余额与存续回购业务表正回购金额有差异，差异值50",
                    "rows": [
                        {
                            "index": "①",
                            "account_type": "正回购",
                            "account_name": "正回购",
                            "account_code": "2111.12.34.01",
                            "account_tail": "",
                            "market_value": "200",
                            "business_amount": "150",
                            "direction": "差异",
                            "check_result": "金额差异",
                            "reason": "正回购：FA科目余额与存续回购业务表正回购金额有差异，差异值50",
                        }
                    ],
                },
            ),
        ],
    )

    details = build_display_details(result)

    assert [detail["title"] for detail in details] == ["最终判断结果", "负债及权益科目细分"]
    assert details[1]["table"]["headers"] == ["序号", "科目类型", "科目名称", "FA科目编码", "科目尾段", "FA科目金额", "存续回购业务表金额", "差异方向", "核查结果", "原因"]
    assert details[1]["table"]["rows"][0] == [
        "①",
        "正回购",
        "正回购",
        "2111.12.34.01",
        "",
        "200",
        "150",
        "差异",
        "金额差异",
        "正回购：FA科目余额与存续回购业务表正回购金额有差异，差异值50",
    ]


def test_post_test_connection_returns_source_status(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle("POST", "/api/test-connection", {})

    assert status == 200
    assert payload["dws"]["ok"] is True
    assert payload["business"]["ok"] is True


def test_connection_error_message_replaces_mojibake_from_database_driver():
    raw = "connection failed: server at 127.0.0.1, port 5432 failed: \ufffd\ufffd\ufffd\ufffd\ufffd\ufffd: \ufffd\ufffd root Password \ufffd\ufffd\ufffd\ufffd"

    message = _connection_error_message(raw)

    assert "\ufffd" not in message
    assert message == "\u8fde\u63a5\u5931\u8d25\uff1a\u6570\u636e\u5e93\u8fd4\u56de\u9519\u8bef\u4fe1\u606f\u7f16\u7801\u5f02\u5e38\uff0c\u8bf7\u68c0\u67e5\u7528\u6237\u540d\u3001\u5bc6\u7801\u3001\u5730\u5740\u548c\u7aef\u53e3"


def test_database_error_messages_explain_common_failure_reasons():
    unknown_db = _runtime_error_message('(1049, "Unknown database \'ass_man_reg\'")')
    assert "数据库不存在：ass_man_reg" in unknown_db
    assert "ass_man_reg.ex_pledge_back" in unknown_db
    assert "连接到的 MySQL 实例包含 ass_man_reg 库" in unknown_db
    assert "原始错误" in unknown_db

    missing_table = _runtime_error_message('(1146, "Table \'reg-report-analysis.zf_detail_2024\' doesn\'t exist")')
    assert "数据表不存在：reg-report-analysis.zf_detail_2024" in missing_table
    assert "所选数据源连接到了正确实例" in missing_table

    missing_column = _runtime_error_message('(1054, "Unknown column \'a0001\' in \'field list\'")')
    assert "字段不存在：a0001" in missing_column
    assert "表结构与当前规则版本一致" in missing_column

    denied = _connection_error_message("(1045, \"Access denied for user 'auto'@'localhost' (using password: YES)\")")
    assert "数据库权限不足：用户 auto" in denied
    assert "用户名、密码" in denied

    refused = _runtime_error_message("Can't connect to MySQL server on '127.0.0.1' ([WinError 10061])")
    assert "数据库连接失败" in refused
    assert "地址或端口" in refused


def test_runtime_error_message_keeps_postgres_missing_table_details_with_sql_context():
    message = _runtime_error_message('关系 "test.zf_detail_2024" 不存在\nLINE 8: FROM "test"."zf_detail_2024"')

    assert "数据表不存在：test.zf_detail_2024" in message
    assert "表字段配置" in message
    assert "操作失败" not in message


def test_get_connection_status_returns_source_status(tmp_path):
    router = ApiRouter(
        config_path=tmp_path / "config.json",
        runner_factory=lambda config: FakeRunner(),
        connection_tester=FakeConnectionTester(),
    )

    status, payload = router.handle("GET", "/api/connection-status", None)

    assert status == 200
    assert payload["dws"]["ok"] is True
    assert payload["business"]["ok"] is True

def test_previous_month_end_handles_year_boundary():
    assert previous_month_end("2026-05-28") == "2026-04-30"
    assert previous_month_end("2026-01-10") == "2025-12-31"
