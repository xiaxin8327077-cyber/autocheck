from __future__ import annotations

import csv
import io
import re
import tempfile
import zipfile
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Iterator, Sequence


SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
SUPPORTED_ARCHIVE_EXTENSIONS = {".zip", ".rar", ".7z"}
SUPPORTED_UPLOAD_EXTENSIONS = SUPPORTED_EXTENSIONS | SUPPORTED_ARCHIVE_EXTENSIONS
_IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
_PREVIEW_ROW_LIMIT = 100
_HEADER_SCAN_LIMIT = 100


def _normalize_cell(value: Any) -> Any:
    """Convert empty strings to None so database can treat them as NULL."""
    if value == "" or value is None:
        return None
    return value


@dataclass(frozen=True)
class ImportFileInspection:
    name: str
    file_type: str
    columns: list[str]
    header_row: int = 1
    data_start_row: int = 2
    detection: str = "default"
    matched_columns: list[str] | None = None


@dataclass(frozen=True)
class ZipInspection:
    files: list[ImportFileInspection]
    columns: list[str]


@dataclass(frozen=True)
class TableColumn:
    name: str
    comment: str = ""


@dataclass(frozen=True)
class HeaderLayout:
    headers: list[str]
    indexes: list[int]


@dataclass(frozen=True)
class ColumnMapping:
    source_column: str
    target_column: str
    target_comment: str = ""
    score: float = 0


@dataclass(frozen=True)
class TableRef:
    parts: tuple[str, ...]

    def quoted(self, db_type: str) -> str:
        if db_type == "postgresql":
            return ".".join('"' + part.replace('"', '""') + '"' for part in self.parts)
        if db_type == "mysql":
            return ".".join("`" + part.replace("`", "``") + "`" for part in self.parts)
        raise ValueError(f"Unsupported database type: {db_type}")


def parse_table_ref(value: str) -> TableRef:
    parts = tuple(part.strip() for part in str(value or "").split("."))
    if not parts or any(not part for part in parts):
        raise ValueError("target table is required")
    if len(parts) > 3:
        raise ValueError("target table supports at most three identifier parts")
    for part in parts:
        if not _IDENTIFIER_RE.match(part):
            raise ValueError(f"unsafe table identifier: {part}")
    return TableRef(parts=parts)


def inspect_zip_headers(zip_path: str | Path, *, max_member_bytes: int | None = None) -> ZipInspection:
    return _inspect_payloads(_iter_zip_payloads(zip_path, max_member_bytes=max_member_bytes), empty_message="zip file does not contain csv, xlsx, or xls files")


def inspect_import_upload(upload_path: str | Path, *, display_name: str | None = None, max_member_bytes: int | None = None) -> ZipInspection:
    return _inspect_payloads(_iter_upload_payloads(upload_path, display_name=display_name, max_member_bytes=max_member_bytes), empty_message="upload does not contain csv, xlsx, or xls files")


def inspect_import_upload_with_target_columns(
    upload_path: str | Path | Sequence[str | Path],
    target_columns: Sequence[TableColumn],
    *,
    max_member_bytes: int | None = None,
) -> ZipInspection:
    payloads = _iter_import_payloads_with_limits(upload_path, max_member_bytes=max_member_bytes)
    return _inspect_payloads(
        payloads,
        empty_message="upload does not contain csv, xlsx, or xls files",
        target_columns=target_columns,
    )


def _inspect_payloads(
    payloads: Iterable[tuple[str, str, bytes]],
    *,
    empty_message: str,
    target_columns: Sequence[TableColumn] = (),
) -> ZipInspection:
    files: list[ImportFileInspection] = []
    columns: list[str] = []
    seen = set()
    for display_name, ext, payload in payloads:
        file_inspection = _inspect_file_payload(display_name, ext, payload, target_columns=target_columns)
        files.append(file_inspection)
        for column in file_inspection.columns:
            if column not in seen:
                seen.add(column)
                columns.append(column)
    if not files:
        raise ValueError(empty_message)
    return ZipInspection(files=files, columns=columns)


def iter_projected_rows(
    zip_path: str | Path | Sequence[str | Path],
    *,
    columns: Sequence[str],
    drop_columns: Iterable[str] = (),
    column_order: Sequence[str] | None = None,
    file_layouts: Sequence[ImportFileInspection] = (),
) -> Iterator[tuple[Any, ...]]:
    final_columns = projected_columns(columns, drop_columns=drop_columns, column_order=column_order)
    layout_by_name = _layout_by_name(file_layouts)
    for display_name, ext, payload in _iter_import_payloads(zip_path):
        layout = layout_by_name.get(display_name)
        for row in _iter_rows_from_bytes(
            display_name,
            ext,
            payload,
            header_row=layout.header_row if layout else 1,
            data_start_row=layout.data_start_row if layout else None,
        ):
            yield tuple(_normalize_cell(row.get(column, "")) for column in final_columns)


def iter_mapped_rows(
    zip_path: str | Path | Sequence[str | Path],
    mappings: Sequence[ColumnMapping],
    *,
    file_layouts: Sequence[ImportFileInspection] = (),
) -> Iterator[tuple[Any, ...]]:
    active = [mapping for mapping in mappings if mapping.target_column]
    if not active:
        raise ValueError("at least one mapped column is required")
    layout_by_name = _layout_by_name(file_layouts)
    for display_name, ext, payload in _iter_import_payloads(zip_path):
        layout = layout_by_name.get(display_name)
        for row in _iter_rows_from_bytes(
            display_name,
            ext,
            payload,
            header_row=layout.header_row if layout else 1,
            data_start_row=layout.data_start_row if layout else None,
        ):
            yield tuple(_normalize_cell(row.get(mapping.source_column, "")) for mapping in active)


def build_column_mappings(source_columns: Sequence[str], target_columns: Sequence[TableColumn]) -> list[ColumnMapping]:
    mappings: list[ColumnMapping] = []
    used_targets: set[str] = set()
    for source in source_columns:
        best_column: TableColumn | None = None
        best_score = 0.0
        for target in target_columns:
            if target.name in used_targets:
                continue
            score = _semantic_score(source, target)
            if score > best_score:
                best_score = score
                best_column = target
        if best_column is not None and best_score >= 0.55:
            used_targets.add(best_column.name)
            mappings.append(ColumnMapping(source_column=source, target_column=best_column.name, target_comment=best_column.comment, score=round(best_score, 4)))
        else:
            mappings.append(ColumnMapping(source_column=source, target_column="", target_comment="", score=0))
    return mappings


def projected_columns(
    columns: Sequence[str],
    *,
    drop_columns: Iterable[str] = (),
    column_order: Sequence[str] | None = None,
) -> list[str]:
    drop_set = set(drop_columns or ())
    available = [column for column in columns if column and column not in drop_set]
    ordered: list[str] = []
    for column in column_order or ():
        if column in available and column not in ordered:
            ordered.append(column)
    ordered.extend(column for column in available if column not in ordered)
    if not ordered:
        raise ValueError("at least one column must be selected")
    return ordered


def mapped_target_columns(mappings: Sequence[ColumnMapping]) -> list[str]:
    columns = [mapping.target_column for mapping in mappings if mapping.target_column]
    if not columns:
        raise ValueError("at least one mapped column is required")
    return columns


def _suffix(name: str) -> str:
    return Path(name).suffix.lower()


def _display_zip_name(info: zipfile.ZipInfo) -> str:
    if info.flag_bits & 0x800:
        return info.filename
    try:
        return info.filename.encode("cp437").decode("gbk")
    except UnicodeError:
        return info.filename


def _read_header(archive: zipfile.ZipFile, name: str, ext: str) -> list[str]:
    rows = _iter_rows_from_bytes(name, ext, archive.read(name), limit=0, include_header=True)
    try:
        return next(rows)
    except StopIteration as exc:
        raise ValueError(f"{name} does not contain a header row") from exc


def _read_header_from_bytes(name: str, ext: str, payload: bytes) -> list[str]:
    rows = _iter_rows_from_bytes(name, ext, payload, limit=0, include_header=True)
    try:
        return next(rows)
    except StopIteration as exc:
        raise ValueError(f"{name} does not contain a header row") from exc


def _inspect_file_payload(
    name: str,
    ext: str,
    payload: bytes,
    *,
    target_columns: Sequence[TableColumn] = (),
) -> ImportFileInspection:
    if target_columns:
        return _detect_file_layout(name, ext, payload, target_columns)
    return _detect_generic_file_layout(name, ext, payload)


def _detect_generic_file_layout(name: str, ext: str, payload: bytes) -> ImportFileInspection:
    rows = _read_preview_rows(name, ext, payload, limit=_PREVIEW_ROW_LIMIT)
    if not rows:
        raise ValueError(f"{name} does not contain a header row")
    layout = _best_generic_layout(name, ext, rows)
    if layout is not None:
        return layout
    raise ValueError(f"{name} does not contain a valid header row")


def _detect_file_layout(name: str, ext: str, payload: bytes, target_columns: Sequence[TableColumn]) -> ImportFileInspection:
    rows = _read_preview_rows(name, ext, payload, limit=_PREVIEW_ROW_LIMIT)
    if not rows:
        raise ValueError(f"{name} does not contain a header row")
    default = _candidate_layout(name, ext, rows, 0, target_columns, detection="default")
    if default is not None and _is_reliable_default_header(default):
        return default
    scored: list[tuple[float, ImportFileInspection]] = []
    for index in range(min(len(rows), _HEADER_SCAN_LIMIT)):
        candidate = _candidate_layout(name, ext, rows, index, target_columns, detection="smart")
        if candidate is None:
            continue
        match_score = _header_match_score(candidate.columns, target_columns)
        if match_score["matched_count"] <= 0:
            continue
        nonempty_ratio = _row_nonempty_ratio(rows[index])
        data_score = _data_region_score(rows[index + 1:index + 6])
        title_bonus = min(index, 6) * 0.08
        score = float(match_score["matched_count"]) * 10 + float(match_score["score_total"]) + nonempty_ratio + data_score + title_bonus
        scored.append((score, candidate))
    if scored:
        scored.sort(key=lambda item: item[0], reverse=True)
        return scored[0][1]
    generic = _best_generic_layout(name, ext, rows)
    if generic is not None:
        return generic
    raise ValueError(f"{name} does not contain a valid header row")


def _best_generic_layout(name: str, ext: str, rows: Sequence[Sequence[Any]]) -> ImportFileInspection | None:
    default = _candidate_generic_layout(name, ext, rows, 0, detection="default")
    if default is not None and len(default.columns) >= 2 and _row_nonempty_ratio(rows[0]) >= 0.5:
        return default
    for index in range(min(len(rows), _HEADER_SCAN_LIMIT)):
        candidate = _candidate_generic_layout(name, ext, rows, index, detection="smart")
        if candidate is None or len(candidate.columns) < 2:
            continue
        nonempty_ratio = _row_nonempty_ratio(rows[index])
        data_score = _data_region_score(rows[index + 1:index + 6])
        if nonempty_ratio >= 0.5 and data_score > 0:
            return candidate
    if default is not None:
        return default
    return None


def _candidate_layout(
    name: str,
    ext: str,
    rows: Sequence[Sequence[Any]],
    header_index: int,
    target_columns: Sequence[TableColumn],
    *,
    detection: str,
) -> ImportFileInspection | None:
    try:
        columns = _clean_headers(rows[header_index])
    except ValueError:
        return None
    match_score = _header_match_score(columns, target_columns)
    data_start_index = _first_data_row_index(rows, header_index + 1)
    return ImportFileInspection(
        name=name,
        file_type=ext.lstrip("."),
        columns=columns,
        header_row=header_index + 1,
        data_start_row=data_start_index + 1,
        detection=detection,
        matched_columns=list(match_score["matched_columns"]),
    )


def _candidate_generic_layout(
    name: str,
    ext: str,
    rows: Sequence[Sequence[Any]],
    header_index: int,
    *,
    detection: str,
) -> ImportFileInspection | None:
    try:
        columns = _clean_headers(rows[header_index])
    except ValueError:
        return None
    data_start_index = _first_data_row_index(rows, header_index + 1)
    return ImportFileInspection(
        name=name,
        file_type=ext.lstrip("."),
        columns=columns,
        header_row=header_index + 1,
        data_start_row=data_start_index + 1,
        detection=detection,
        matched_columns=[],
    )


def _is_reliable_default_header(layout: ImportFileInspection) -> bool:
    matched = layout.matched_columns or []
    if any(kind == "exact" for kind in matched):
        return True
    return len(matched) >= 2


def _header_match_score(columns: Sequence[str], target_columns: Sequence[TableColumn]) -> dict[str, Any]:
    matched_columns: list[str] = []
    score_total = 0.0
    for source in columns:
        best_score = 0.0
        best_kind = ""
        for target in target_columns:
            source_norm = _normalize_label(source)
            names = [_normalize_label(target.name), _normalize_label(target.comment)]
            if source_norm and source_norm in names:
                best_score = 1.0
                best_kind = "exact"
                break
            semantic = _semantic_score(source, target)
            if semantic > best_score:
                best_score = semantic
                best_kind = "semantic"
        if best_score >= 0.72:
            matched_columns.append(best_kind)
            score_total += best_score
    return {"matched_count": len(matched_columns), "matched_columns": matched_columns, "score_total": score_total}


def _row_nonempty_ratio(row: Sequence[Any]) -> float:
    values = list(row)
    if not values:
        return 0
    return sum(1 for value in values if str(value or "").strip()) / len(values)


def _data_region_score(rows: Sequence[Sequence[Any]]) -> float:
    nonempty_rows = [row for row in rows if any(str(value or "").strip() for value in row)]
    if not nonempty_rows:
        return 0
    return min(len(nonempty_rows), 5) * 0.2


def _first_data_row_index(rows: Sequence[Sequence[Any]], start_index: int) -> int:
    for index in range(start_index, len(rows)):
        if any(str(value or "").strip() for value in rows[index]):
            return index
    return start_index


def _layout_by_name(file_layouts: Sequence[ImportFileInspection]) -> dict[str, ImportFileInspection]:
    return {layout.name: layout for layout in file_layouts}


def _iter_import_payloads(upload_path: str | Path | Sequence[str | Path]) -> Iterator[tuple[str, str, bytes]]:
    if isinstance(upload_path, (str, Path)):
        yield from _iter_upload_payloads(upload_path)
        return
    for path in upload_path:
        yield from _iter_upload_payloads(path)


def _iter_import_payloads_with_limits(upload_path: str | Path | Sequence[str | Path], *, max_member_bytes: int | None = None) -> Iterator[tuple[str, str, bytes]]:
    if isinstance(upload_path, (str, Path)):
        yield from _iter_upload_payloads(upload_path, max_member_bytes=max_member_bytes)
        return
    for path in upload_path:
        yield from _iter_upload_payloads(path, max_member_bytes=max_member_bytes)


def _iter_rows_from_bytes(
    name: str,
    ext: str,
    payload: bytes,
    *,
    limit: int | None = None,
    include_header: bool = False,
    header_row: int = 1,
    data_start_row: int | None = None,
) -> Iterator[dict[str, Any] | list[str]]:
    if ext == ".csv":
        yield from _iter_csv_rows(payload, limit=limit, include_header=include_header, header_row=header_row, data_start_row=data_start_row)
        return
    if ext == ".xlsx":
        yield from _iter_xlsx_rows(payload, limit=limit, include_header=include_header, header_row=header_row, data_start_row=data_start_row)
        return
    if ext == ".xls":
        yield from _iter_xls_rows(payload, limit=limit, include_header=include_header, header_row=header_row, data_start_row=data_start_row)
        return
    raise ValueError(f"unsupported file type: {name}")


def _iter_csv_rows(
    payload: bytes,
    *,
    limit: int | None,
    include_header: bool,
    header_row: int,
    data_start_row: int | None,
) -> Iterator[dict[str, str] | list[str]]:
    encoding = _guess_csv_encoding_bytes(payload)
    text = io.StringIO(payload.decode(encoding, errors="replace"), newline="")
    reader = csv.reader(text)
    try:
        for _ in range(max(0, header_row - 1)):
            next(reader)
        header_layout = _clean_header_layout(next(reader))
    except StopIteration:
        return
    headers = header_layout.headers
    if include_header:
        yield headers
        return
    start_row = data_start_row if data_start_row is not None else header_row + 1
    for _ in range(max(0, start_row - header_row - 1)):
        try:
            next(reader)
        except StopIteration:
            return
    emitted = 0
    for row in reader:
        if not any(str(value or "").strip() for value in row):
            continue
        yield _row_dict(header_layout, row)
        emitted += 1
        if limit is not None and emitted >= limit:
            return


def _iter_xlsx_rows(
    payload: bytes,
    *,
    limit: int | None,
    include_header: bool,
    header_row: int,
    data_start_row: int | None,
) -> Iterator[dict[str, Any] | list[str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read xlsx files") from exc
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        try:
            for _ in range(max(0, header_row - 1)):
                next(rows)
            header_layout = _clean_header_layout(next(rows))
        except StopIteration:
            return
        headers = header_layout.headers
        if include_header:
            yield headers
            return
        start_row = data_start_row if data_start_row is not None else header_row + 1
        for _ in range(max(0, start_row - header_row - 1)):
            try:
                next(rows)
            except StopIteration:
                return
        emitted = 0
        for row in rows:
            if not any(str(value or "").strip() for value in row):
                continue
            yield _row_dict(header_layout, row)
            emitted += 1
            if limit is not None and emitted >= limit:
                return
    finally:
        workbook.close()


def _iter_xls_rows(
    payload: bytes,
    *,
    limit: int | None,
    include_header: bool,
    header_row: int,
    data_start_row: int | None,
) -> Iterator[dict[str, Any] | list[str]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required to read xls files") from exc
    workbook = xlrd.open_workbook(file_contents=payload)
    sheet = workbook.sheet_by_index(0)
    header_index = max(0, header_row - 1)
    if sheet.nrows <= header_index:
        return
    header_layout = _clean_header_layout(sheet.row_values(header_index))
    headers = header_layout.headers
    if include_header:
        yield headers
        return
    emitted = 0
    start_index = (data_start_row - 1) if data_start_row is not None else header_index + 1
    for row_idx in range(max(0, start_index), sheet.nrows):
        values = sheet.row_values(row_idx)
        if not any(str(value or "").strip() for value in values):
            continue
        yield _row_dict(header_layout, values)
        emitted += 1
        if limit is not None and emitted >= limit:
            return


def _read_preview_rows(name: str, ext: str, payload: bytes, *, limit: int) -> list[list[Any]]:
    if ext == ".csv":
        return _read_csv_preview_rows(payload, limit=limit)
    if ext == ".xlsx":
        return _read_xlsx_preview_rows(payload, limit=limit)
    if ext == ".xls":
        return _read_xls_preview_rows(payload, limit=limit)
    raise ValueError(f"unsupported file type: {name}")


def _read_csv_preview_rows(payload: bytes, *, limit: int) -> list[list[Any]]:
    encoding = _guess_csv_encoding_bytes(payload)
    text = io.StringIO(payload.decode(encoding, errors="replace"), newline="")
    rows: list[list[Any]] = []
    for row in csv.reader(text):
        rows.append(list(row))
        if len(rows) >= limit:
            break
    return rows


def _read_xlsx_preview_rows(payload: bytes, *, limit: int) -> list[list[Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read xlsx files") from exc
    workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= limit:
                break
        return rows
    finally:
        workbook.close()


def _read_xls_preview_rows(payload: bytes, *, limit: int) -> list[list[Any]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is required to read xls files") from exc
    workbook = xlrd.open_workbook(file_contents=payload)
    sheet = workbook.sheet_by_index(0)
    return [sheet.row_values(row_idx) for row_idx in range(min(sheet.nrows, limit))]


def _clean_headers(values: Iterable[Any]) -> list[str]:
    return _clean_header_layout(values).headers


def _clean_header_layout(values: Iterable[Any]) -> HeaderLayout:
    indexed_headers = [(index, str(value or "").strip().lstrip("\ufeff")) for index, value in enumerate(values)]
    while indexed_headers and not indexed_headers[-1][1]:
        indexed_headers.pop()
    indexed_headers = [(index, header) for index, header in indexed_headers if header]
    if not indexed_headers:
        raise ValueError("header row contains blank columns")
    return HeaderLayout(
        headers=[header for _, header in indexed_headers],
        indexes=[index for index, _ in indexed_headers],
    )


def _row_dict(header_layout: HeaderLayout, row: Sequence[Any]) -> dict[str, Any]:
    return {
        header: row[index] if index < len(row) else ""
        for header, index in zip(header_layout.headers, header_layout.indexes)
    }


def _guess_csv_encoding_bytes(payload: bytes) -> str:
    sample = payload.splitlines(keepends=True)[0] if payload else b""
    try:
        sample.decode("utf-8-sig")
        return "utf-8-sig"
    except UnicodeDecodeError:
        return "gb18030"


def _iter_upload_payloads(upload_path: str | Path, *, display_name: str | None = None, max_member_bytes: int | None = None) -> Iterator[tuple[str, str, bytes]]:
    path = Path(upload_path)
    ext = _suffix(path.name)
    if ext in SUPPORTED_EXTENSIONS:
        _ensure_size(path.stat().st_size, max_member_bytes, display_name or path.name)
        yield _archive_file_name(display_name or path.name), ext, path.read_bytes()
        return
    if ext == ".zip":
        yield from _iter_zip_payloads(path, max_member_bytes=max_member_bytes)
        return
    if ext == ".7z":
        yield from _iter_7z_payloads(path, max_member_bytes=max_member_bytes)
        return
    if ext == ".rar":
        yield from _iter_rar_payloads(path, max_member_bytes=max_member_bytes)
        return
    raise ValueError(f"unsupported upload type: {ext}")


def _iter_zip_payloads(zip_path: str | Path, *, max_member_bytes: int | None = None) -> Iterator[tuple[str, str, bytes]]:
    with zipfile.ZipFile(zip_path) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            display_name = _display_zip_name(info)
            ext = _suffix(display_name)
            if ext not in SUPPORTED_EXTENSIONS:
                continue
            _ensure_size(info.file_size, max_member_bytes, display_name)
            yield _archive_file_name(display_name), ext, archive.read(info.filename)


def _iter_7z_payloads(archive_path: str | Path, *, max_member_bytes: int | None = None) -> Iterator[tuple[str, str, bytes]]:
    try:
        import py7zr
    except ImportError as exc:
        raise RuntimeError("py7zr is required to read 7z files") from exc
    with tempfile.TemporaryDirectory() as temp_dir_name:
        temp_dir = Path(temp_dir_name).resolve()
        with py7zr.SevenZipFile(archive_path, mode="r") as archive:
            names = archive.getnames()
            archive.extractall(path=temp_dir)
        for name in names:
            ext = _suffix(name)
            if ext not in SUPPORTED_EXTENSIONS:
                continue
            target = (temp_dir / name).resolve()
            if not target.is_file() or not str(target).startswith(str(temp_dir)):
                continue
            _ensure_size(target.stat().st_size, max_member_bytes, name)
            yield _archive_file_name(name), ext, target.read_bytes()


def _iter_rar_payloads(archive_path: str | Path, *, max_member_bytes: int | None = None) -> Iterator[tuple[str, str, bytes]]:
    try:
        import rarfile
    except ImportError as exc:
        raise RuntimeError("rarfile is required to read rar files") from exc
    with rarfile.RarFile(archive_path) as archive:
        for info in archive.infolist():
            if info.isdir():
                continue
            ext = _suffix(info.filename)
            if ext not in SUPPORTED_EXTENSIONS:
                continue
            _ensure_size(info.file_size, max_member_bytes, info.filename)
            with archive.open(info) as raw:
                yield _archive_file_name(info.filename), ext, raw.read()


def _archive_file_name(name: str) -> str:
    return PurePosixPath(str(name or "").replace("\\", "/")).name


def _ensure_size(size: int, limit: int | None, name: str) -> None:
    if limit is not None and size > limit:
        raise ValueError(f"{_archive_file_name(name)} is too large")


def _semantic_score(source: str, target: TableColumn) -> float:
    source_norm = _normalize_label(source)
    target_name = _normalize_label(target.name)
    target_comment = _normalize_label(target.comment)
    candidates = [candidate for candidate in [target_comment, target_name] if candidate]
    if not source_norm or not candidates:
        return 0
    best = 0.0
    for candidate in candidates:
        if source_norm == candidate:
            best = max(best, 1.0)
        elif source_norm in candidate or candidate in source_norm:
            shorter = min(len(source_norm), len(candidate))
            longer = max(len(source_norm), len(candidate))
            best = max(best, 0.72 + 0.18 * (shorter / longer))
        else:
            best = max(best, SequenceMatcher(None, source_norm, candidate).ratio())
    return best


def _normalize_label(value: str) -> str:
    text = str(value or "").lower()
    text = re.sub(r"[\s_\-./()（）【】\[\]:：,，]+", "", text)
    for token in ["字段", "名称", "代码", "编号"]:
        if len(text) > len(token) + 1:
            text = text.replace(token, token)
    return text
