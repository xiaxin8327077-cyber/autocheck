from __future__ import annotations

import json
import re
import sqlite3
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auto_check.app.local_store import _connect, db_path_for_config
from auto_check.app.storage_schema import CURRENT_SCHEMA_VERSION, backup_database_if_exists


SCHEMA_EXPORT_FILENAME = "本地数据库表结构.xlsx"
ALLOWED_PAGE_SIZES = {10, 20, 50, 100}
DEFAULT_PAGE_SIZE = 20
MASKED_VALUE = "******"
SENSITIVE_KEY_RE = re.compile(
    r"(password|password_hash|password_encrypted|token|secret|credential|host|username|source_path|download_path|excel_path)",
    re.I,
)


@dataclass(frozen=True)
class StorageField:
    name: str
    cn_name: str
    display: str = "原文"
    sensitive: bool = False


@dataclass(frozen=True)
class StorageTable:
    name: str
    cn_name: str
    category: str
    primary_key: tuple[str, ...]
    purpose: str
    fields: tuple[StorageField, ...]
    default_order_by: tuple[tuple[str, str], ...] = ()
    internal: bool = False
    allow_rows: bool = True


def _field(name: str, cn_name: str, display: str = "原文", *, sensitive: bool = False) -> StorageField:
    return StorageField(name=name, cn_name=cn_name, display=display, sensitive=sensitive)


STORAGE_TABLES: tuple[StorageTable, ...] = (
    StorageTable(
        name="data_sources",
        cn_name="数据源配置表",
        category="配置与用户",
        primary_key=("id",),
        purpose="保存 DWS、报表库等本地数据源连接配置和默认标记。",
        fields=(
            _field("id", "数据源 ID"),
            _field("name", "数据源名称"),
            _field("db_type", "数据库类型", "标签"),
            _field("host", "主机地址", "脱敏", sensitive=True),
            _field("port", "端口", "数字"),
            _field("database_name", "数据库名"),
            _field("schema_name", "Schema"),
            _field("username", "用户名", "脱敏", sensitive=True),
            _field("password_encrypted", "加密密码", "脱敏", sensitive=True),
            _field("is_default", "是否默认", "标签"),
            _field("created_at", "创建时间"),
            _field("updated_at", "更新时间"),
        ),
        default_order_by=(("updated_at", "DESC"), ("name", "ASC")),
    ),
    StorageTable(
        name="app_settings",
        cn_name="应用设置表",
        category="配置与用户",
        primary_key=("key",),
        purpose="保存系统设置、默认设置、人行逐笔校验设置、流程工具设置等结构化配置。",
        fields=(
            _field("key", "设置键"),
            _field("value_json", "设置内容", "JSON 抽屉"),
            _field("updated_at", "更新时间"),
        ),
        default_order_by=(("updated_at", "DESC"), ("key", "ASC")),
    ),
    StorageTable(
        name="users",
        cn_name="用户账号表",
        category="配置与用户",
        primary_key=("id",),
        purpose="保存用户、角色、状态、密码哈希和登录时间。",
        fields=(
            _field("id", "用户 ID"),
            _field("username", "登录账号", "脱敏", sensitive=True),
            _field("display_name", "展示名"),
            _field("role", "角色", "标签"),
            _field("password_hash", "密码哈希", "脱敏", sensitive=True),
            _field("enabled", "是否启用", "标签"),
            _field("created_at", "创建时间"),
            _field("updated_at", "更新时间"),
            _field("last_login_at", "最近登录时间"),
        ),
        default_order_by=(("updated_at", "DESC"), ("username", "ASC")),
    ),
    StorageTable(
        name="run_headers",
        cn_name="历史运行头表",
        category="历史公共",
        primary_key=("id",),
        purpose="保存各类历史运行的公共字段和完整 payload 快照。",
        fields=(
            _field("id", "运行 ID"),
            _field("kind", "历史类型", "标签"),
            _field("run_date", "业务日期"),
            _field("run_at", "开始时间"),
            _field("finished_at", "完成时间"),
            _field("status", "状态", "标签"),
            _field("executor_id", "执行人 ID"),
            _field("executor_username", "执行账号", "脱敏", sensitive=True),
            _field("executor_name", "执行人"),
            _field("config_fingerprint", "配置指纹"),
            _field("payload_json", "完整快照", "JSON 抽屉"),
        ),
        default_order_by=(("run_date", "DESC"), ("run_at", "DESC")),
    ),
    StorageTable(
        name="reconcile_runs",
        cn_name="自动对数运行表",
        category="自动对数",
        primary_key=("id",),
        purpose="保存自动对数运行摘要、配置名称、规则版本和增量数量。",
        fields=(
            _field("id", "运行 ID"),
            _field("config_name", "配置名称"),
            _field("dws_source_name", "DWS 数据源"),
            _field("rule_version", "规则版本"),
            _field("baseline_id", "基准运行 ID"),
            _field("baseline_run_at", "基准运行时间"),
            _field("baseline_count", "基准条数", "数字"),
            _field("total_count", "差异总数", "数字"),
            _field("added_count", "新增差异数", "数字"),
            _field("removed_count", "减少差异数", "数字"),
        ),
        default_order_by=(("id", "DESC"),),
    ),
    StorageTable(
        name="reconcile_run_counts",
        cn_name="自动对数运行统计表",
        category="自动对数",
        primary_key=("run_id", "count_type", "label"),
        purpose="保存匹配状态、差异类型等聚合统计。",
        fields=(
            _field("run_id", "运行 ID"),
            _field("count_type", "统计类型", "标签"),
            _field("label", "统计标签"),
            _field("count_value", "统计值", "数字"),
        ),
        default_order_by=(("run_id", "DESC"), ("count_type", "ASC"), ("label", "ASC")),
    ),
    StorageTable(
        name="reconcile_results",
        cn_name="自动对数结果明细表",
        category="自动对数",
        primary_key=("id",),
        purpose="保存项目编号、差异类型、匹配状态、差异金额等结果热字段。",
        fields=(
            _field("id", "结果 ID", "数字"),
            _field("run_id", "运行 ID"),
            _field("result_order", "结果顺序", "数字"),
            _field("project_code", "项目编号"),
            _field("project_name", "项目名称"),
            _field("asset_total", "资产合计"),
            _field("liability_equity_total", "负债及权益合计"),
            _field("received_trust_balance", "实收信托余额"),
            _field("difference", "差异金额"),
            _field("direction", "差异方向"),
            _field("difference_reason", "差异类型", "标签"),
            _field("match_status", "匹配状态", "标签"),
            _field("valuation_asset_total", "估值表资产合计"),
            _field("payload_json", "结果快照", "JSON 抽屉"),
        ),
        default_order_by=(("run_id", "DESC"), ("result_order", "ASC")),
    ),
    StorageTable(
        name="reconcile_result_details",
        cn_name="自动对数结果详情表",
        category="自动对数",
        primary_key=("id",),
        purpose="保存结构化详情类型、具体原因和详情 payload。",
        fields=(
            _field("id", "详情 ID", "数字"),
            _field("result_id", "结果 ID", "数字"),
            _field("detail_order", "详情顺序", "数字"),
            _field("kind", "详情类型", "标签"),
            _field("specific_reason", "具体原因"),
            _field("data_json", "详情数据", "JSON 抽屉"),
        ),
        default_order_by=(("result_id", "DESC"), ("detail_order", "ASC")),
    ),
    StorageTable(
        name="reconcile_delta_results",
        cn_name="自动对数增量差异表",
        category="自动对数",
        primary_key=("run_id", "delta_type", "result_order"),
        purpose="保存本次新增差异和减少差异快照。",
        fields=(
            _field("run_id", "运行 ID"),
            _field("delta_type", "增量类型", "标签"),
            _field("result_order", "结果顺序", "数字"),
            _field("payload_json", "增量快照", "JSON 抽屉"),
        ),
        default_order_by=(("run_id", "DESC"), ("result_order", "ASC")),
    ),
    StorageTable(
        name="db_validation_runs",
        cn_name="人行逐笔校验运行表",
        category="人行逐笔校验",
        primary_key=("id",),
        purpose="保存逐笔校验报告期、结果数、告警数、校验开关和下载路径。",
        fields=(
            _field("id", "运行 ID"),
            _field("report_date", "报告期"),
            _field("result_count", "结果数", "数字"),
            _field("warning_count", "告警数", "数字"),
            _field("table_count", "选表数量", "数字"),
            _field("enable_public_info_check", "公开信息校验", "标签"),
            _field("enable_template_check", "模板校验", "标签"),
            _field("excel_filename", "结果文件名"),
            _field("excel_path", "结果文件路径", "脱敏", sensitive=True),
            _field("download_url", "下载地址", "脱敏", sensitive=True),
        ),
        default_order_by=(("report_date", "DESC"), ("id", "DESC")),
    ),
    StorageTable(
        name="db_validation_selected_tables",
        cn_name="人行逐笔校验选表明细表",
        category="人行逐笔校验",
        primary_key=("run_id", "table_order"),
        purpose="保存一次逐笔校验运行中勾选的 ZG 表清单。",
        fields=(
            _field("run_id", "运行 ID"),
            _field("table_order", "选表顺序", "数字"),
            _field("table_code", "ZG 表编号", "标签"),
        ),
        default_order_by=(("run_id", "DESC"), ("table_order", "ASC")),
    ),
    StorageTable(
        name="db_validation_warnings",
        cn_name="人行逐笔校验告警表",
        category="人行逐笔校验",
        primary_key=("run_id", "warning_order"),
        purpose="保存一次逐笔校验运行产生的告警信息。",
        fields=(
            _field("run_id", "运行 ID"),
            _field("warning_order", "告警顺序", "数字"),
            _field("message", "告警内容"),
        ),
        default_order_by=(("run_id", "DESC"), ("warning_order", "ASC")),
    ),
    StorageTable(
        name="db_validation_result_rows",
        cn_name="人行逐笔校验结果行表",
        category="人行逐笔校验",
        primary_key=("id",),
        purpose="保存逐笔校验结果行的表号、规则、级别、消息和完整行快照。",
        fields=(
            _field("id", "结果行 ID", "数字"),
            _field("run_id", "运行 ID"),
            _field("row_order", "行顺序", "数字"),
            _field("table_code", "ZG 表编号", "标签"),
            _field("rule_id", "规则编号"),
            _field("severity", "级别", "标签"),
            _field("message", "消息"),
            _field("detail", "详情"),
            _field("payload_json", "行快照", "JSON 抽屉"),
        ),
        default_order_by=(("run_id", "DESC"), ("row_order", "ASC")),
    ),
    StorageTable(
        name="flow_chain_runs",
        cn_name="流程链执行运行表",
        category="流程链",
        primary_key=("id",),
        purpose="保存流程链名称、触发方式、执行人、状态、错误、步骤数和总耗时。",
        fields=(
            _field("id", "运行 ID"),
            _field("chain_id", "链路编号"),
            _field("chain_name", "链路名称"),
            _field("is_multi_chain", "是否多链路", "标签"),
            _field("trigger_type", "触发方式", "标签"),
            _field("executor_name", "执行人"),
            _field("status", "状态", "标签"),
            _field("error", "错误信息"),
            _field("step_count", "步骤数", "数字"),
            _field("duration_seconds", "耗时秒数", "数字"),
        ),
        default_order_by=(("id", "DESC"),),
    ),
    StorageTable(
        name="flow_chain_run_steps",
        cn_name="流程链执行步骤表",
        category="流程链",
        primary_key=("id",),
        purpose="保存每个流程步骤的流程编号、名称、状态、任务号和起止时间。",
        fields=(
            _field("id", "步骤 ID", "数字"),
            _field("run_id", "运行 ID"),
            _field("step_order", "步骤顺序", "数字"),
            _field("flow_id", "流程编号"),
            _field("name", "步骤名称"),
            _field("status", "状态", "标签"),
            _field("sp_task_id", "申报平台任务号"),
            _field("start_time", "开始时间"),
            _field("end_time", "结束时间"),
            _field("duration_seconds", "耗时秒数", "数字"),
            _field("payload_json", "步骤快照", "JSON 抽屉"),
        ),
        default_order_by=(("run_id", "DESC"), ("step_order", "ASC")),
    ),
    StorageTable(
        name="flow_chain_run_logs",
        cn_name="流程链执行日志表",
        category="流程链",
        primary_key=("id",),
        purpose="保存流程链执行过程中的日志、进度和当前步骤。",
        fields=(
            _field("id", "日志 ID", "数字"),
            _field("run_id", "运行 ID"),
            _field("log_order", "日志顺序", "数字"),
            _field("log_time", "日志时间"),
            _field("message", "日志内容"),
            _field("progress", "进度", "数字"),
            _field("step", "当前步骤"),
            _field("payload_json", "日志快照", "JSON 抽屉"),
        ),
        default_order_by=(("run_id", "DESC"), ("log_order", "ASC")),
    ),
    StorageTable(
        name="flow_chain_run_details",
        cn_name="流程链执行链路明细表",
        category="流程链",
        primary_key=("id",),
        purpose="保存单链路或多链路合并历史中的链路详情。",
        fields=(
            _field("id", "链路明细 ID", "数字"),
            _field("run_id", "运行 ID"),
            _field("chain_order", "链路顺序", "数字"),
            _field("chain_name", "链路名称"),
            _field("status", "状态", "标签"),
            _field("step_count", "步骤数", "数字"),
            _field("duration_seconds", "耗时秒数", "数字"),
            _field("error", "错误信息"),
            _field("payload_json", "链路快照", "JSON 抽屉"),
        ),
        default_order_by=(("run_id", "DESC"), ("chain_order", "ASC")),
    ),
    StorageTable(
        name="app_kv",
        cn_name="旧版键值快照表",
        category="兼容与迁移",
        primary_key=("key",),
        purpose="保留旧版 config_store、auth 等键值快照，作为兼容回退来源。",
        fields=(
            _field("key", "键"),
            _field("value", "值", "JSON 抽屉"),
            _field("updated_at", "更新时间"),
        ),
        default_order_by=(("updated_at", "DESC"), ("key", "ASC")),
        internal=True,
    ),
    StorageTable(
        name="history_runs",
        cn_name="旧版历史兼容表",
        category="兼容与迁移",
        primary_key=("kind", "id"),
        purpose="保留旧版历史 payload，按 kind 区分 reconcile、db_validation、flow_chain。",
        fields=(
            _field("kind", "历史类型", "标签"),
            _field("id", "历史 ID"),
            _field("payload", "历史快照", "JSON 抽屉"),
            _field("run_date", "业务日期"),
            _field("run_at", "执行时间"),
            _field("config_fingerprint", "配置指纹"),
        ),
        default_order_by=(("run_date", "DESC"), ("run_at", "DESC")),
        internal=True,
    ),
    StorageTable(
        name="config_snapshots",
        cn_name="配置兼容快照表",
        category="兼容与迁移",
        primary_key=("id",),
        purpose="保存完整配置 payload 快照，便于兼容旧结构和后续排查。",
        fields=(
            _field("id", "快照 ID", "数字"),
            _field("fingerprint", "配置指纹"),
            _field("payload_json", "配置快照", "JSON 抽屉"),
            _field("created_at", "创建时间"),
        ),
        default_order_by=(("created_at", "DESC"), ("id", "DESC")),
        internal=True,
    ),
    StorageTable(
        name="schema_migrations",
        cn_name="存储结构版本表",
        category="兼容与迁移",
        primary_key=("version",),
        purpose="记录本地 SQLite schema 已执行到的版本。",
        fields=(
            _field("version", "版本号", "数字"),
            _field("applied_at", "应用时间"),
        ),
        default_order_by=(("version", "DESC"),),
        internal=True,
    ),
    StorageTable(
        name="storage_migration_runs",
        cn_name="数据迁移记录表",
        category="兼容与迁移",
        primary_key=("id",),
        purpose="记录旧 SQLite/旧 JSON 来源的迁移路径、指纹、条数和状态。",
        fields=(
            _field("id", "迁移 ID", "数字"),
            _field("source_type", "来源类型", "标签"),
            _field("source_path", "来源路径", "脱敏", sensitive=True),
            _field("source_key", "来源子类"),
            _field("source_fingerprint", "来源指纹"),
            _field("migrated_count", "迁移条数", "数字"),
            _field("skipped_count", "跳过条数", "数字"),
            _field("status", "状态", "标签"),
            _field("message", "消息"),
            _field("started_at", "开始时间"),
            _field("finished_at", "完成时间"),
        ),
        default_order_by=(("id", "DESC"),),
        internal=True,
    ),
    StorageTable(
        name="sqlite_sequence",
        cn_name="SQLite 自增序列表",
        category="兼容与迁移",
        primary_key=(),
        purpose="SQLite 内部表，用于维护 AUTOINCREMENT 序列。",
        fields=(
            _field("name", "表名"),
            _field("seq", "当前序号", "数字"),
        ),
        default_order_by=(("name", "ASC"),),
        internal=True,
    ),
)

TABLE_BY_NAME = {table.name: table for table in STORAGE_TABLES}


def build_storage_health(config_path: str | Path) -> dict[str, Any]:
    db_path = db_path_for_config(config_path)
    with _connect(db_path) as connection:
        integrity_rows = [str(row[0]) for row in connection.execute("PRAGMA integrity_check").fetchall()]
        foreign_key_rows = connection.execute("PRAGMA foreign_key_check").fetchall()
        schema_version = _schema_version(connection)
        table_counts = {
            table.name: _table_count(connection, table.name) if _table_exists(connection, table.name) else 0
            for table in STORAGE_TABLES
        }
        history_counts = {
            str(row["kind"]): int(row["count"])
            for row in connection.execute(
                "SELECT kind, COUNT(*) AS count FROM run_headers GROUP BY kind ORDER BY kind"
            ).fetchall()
        }
        latest_migration = _latest_migration(connection)

    return {
        "schema_version": schema_version,
        "expected_schema_version": CURRENT_SCHEMA_VERSION,
        "integrity_check": "ok" if integrity_rows == ["ok"] else "; ".join(integrity_rows),
        "foreign_key_issues": len(foreign_key_rows),
        "database": _database_info(db_path),
        "latest_backup": _latest_backup(db_path.parent),
        "business_table_count": sum(1 for table in STORAGE_TABLES if not table.internal),
        "table_counts": table_counts,
        "history_counts": history_counts,
        "latest_migration": latest_migration,
    }


def list_storage_tables(config_path: str | Path) -> list[dict[str, Any]]:
    db_path = db_path_for_config(config_path)
    with _connect(db_path) as connection:
        return [_table_catalog_item(connection, table) for table in STORAGE_TABLES]


def get_storage_table_schema(config_path: str | Path, table_name: str) -> dict[str, Any]:
    table = _require_table(table_name)
    db_path = db_path_for_config(config_path)
    with _connect(db_path) as connection:
        if not _table_exists(connection, table.name):
            raise LookupError("table not found")
        fields = _table_fields(connection, table)
    return {"table": _table_info(table, len(fields), 0), "fields": fields}


def get_storage_table_rows(
    config_path: str | Path,
    table_name: str,
    *,
    page: int = 1,
    page_size: int = DEFAULT_PAGE_SIZE,
) -> dict[str, Any]:
    table = _require_table(table_name)
    if not table.allow_rows:
        raise PermissionError("table rows are not viewable")
    page = max(1, int(page or 1))
    page_size = int(page_size or DEFAULT_PAGE_SIZE)
    if page_size not in ALLOWED_PAGE_SIZES:
        raise ValueError("page_size must be one of 10, 20, 50, 100")

    db_path = db_path_for_config(config_path)
    with _connect(db_path) as connection:
        if not _table_exists(connection, table.name):
            raise LookupError("table not found")
        fields = _table_fields(connection, table)
        field_names = [field["name"] for field in fields]
        total = _table_count(connection, table.name)
        rows = _select_rows(connection, table, field_names, page=page, page_size=page_size)
    field_meta = {
        field["name"]: {
            "display": field["display"],
            "cn_name": field["cn_name"],
            "sensitive": field["sensitive"],
        }
        for field in fields
    }
    return {
        "table": _table_info(table, len(fields), total),
        "rows": rows,
        "total": total,
        "page": page,
        "page_size": page_size,
        "fields": field_meta,
    }


def build_storage_schema_workbook(config_path: str | Path) -> tuple[str, bytes]:
    db_path = db_path_for_config(config_path)
    with _connect(db_path) as connection:
        catalog_items = [_table_catalog_item(connection, table) for table in STORAGE_TABLES]
        table_fields = {
            table.name: _table_fields(connection, table) if _table_exists(connection, table.name) else []
            for table in STORAGE_TABLES
        }

    workbook = Workbook()
    catalog_sheet = workbook.active
    catalog_sheet.title = "表目录"
    _write_rows(
        catalog_sheet,
        [
            ["分类", "表名", "中文名", "记录数", "字段数", "主键", "用途"],
            *[
                [
                    item["category"],
                    item["name"],
                    item["cn_name"],
                    item["record_count"],
                    item["field_count"],
                    ", ".join(item["primary_key"]) or "-",
                    item["purpose"],
                ]
                for item in catalog_items
            ],
        ],
    )

    used_sheet_names = {"表目录"}
    for table in STORAGE_TABLES:
        sheet = workbook.create_sheet(_safe_sheet_name(table.name, used_sheet_names))
        rows = [
            ["字段序号", "字段名", "类型", "约束", "中文说明", "展示策略"],
            *[
                [
                    index,
                    field["name"],
                    field["type"],
                    _constraint_text(field),
                    field["cn_name"],
                    field["display"],
                ]
                for index, field in enumerate(table_fields[table.name], start=1)
            ],
        ]
        _write_rows(sheet, rows)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return SCHEMA_EXPORT_FILENAME, buffer.getvalue()


def build_storage_table_data_workbook(config_path: str | Path, table_name: str) -> tuple[str, bytes]:
    table = _require_table(table_name)
    if not table.allow_rows:
        raise PermissionError("table rows are not viewable")

    db_path = db_path_for_config(config_path)
    with _connect(db_path) as connection:
        if not _table_exists(connection, table.name):
            raise LookupError("table not found")
        fields = _table_fields(connection, table)
        field_names = [field["name"] for field in fields]
        total = _table_count(connection, table.name)
        rows = _select_rows(connection, table, field_names, page=1, page_size=max(total, 1))

    workbook = Workbook()
    sheet = workbook.active
    used_sheet_names: set[str] = set()
    sheet.title = _safe_sheet_name(table.name, used_sheet_names)
    _write_rows(
        sheet,
        [
            [field["cn_name"] or field["name"] for field in fields],
            *[
                [_workbook_cell_value(row.get(field["name"])) for field in fields]
                for row in rows
            ],
        ],
    )

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return f"{table.name}-表数据.xlsx", buffer.getvalue()


def generate_storage_backup(config_path: str | Path) -> dict[str, Any]:
    db_path = db_path_for_config(config_path)
    with _connect(db_path):
        pass
    backup_path = backup_database_if_exists(db_path)
    if backup_path is None:
        raise FileNotFoundError("database file not found")
    return {
        "filename": backup_path.name,
        "display_path": _masked_path(backup_path),
        "created_at": _format_mtime(backup_path),
        "size_bytes": backup_path.stat().st_size,
        "size_text": _format_size(backup_path.stat().st_size),
    }


def _require_table(table_name: str) -> StorageTable:
    table = TABLE_BY_NAME.get(str(table_name or ""))
    if table is None:
        raise LookupError("table not found")
    return table


def _table_catalog_item(connection: sqlite3.Connection, table: StorageTable) -> dict[str, Any]:
    field_count = len(_table_fields(connection, table)) if _table_exists(connection, table.name) else len(table.fields)
    record_count = _table_count(connection, table.name) if _table_exists(connection, table.name) else 0
    return _table_info(table, field_count, record_count)


def _table_info(table: StorageTable, field_count: int, record_count: int) -> dict[str, Any]:
    return {
        "name": table.name,
        "cn_name": table.cn_name,
        "category": table.category,
        "primary_key": list(table.primary_key),
        "field_count": int(field_count),
        "record_count": int(record_count),
        "purpose": table.purpose,
        "internal": table.internal,
        "allow_rows": table.allow_rows,
    }


def _schema_version(connection: sqlite3.Connection) -> int:
    row = connection.execute("SELECT MAX(version) FROM schema_migrations").fetchone()
    if row is None or row[0] is None:
        return 0
    return int(row[0])


def _latest_migration(connection: sqlite3.Connection) -> dict[str, Any] | None:
    row = connection.execute(
        """
        SELECT source_type, source_key, status, migrated_count, skipped_count, message, finished_at
        FROM storage_migration_runs
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    return dict(row) if row is not None else None


def _table_exists(connection: sqlite3.Connection, table_name: str) -> bool:
    row = connection.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table_name,),
    ).fetchone()
    return row is not None


def _table_count(connection: sqlite3.Connection, table_name: str) -> int:
    quoted = _quote_identifier(table_name)
    row = connection.execute(f"SELECT COUNT(*) FROM {quoted}").fetchone()
    return int(row[0]) if row is not None else 0


def _table_fields(connection: sqlite3.Connection, table: StorageTable) -> list[dict[str, Any]]:
    rows = connection.execute(f"PRAGMA table_info({_quote_identifier(table.name)})").fetchall()
    meta_by_name = {field.name: field for field in table.fields}
    return [_field_info(row, meta_by_name.get(str(row["name"]))) for row in rows]


def _field_info(row: sqlite3.Row, meta: StorageField | None) -> dict[str, Any]:
    name = str(row["name"])
    sensitive = bool(meta.sensitive) if meta else _is_sensitive_key(name)
    if _is_sensitive_key(name):
        sensitive = True
    display = meta.display if meta else "原文"
    return {
        "name": name,
        "type": str(row["type"] or ""),
        "primary_key": bool(row["pk"]),
        "not_null": bool(row["notnull"]),
        "nullable": not bool(row["notnull"]),
        "default": row["dflt_value"],
        "cn_name": meta.cn_name if meta else name,
        "display": display,
        "sensitive": sensitive,
    }


def _select_rows(
    connection: sqlite3.Connection,
    table: StorageTable,
    field_names: list[str],
    *,
    page: int,
    page_size: int,
) -> list[dict[str, Any]]:
    columns = ", ".join(_quote_identifier(field) for field in field_names)
    order_clause = _order_clause(table, field_names)
    offset = (page - 1) * page_size
    rows = connection.execute(
        f"SELECT {columns} FROM {_quote_identifier(table.name)} {order_clause} LIMIT ? OFFSET ?",
        (page_size, offset),
    ).fetchall()
    field_info = {field.name: field for field in table.fields}
    return [_serialize_row(row, field_names, field_info) for row in rows]


def _serialize_row(row: sqlite3.Row, field_names: Iterable[str], field_info: dict[str, StorageField]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for field_name in field_names:
        meta = field_info.get(field_name)
        value = row[field_name]
        if meta and meta.display == "JSON 抽屉":
            value = _parse_json_value(value)
        payload[field_name] = _mask_value(field_name, value, bool(meta.sensitive) if meta else False)
    return payload


def _parse_json_value(value: Any) -> Any:
    if value is None or isinstance(value, (dict, list)):
        return value
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return ""
    if not (stripped.startswith("{") or stripped.startswith("[")):
        return value
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        return value


def _workbook_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, indent=2)
    return value


def _mask_value(key: str, value: Any, sensitive: bool = False) -> Any:
    if isinstance(value, dict):
        return {item_key: _mask_value(item_key, item_value, _is_sensitive_key(item_key)) for item_key, item_value in value.items()}
    if isinstance(value, list):
        return [_mask_value(key, item, sensitive) for item in value]
    if not (sensitive or _is_sensitive_key(key)):
        return value
    if value in (None, ""):
        return value
    text = str(value)
    if "path" in key.lower():
        return _masked_path(Path(text))
    if "host" in key.lower():
        return _mask_host(text)
    if "username" in key.lower():
        return _mask_username(text)
    return MASKED_VALUE


def _is_sensitive_key(key: str) -> bool:
    return bool(SENSITIVE_KEY_RE.search(str(key or "")))


def _mask_host(value: str) -> str:
    if re.match(r"^\d{1,3}(?:\.\d{1,3}){3}$", value):
        return re.sub(r"\.\d{1,3}$", ".***", value)
    if len(value) <= 3:
        return "***"
    return f"{value[:2]}***"


def _mask_username(value: str) -> str:
    if len(value) <= 2:
        return "***"
    return f"{value[:2]}***"


def _masked_path(path: Path) -> str:
    name = path.name or "auto-check.db"
    parent_name = path.parent.name
    if parent_name:
        return f".../{parent_name}/{name}"
    return f".../{name}"


def _database_info(db_path: Path) -> dict[str, Any]:
    exists = db_path.exists()
    size = db_path.stat().st_size if exists else 0
    return {
        "filename": db_path.name,
        "exists": exists,
        "display_path": _masked_path(db_path),
        "size_bytes": size,
        "size_text": _format_size(size),
        "modified_at": _format_mtime(db_path) if exists else "",
    }


def _latest_backup(directory: Path) -> dict[str, Any] | None:
    candidates: list[Path] = []
    if directory.exists():
        for backup_dir in directory.glob("backup-before-storage-v2-*"):
            backup_file = backup_dir / "auto-check.db"
            if backup_file.exists():
                candidates.append(backup_file)
    if not candidates:
        return None
    latest = max(candidates, key=lambda item: item.stat().st_mtime)
    return {
        "filename": latest.name,
        "display_path": _masked_path(latest),
        "created_at": _format_mtime(latest),
        "size_bytes": latest.stat().st_size,
        "size_text": _format_size(latest.stat().st_size),
    }


def _format_mtime(path: Path) -> str:
    from datetime import datetime

    return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")


def _format_size(size: int) -> str:
    units = ["B", "KB", "MB", "GB"]
    value = float(size)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{size} B"


def _order_clause(table: StorageTable, field_names: list[str]) -> str:
    clauses: list[str] = []
    available = set(field_names)
    for field_name, direction in table.default_order_by:
        if field_name in available:
            safe_direction = "DESC" if str(direction).upper() == "DESC" else "ASC"
            clauses.append(f"{_quote_identifier(field_name)} {safe_direction}")
    if not clauses:
        for field_name in table.primary_key:
            if field_name in available:
                clauses.append(f"{_quote_identifier(field_name)} ASC")
    return f"ORDER BY {', '.join(clauses)}" if clauses else ""


def _quote_identifier(identifier: str) -> str:
    return '"' + str(identifier).replace('"', '""') + '"'


def _constraint_text(field: dict[str, Any]) -> str:
    parts: list[str] = []
    if field["primary_key"]:
        parts.append("PK")
    if field["not_null"]:
        parts.append("NOT NULL")
    if field["default"] not in (None, ""):
        parts.append(f"DEFAULT {field['default']}")
    return " / ".join(parts) or "-"


def _safe_sheet_name(base_name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\[\]\*\/\\\?:]", "_", str(base_name or "sheet"))[:31] or "sheet"
    candidate = cleaned
    index = 2
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def _write_rows(sheet: Any, rows: list[list[Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(bold=True, color="0F172A")
    for row_index, values in enumerate(rows, start=1):
        sheet.append(values)
        for cell in sheet[row_index]:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_index == 1:
                cell.fill = header_fill
                cell.font = header_font
    sheet.freeze_panes = "A2"
    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 42)
