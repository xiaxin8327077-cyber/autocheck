from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auto_check.db_validation.legacy_rules import ACTIVE_LEGACY_RULES, LegacyRule


RULE_DOCUMENT_FILENAME = "数据库校验规则说明.xlsx"


FORM_NAMES: dict[str, str] = {
    "ZG01": "资管产品基本信息",
    "ZG02": "资管产品初始募集信息",
    "ZG03": "资管产品终止信息",
    "ZG04": "资管产品存续募集信息",
    "ZG05": "资管产品资产负债信息",
    "ZG06": "资产收益权明细信息",
    "ZG07": "除回购和拆借外贷款明细信息",
    "ZG08": "特定目的载体交易对手明细信息",
    "ZG09": "资产负债剩余期限信息",
    "ZG10": "债券等资产配置情况信息",
    "ZG11": "行业投向信息",
    "ZG12": "地方政府融资平台及房地产明细信息",
    "ZG13": "其他股权投资明细信息",
}

DOCUMENT_ENABLED_RULE_IDS: frozenset[str] = frozenset({"Zg09_Rule3", "Zg10_Rule1"})
TEMPLATE_RULE_NOTES: dict[str, str] = {
    "Zg09_Rule3": (
        "\u53d7\u6267\u884c\u754c\u9762\u7684\u201c\u6a21\u677f\u6821\u9a8c\u201d\u52fe\u9009\u9879\u63a7\u5236\uff1b"
        "cpkj=1 \u5bf9\u6bd4 balance_sheet_info\uff1b"
        "cpkj=2 \u5bf9\u6bd4 balance_sheet_info_zcglxt\uff082a\uff09\u3002"
    ),
    "Zg10_Rule1": (
        "\u53d7\u6267\u884c\u754c\u9762\u7684\u201c\u6a21\u677f\u6821\u9a8c\u201d\u52fe\u9009\u9879\u63a7\u5236\uff1b"
        "cpkj=1 \u5bf9\u6bd4 balance_sheet_info2\uff1b"
        "cpkj=2 \u5bf9\u6bd4 balance_sheet_info2_zcglxt\uff082a\uff09\u3002"
    ),
}


@dataclass(frozen=True)
class UserRule:
    form_code: str
    form_name: str
    rule_id: str
    category: str
    status: str
    check_item: str
    trigger_condition: str
    result_message: str
    cross_period: str
    note: str = ""


def build_rules_document() -> tuple[str, bytes]:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "使用说明"
    summary = workbook.create_sheet("规则清单")
    detail = workbook.create_sheet("规则明细")

    rules = tuple(_user_rule(rule) for rule in ACTIVE_LEGACY_RULES)
    _write_overview(overview, rules)
    _write_summary(summary, rules)
    _write_detail(detail, rules)

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return RULE_DOCUMENT_FILENAME, buffer.getvalue()


def _rule_document_enabled(rule: LegacyRule) -> bool:
    return rule.enabled or rule.rule_id in DOCUMENT_ENABLED_RULE_IDS


def _user_rule(rule: LegacyRule) -> UserRule:
    message = _message(rule)
    note = _note(rule)
    return UserRule(
        form_code=rule.zg_code,
        form_name=FORM_NAMES.get(rule.zg_code, rule.zg_code),
        rule_id=rule.rule_id,
        category=rule.category,
        status="启用" if _rule_document_enabled(rule) else "停用",
        check_item=_check_item(rule, message),
        trigger_condition=_trigger_condition(rule, message),
        result_message=message,
        cross_period="是" if _is_cross_period(message) else "否",
        note=note,
    )


def _message(rule: LegacyRule) -> str:
    return rule.rule_text.split(":", 1)[-1].split("：", 1)[-1].strip()


def _check_item(rule: LegacyRule, message: str) -> str:
    if rule.is_template_rule:
        return "将数据平台填报值与模板数据进行一致性比对。"
    if rule.is_public_info_rule:
        return "将逐笔数据与公开信息表或交易对手填报信息进行交叉比对。"
    if _is_cross_period(message):
        return "检查本期数据与上期数据之间是否保持旧程序要求的一致性或衔接关系。"
    if "人民币合计" in message:
        return "检查人民币合计、人民币金额或折人民币金额之间是否一致。"
    if "地区代码" in message:
        return "检查地区代码是否按旧程序地区字典填报到区县一级。"
    if "代码" in message and "编码规则" in message:
        return "检查机构、客户、借款人或交易场所代码是否符合旧程序编码规则。"
    if "同一" in message and "不一致" in message:
        return "检查同一主体或同一业务编号下的关键字段是否一致。"
    return f"检查{_strip_suffix(message)}。"


def _trigger_condition(rule: LegacyRule, message: str) -> str:
    body = _strip_suffix(message)
    if rule.is_template_rule:
        return f"启用模板校验并取得模板数据后，旧程序发现{body}时提示。"
    if rule.is_public_info_rule:
        return f"勾选公开信息校验后，旧程序发现{body}时提示。"
    if _is_cross_period(message):
        return f"按旧程序匹配键对本期和上期数据进行比对，发现{body}时提示。"
    return f"旧程序发现{body}时提示。"


def _note(rule: LegacyRule) -> str:
    notes: list[str] = []
    if rule.is_public_info_rule:
        notes.append("受执行界面的“公开信息校验”勾选项控制。")
    if rule.rule_id in TEMPLATE_RULE_NOTES:
        notes.append(TEMPLATE_RULE_NOTES[rule.rule_id])
    elif rule.is_template_rule:
        notes.append("模板数据源已在配置中预留，启用模板校验后执行。")
    if not _rule_document_enabled(rule) and rule.disabled_reason:
        notes.append(rule.disabled_reason)
    return "".join(notes)


def _rule_label(rule: UserRule) -> str:
    labels = [rule.rule_id, f"[{rule.category}]"]
    if rule.status == "停用":
        labels.append("[停用]")
    return "".join(labels)


def _is_cross_period(message: str) -> bool:
    return "跨期" in message or "上期" in message or "本期" in message and "上期" in message


def _strip_suffix(message: str) -> str:
    cleaned = message
    for suffix in ("，需核实。", "，需核实", "需核实。", "需核实"):
        if cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
            break
    return cleaned.rstrip("，。")


def _write_overview(sheet, rules: tuple[UserRule, ...]) -> None:
    public_count = sum(1 for rule in rules if "公开信息" in rule.result_message)
    template_count = sum(1 for rule in rules if "模板" in rule.result_message)
    rows = [
        ["数据库校验规则说明"],
        ["用途", "用于说明数据库校验引擎当前对接的旧程序业务校验规则，便于业务人员理解校验结果。"],
        ["适用范围", "本说明覆盖旧程序已迁移到数据库校验引擎的全部活动规则，输出结果仍以旧程序格式 Excel 为准。"],
        ["规则数量", f"共 {len(rules)} 条；其中公开信息交叉校验 {public_count} 条，模板交叉校验 {template_count} 条。"],
        ["读取方法", "先看“规则清单”了解每张表包含哪些规则，再看“规则明细”理解每条规则的触发含义。"],
        ["说明", "文档使用业务名称描述规则，不展示数据库字段名、技术实现细节或查询语句。"],
    ]
    for row in rows:
        sheet.append(row)
    sheet.merge_cells("A1:B1")
    sheet["A1"].font = Font(size=16, bold=True, color="1F2937")
    sheet["A1"].fill = PatternFill("solid", fgColor="EAF2FF")
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 96
    _style_sheet(sheet, header_row=1)


def _write_summary(sheet, rules: Iterable[UserRule]) -> None:
    sheet.append(["表单编号", "表单名称", "规则数量", "包含规则"])
    grouped: dict[tuple[str, str], list[UserRule]] = {}
    for rule in rules:
        grouped.setdefault((rule.form_code, rule.form_name), []).append(rule)
    for (form_code, form_name), form_rules in grouped.items():
        sheet.append([form_code, form_name, len(form_rules), "、".join(_rule_label(rule) for rule in form_rules)])
    _style_sheet(sheet)
    _set_widths(sheet, [12, 34, 12, 110])


def _write_detail(sheet, rules: Iterable[UserRule]) -> None:
    headers = ["序号", "表单编号", "表单名称", "规则编号", "检查内容", "什么情况下提示", "结果中的提示", "是否跨期", "规则类型", "执行状态", "备注"]
    sheet.append(headers)
    for index, rule in enumerate(rules, start=1):
        sheet.append([
            index,
            rule.form_code,
            rule.form_name,
            rule.rule_id,
            rule.check_item,
            rule.trigger_condition,
            rule.result_message,
            rule.cross_period,
            rule.category,
            rule.status,
            rule.note,
        ])
    _style_sheet(sheet)
    _set_widths(sheet, [8, 10, 28, 16, 42, 68, 48, 10, 16, 12, 56])


def _style_sheet(sheet, *, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="DCEBFF")
    header_font = Font(bold=True, color="111827")
    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = f"A{header_row + 1}"


def _set_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
