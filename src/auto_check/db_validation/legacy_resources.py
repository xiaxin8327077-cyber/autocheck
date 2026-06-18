from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from importlib import resources
from typing import Any

from openpyxl import load_workbook


RESOURCE_PACKAGE = "auto_check.resources.data"


@dataclass(frozen=True)
class OrgInfo:
    org_code: str
    org_name: str
    manager_org: str


@dataclass(frozen=True)
class Zg05Zg08Mapping:
    zg05_indicator: str
    zg08_counterparty_type: str


def _resource_workbook(name: str):
    resource = resources.files(RESOURCE_PACKAGE).joinpath(name)
    with resources.as_file(resource) as path:
        return load_workbook(path, read_only=True, data_only=True)


def _code_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    raw = str(value).strip()
    if raw.endswith(".0") and raw[:-2].isdigit():
        return raw[:-2]
    return raw


@lru_cache(maxsize=1)
def get_non_county_level_area_codes() -> frozenset[str]:
    workbook = _resource_workbook("FileName.xlsx")
    sheet = workbook.worksheets[1]
    codes: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = _code_text(row[1])
        flag = str(row[3] or "").strip()
        if code and flag == "Y":
            codes.add(code.zfill(6) if code.isdigit() and len(code) < 6 else code)
    workbook.close()
    return frozenset(codes)


@lru_cache(maxsize=16)
def get_indicator_names(zg_code: str) -> tuple[str, ...]:
    workbook = _resource_workbook("FileName.xlsx")
    sheet = workbook["指标信息"]
    headers = [_code_text(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        index = headers.index(zg_code.upper())
    except ValueError:
        workbook.close()
        return ()

    names: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = _code_text(row[index] if index < len(row) else "")
        if name:
            names.append(name)
    workbook.close()
    return tuple(names)


@lru_cache(maxsize=1)
def get_zg05_zg08_mappings() -> tuple[Zg05Zg08Mapping, ...]:
    workbook = _resource_workbook("FileName.xlsx")
    sheet = workbook["ZG05与ZG08指标对照表"]
    mappings: list[Zg05Zg08Mapping] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        indicator = _code_text(row[0])
        counterparty_type = _code_text(row[1])
        if indicator and counterparty_type and counterparty_type != "0":
            mappings.append(Zg05Zg08Mapping(indicator, counterparty_type))
    workbook.close()
    return tuple(mappings)


@lru_cache(maxsize=1)
def _org_info_by_code() -> dict[str, OrgInfo]:
    workbook = _resource_workbook("RefInfo.xlsx")
    sheet = workbook.worksheets[0]
    result: dict[str, OrgInfo] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        org_code = _code_text(row[1])
        if not org_code:
            continue
        result[org_code] = OrgInfo(
            org_code=org_code,
            org_name=str(row[2] or "").strip(),
            manager_org=str(row[4] or "").strip(),
        )
    workbook.close()
    return result


def get_org_info(org_code: str) -> OrgInfo:
    return _org_info_by_code().get(org_code, OrgInfo(org_code=org_code, org_name="", manager_org=""))


@lru_cache(maxsize=1)
def get_org_codes() -> frozenset[str]:
    return frozenset(_org_info_by_code())
