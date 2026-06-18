from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from auto_check.db_validation import VERSION
from auto_check.db_validation.models import ValidationResultRow
from auto_check.db_validation.tables import report_date_token


HEADERS = [
    "数据日期",
    "金融机构编码",
    "法人金融机构名称",
    "数据管理机构",
    "明细数据相关信息",
    "校验表单",
    "数据值1",
    "数据值2",
    "校验标识",
    "校验规则",
    "错误描述",
    "情况说明",
]


def result_filename(report_date: date) -> str:
    return f"{report_date_token(report_date)}-资管产品数据审核结果-模板校验（否）-公开信息校验（否）({VERSION}).xlsx"


def result_filename(
    report_date: date,
    *,
    enable_public_info_check: bool = False,
    enable_template_check: bool = False,
) -> str:
    template_flag = "\u662f" if enable_template_check else "\u5426"
    public_flag = "\u662f" if enable_public_info_check else "\u5426"
    return (
        f"{report_date_token(report_date)}-"
        f"\u8d44\u7ba1\u4ea7\u54c1\u6570\u636e\u5ba1\u6838\u7ed3\u679c-"
        f"\u6a21\u677f\u6821\u9a8c\uff08{template_flag}\uff09-"
        f"\u516c\u5f00\u4fe1\u606f\u6821\u9a8c\uff08{public_flag}\uff09"
        f"({VERSION}).xlsx"
    )


def write_result_excel(path: str | Path, rows: list[ValidationResultRow]) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for column in range(1, len(HEADERS) + 1):
        ws.column_dimensions[ws.cell(1, column).column_letter].width = 13
    for row in rows:
        ws.append(row.to_excel_row())
    wb.save(output)
    return output
